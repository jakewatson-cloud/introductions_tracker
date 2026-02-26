"""
Scan INVESTMENT COMPARABLES MASTER.xlsx for duplicate rows using the
fuzzy matching logic (price ±5% if both present, quarter ±1, fuzzy address).

Price rule: if both have a price → must be within ±5%. If one has a price
and the other doesn't → no match. If neither has a price → match on
address + quarter alone.

Reports duplicate pairs, merges any extra data from the duplicate into
the kept row, then deletes the duplicate row.

Usage:
    python find_inv_dupes.py          # scan only (dry run)
    python find_inv_dupes.py --fix    # merge + delete duplicates
"""

import argparse
import difflib
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# Reuse the helpers from excel_writer
sys.path.insert(0, str(Path(__file__).resolve().parent))
from email_pipeline.excel_writer import _normalize_name, _significant_words


# --- Column layout (B=2 through T=20) ---
DATA_START = 3
COL_RANGE = range(2, 21)  # B through T
COL_QUARTER = 3    # C
COL_ADDR = 6       # F
COL_PRICE = 12     # L
COL_TOWN = 4       # D
COL_SOURCE = 19    # S

COL_NAMES = {
    2: "Date", 3: "Quarter", 4: "Town", 5: "Style", 6: "Address",
    7: "Units", 8: "Area", 9: "Rent PA", 10: "Rent PSF", 11: "AWULTC",
    12: "Price", 13: "NIY", 14: "RY", 15: "Cap Val PSF",
    16: "Vendor", 17: "Purchaser", 18: "Comment", 19: "Source", 20: "Link",
}


# --- Matching helpers (same as InvestmentCompsWriter) ---

def parse_quarter(q: str) -> Optional[int]:
    s = (q or "").strip()
    # New format: "2025 Q1"
    m = re.match(r"(\d{4})\s*Q([1-4])", s, re.IGNORECASE)
    if m:
        return int(m.group(1)) * 4 + int(m.group(2))
    # Old format: "Q1 2025"
    m = re.match(r"Q([1-4])\s*(\d{4})", s, re.IGNORECASE)
    if m:
        return int(m.group(2)) * 4 + int(m.group(1))
    return None


def is_price_close(a: Optional[float], b: Optional[float], tol: float = 0.05) -> bool:
    if not a or not b:
        return False
    avg = (a + b) / 2
    return abs(a - b) / avg <= tol


def is_address_close(a: str, b: str) -> bool:
    na, nb = _normalize_name(a), _normalize_name(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    shorter = na if len(na) <= len(nb) else nb
    if len(shorter.split()) >= 2 and (na in nb or nb in na):
        return True
    wa, wb = _significant_words(a), _significant_words(b)
    if wa and wb:
        overlap = wa & wb
        shorter_len = min(len(wa), len(wb))
        if len(overlap) >= 2 and (len(overlap) / shorter_len) >= 0.6:
            return True
    if difflib.SequenceMatcher(None, na, nb).ratio() >= 0.85:
        return True
    return False


# --- Main ---

def main():
    parser = argparse.ArgumentParser(description="Find and optionally fix investment comp duplicates")
    parser.add_argument("--fix", action="store_true", help="Merge and delete duplicates (default: dry run)")
    args = parser.parse_args()

    from config import get_investment_comps_path

    path = get_investment_comps_path()
    if not path or not path.exists():
        print(f"File not found: {path}")
        return

    print(f"Scanning: {path.name}")
    print()

    # Work on a temp copy to avoid lock issues
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    shutil.copy2(str(path), tmp)

    wb = load_workbook(tmp, data_only=True, read_only=True)
    ws = wb["2026 Data"]

    # Read all rows into memory
    rows = []
    for r in range(DATA_START, ws.max_row + 1):
        addr = str(ws.cell(row=r, column=COL_ADDR).value or "").strip()
        if not addr:
            continue

        price_raw = ws.cell(row=r, column=COL_PRICE).value
        try:
            price = float(price_raw)
        except (TypeError, ValueError):
            price = None

        # Read all cell values for merge
        cells = {}
        for col in COL_RANGE:
            cells[col] = ws.cell(row=r, column=col).value

        rows.append({
            "row": r,
            "town": str(cells[COL_TOWN] or "").strip(),
            "address": addr,
            "price": price,
            "quarter": str(cells[COL_QUARTER] or "").strip(),
            "quarter_ord": parse_quarter(str(cells[COL_QUARTER] or "")),
            "source": str(cells[COL_SOURCE] or "").strip(),
            "cells": cells,
        })

    wb.close()
    try:
        os.unlink(tmp)
    except OSError:
        pass

    print(f"Loaded {len(rows)} data rows")
    print()

    # Find duplicate pairs
    dupes_found = []  # list of (keep_row_data, dupe_row_data)
    seen = set()

    for i, a in enumerate(rows):
        if a["row"] in seen:
            continue
        for j, b in enumerate(rows):
            if j <= i:
                continue
            if b["row"] in seen:
                continue

            # 1. Price check
            if a["price"] and b["price"]:
                # Both have prices — must be within ±5%
                if not is_price_close(a["price"], b["price"]):
                    continue
            elif a["price"] or b["price"]:
                # One has a price, the other doesn't — not a match
                continue
            # else: neither has a price — skip price check

            # 2. Quarter (if both present, within ±1)
            if a["quarter_ord"] is not None and b["quarter_ord"] is not None:
                if abs(a["quarter_ord"] - b["quarter_ord"]) > 1:
                    continue

            # 3. Address (fuzzy)
            if not is_address_close(a["address"], b["address"]):
                continue

            dupes_found.append((a, b))
            seen.add(b["row"])

    if not dupes_found:
        print("No duplicates found.")
        return

    # Report duplicates and what would be merged
    print(f"Found {len(dupes_found)} duplicate pair(s):")
    print("=" * 80)

    for keep, dupe in dupes_found:
        price_k = f"£{keep['price']:,.0f}" if keep["price"] else "N/A"
        price_d = f"£{dupe['price']:,.0f}" if dupe["price"] else "N/A"
        print()
        print(f"  KEEP   row {keep['row']:>4}: {keep['town']}, {keep['address']}")
        print(f"                   {keep['quarter']}  {price_k}  (source: {keep['source'][:40]})")
        print(f"  DUPE   row {dupe['row']:>4}: {dupe['town']}, {dupe['address']}")
        print(f"                   {dupe['quarter']}  {price_d}  (source: {dupe['source'][:40]})")

        # Show what would be merged
        merge_cols = []
        for col in COL_RANGE:
            keep_val = keep["cells"][col]
            dupe_val = dupe["cells"][col]
            keep_empty = keep_val is None or str(keep_val).strip() == ""
            dupe_has = dupe_val is not None and str(dupe_val).strip() != ""
            if keep_empty and dupe_has:
                merge_cols.append((col, dupe_val))

        if merge_cols:
            print(f"  MERGE  {len(merge_cols)} field(s) from dupe -> keep:")
            for col, val in merge_cols:
                display = f"{val}" if not isinstance(val, float) else f"{val:,.2f}"
                print(f"           {COL_NAMES.get(col, f'Col {col}')}: {display}")
        else:
            print(f"  MERGE  nothing to merge (kept row already has all data)")

    print()
    print(f"Total: {len(dupes_found)} duplicate pair(s)")
    print(f"Rows to remove: {sorted(seen)}")

    if not args.fix:
        print()
        print("Dry run — no changes made. Use --fix to merge and delete.")
        return

    # --- Apply: merge + delete ---
    print()
    print("Applying fixes...")

    # Re-open the actual file (not read-only this time)
    wb = load_workbook(str(path), data_only=False)
    ws = wb["2026 Data"]

    # 1. Merge: copy non-empty cells from dupe into keep where keep is empty
    for keep, dupe in dupes_found:
        for col in COL_RANGE:
            keep_cell = ws.cell(row=keep["row"], column=col)
            dupe_cell = ws.cell(row=dupe["row"], column=col)
            keep_empty = keep_cell.value is None or str(keep_cell.value).strip() == ""
            dupe_has = dupe_cell.value is not None and str(dupe_cell.value).strip() != ""
            if keep_empty and dupe_has:
                keep_cell.value = dupe_cell.value
                print(f"  Row {keep['row']}, {COL_NAMES.get(col, f'Col {col}')}: "
                      f"← {dupe_cell.value} (from row {dupe['row']})")

    # 2. Delete duplicate rows (in reverse order to keep row numbers stable)
    rows_to_delete = sorted(seen, reverse=True)
    for row_num in rows_to_delete:
        ws.delete_rows(row_num)
        print(f"  Deleted row {row_num}")

    # 3. Save via temp file (OneDrive safety)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        wb.close()
        shutil.copy2(tmp, str(path))
        print(f"\n  ✓ Saved: {path.name}")
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass

    print(f"  Done: {len(dupes_found)} duplicate(s) merged and removed.")


if __name__ == "__main__":
    main()
