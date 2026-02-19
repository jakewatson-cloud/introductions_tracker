"""
Scan OCCUPATIONAL COMPARABLES.xlsx for duplicate rows and invalid entries.

Six-phase approach:
  1. Normalised tenant name + rent PA within ±0.5% → duplicate
  2. Exact normalised unit + rent PA within ±0.5% → duplicate
  3. Fuzzy tenant name (≥90%) + rent PA within ±0.5% → duplicate
  4. Tenant = "Vacant" → remove
  5. Notes contains investment language (NIY, yield, cap val, etc.) → remove
  6. Both Rent PA and Rent PSF are blank → remove

Reports what would change, then merges extra data and deletes duplicates
when --fix is passed.

Usage:
    python find_occ_dupes.py          # scan only (dry run)
    python find_occ_dupes.py --fix    # merge + delete duplicates + vacants
"""

import argparse
import os
import re
import shutil
import sys
import tempfile
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# Reuse the normaliser from excel_writer
sys.path.insert(0, str(Path(__file__).resolve().parent))
from email_pipeline.excel_writer import _normalize_name


# --- Column layout (1-based, matching OccupationalCompsWriter) ---
COL_SOURCE = 1       # A: Source Deal
COL_ENTRY_TYPE = 2   # B: Entry Type
COL_TENANT = 3       # C: Tenant
COL_UNIT = 4         # D: Unit
COL_ADDRESS = 5      # E: Address
COL_TOWN = 6         # F: Town
COL_POSTCODE = 7     # G: Postcode
COL_SIZE = 8         # H: Size (sqft)
COL_RENT_PA = 9      # I: Rent PA
COL_RENT_PSF = 10    # J: Rent PSF
COL_LEASE_START = 11 # K
COL_LEASE_EXPIRY = 12 # L
COL_BREAK = 13       # M
COL_REVIEW = 14      # N
COL_TERM = 15        # O
COL_COMP_DATE = 16   # P
COL_NOTES = 17       # Q
COL_SOURCE_FILE = 18 # R
COL_EXTRACTION = 19  # S

COL_RANGE = range(1, 20)  # A through S

COL_NAMES = {
    1: "Source Deal", 2: "Entry Type", 3: "Tenant", 4: "Unit",
    5: "Address", 6: "Town", 7: "Postcode", 8: "Size",
    9: "Rent PA", 10: "Rent PSF", 11: "Lease Start", 12: "Lease Expiry",
    13: "Break", 14: "Review", 15: "Term", 16: "Comp Date",
    17: "Notes", 18: "Source File", 19: "Extraction Date",
}


# --- Helpers ---

def normalise_tenant(name: str) -> str:
    """Lowercase, strip punctuation, remove Ltd/Limited/PLC etc."""
    n = _normalize_name(name)
    n = re.sub(r'\b(ltd|limited|plc|inc|llp|llc)\b', '', n)
    return re.sub(r'\s+', ' ', n).strip()


def normalise_unit(name: str) -> str:
    """Lowercase, strip punctuation, remove 'unit'/'plot' prefix, strip leading zeros."""
    n = _normalize_name(name)
    n = re.sub(r'\b0+(\d)', r'\1', n)
    n = re.sub(r'\bunit\b\s*', '', n)
    n = re.sub(r'\bplot\b\s*', '', n)
    return n.strip()


def is_rent_close(a: Optional[float], b: Optional[float], tol: float = 0.005) -> bool:
    """Rent within ±0.5% (handles rounding like £178,875 vs £178,876)."""
    if not a or not b:
        return False
    avg = (a + b) / 2
    if avg == 0:
        return False
    return abs(a - b) / avg <= tol


def _rents_match(a: dict, b: dict, tol: float = 0.005) -> bool:
    """Check if two rows have matching rent — prefer rent PA, fall back to PSF.

    Many comparables have only rent PSF (no PA).  Using PA alone would
    miss obvious duplicates where both rows share the same PSF.
    """
    if is_rent_close(a["rent_pa"], b["rent_pa"], tol):
        return True
    # Fall back to rent PSF when BOTH rows lack rent PA
    if a["rent_pa"] is None and b["rent_pa"] is None:
        return is_rent_close(a["rent_psf"], b["rent_psf"], tol)
    return False


def is_vacant(tenant: str) -> bool:
    """Check if a tenant name contains 'Vacant' in any form.

    Catches "Vacant", "Vacant under offer", "Vacant - 12m rental guarantee",
    etc.  Empty tenant is NOT vacant — comparables naturally have no tenant.
    """
    if not tenant or not tenant.strip():
        return False
    return bool(re.search(r'(?i)\bvacant\b', tenant))


def is_investment_comp(notes: str) -> bool:
    """Check if notes indicate this is an investment comparable (wrong file).

    Matches investment-specific language that doesn't belong in occupational
    comps: "investment comparable", NIY (net initial yield), yield references,
    and capital value ("cap val").
    """
    return bool(
        re.search(r'(?i)investment\s+comp', notes)
        or re.search(r'(?i)\bNIY\b', notes)
        or re.search(r'(?i)\byield\b', notes)
        or re.search(r'(?i)\bcap\s*val', notes)
    )


# --- Core dedup function (callable from GUI or CLI) ---

def dedup_occupational_comps(path: Path, fix: bool = False) -> dict:
    """Scan and optionally fix duplicates in an occupational comps Excel file.

    Parameters
    ----------
    path : Path
        Path to OCCUPATIONAL COMPARABLES.xlsx.
    fix : bool
        If True, merge data from duplicates into keep rows, then delete
        duplicate and vacant rows.  If False, dry-run report only.

    Returns
    -------
    dict
        Summary with keys: 'rows_scanned', 'duplicate_pairs', 'vacant_rows',
        'rows_removed', 'details' (list of str).
    """
    print(f"Scanning: {path.name}")
    print()

    # Work on a temp copy to avoid lock issues
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    shutil.copy2(str(path), tmp)

    # read_only=False is fine here — we're reading a temp copy, not the
    # OneDrive original, so there's no lock concern.  read_only mode is
    # actually *slower* for random cell access (it's designed for streaming
    # with iter_rows) and can inflate max_row.
    wb = load_workbook(tmp, data_only=True)
    ws = wb.active

    # Read all rows into memory
    rows = []
    for r in range(2, ws.max_row + 1):
        source = ws.cell(row=r, column=COL_SOURCE).value
        if source is None or str(source).strip() == "":
            continue

        rent_pa_raw = ws.cell(row=r, column=COL_RENT_PA).value
        try:
            rent_pa = float(rent_pa_raw) if rent_pa_raw else None
        except (TypeError, ValueError):
            rent_pa = None

        rent_psf_raw = ws.cell(row=r, column=COL_RENT_PSF).value
        try:
            rent_psf = float(rent_psf_raw) if rent_psf_raw else None
        except (TypeError, ValueError):
            rent_psf = None

        # Read all cell values for merge
        cells = {}
        for col in COL_RANGE:
            cells[col] = ws.cell(row=r, column=col).value

        tenant_raw = str(cells[COL_TENANT] or "").strip()
        unit_raw = str(cells[COL_UNIT] or "").strip()
        notes_raw = str(cells[COL_NOTES] or "").strip()

        rows.append({
            "row": r,
            "source_deal": str(cells[COL_SOURCE] or "").strip(),
            "entry_type": str(cells[COL_ENTRY_TYPE] or "").strip(),
            "tenant_raw": tenant_raw,
            "tenant_norm": normalise_tenant(tenant_raw),
            "unit_raw": unit_raw,
            "unit_norm": normalise_unit(unit_raw),
            "address": str(cells[COL_ADDRESS] or "").strip(),
            "town": str(cells[COL_TOWN] or "").strip(),
            "notes_raw": notes_raw,
            "rent_pa": rent_pa,
            "rent_psf": rent_psf,
            "cells": cells,
        })

    wb.close()
    try:
        os.unlink(tmp)
    except OSError:
        pass

    print(f"Loaded {len(rows)} data rows")
    print()

    # ============================================================
    # Flag vacant rows, investment comps, and no-rent rows for removal
    # ============================================================
    vacant_rows = set()
    inv_comp_rows = set()
    no_rent_rows = set()
    for r in rows:
        if is_vacant(r["tenant_raw"]):
            vacant_rows.add(r["row"])
        # Also catch vacant units where tenant is blank but notes say "vacant"
        if (not r["tenant_raw"]
                and re.search(r'(?i)\bvacant\b', r["notes_raw"])):
            vacant_rows.add(r["row"])
        if is_investment_comp(r["notes_raw"]):
            inv_comp_rows.add(r["row"])
        if r["rent_pa"] is None and r["rent_psf"] is None:
            no_rent_rows.add(r["row"])

    # ============================================================
    # Phase 1, 2, 3: Find duplicate pairs
    # ============================================================
    dupes_found = []  # list of (keep_row_data, dupe_row_data, reason)
    seen = set()

    skip_rows = vacant_rows | inv_comp_rows | no_rent_rows
    for i, a in enumerate(rows):
        if a["row"] in seen or a["row"] in skip_rows:
            continue
        for j, b in enumerate(rows):
            if j <= i:
                continue
            if b["row"] in seen or b["row"] in skip_rows:
                continue

            # Phase 1: normalised tenant + rent ±0.5% (PA, or PSF fallback)
            if (a["tenant_norm"] and b["tenant_norm"]
                    and a["tenant_norm"] == b["tenant_norm"]
                    and _rents_match(a, b)):
                dupes_found.append((a, b, "tenant+rent"))
                seen.add(b["row"])
                continue

            # Phase 2: normalised unit + rent ±0.5% (PA, or PSF fallback)
            if (a["unit_norm"] and b["unit_norm"]
                    and a["unit_norm"] == b["unit_norm"]
                    and _rents_match(a, b)):
                dupes_found.append((a, b, "unit+rent"))
                seen.add(b["row"])
                continue

            # Phase 3: fuzzy tenant name (≥90%) + rent ±0.5% (PA, or PSF fallback)
            if (a["tenant_norm"] and b["tenant_norm"]
                    and _rents_match(a, b)):
                ratio = SequenceMatcher(
                    None, a["tenant_norm"], b["tenant_norm"]
                ).ratio()
                if ratio >= 0.90:
                    dupes_found.append((a, b, f"fuzzy tenant ({ratio:.0%})+rent"))
                    seen.add(b["row"])
                    continue

    # ============================================================
    # Report
    # ============================================================
    all_rows_to_remove = sorted(seen | vacant_rows | inv_comp_rows | no_rent_rows)
    details = []

    if not dupes_found and not vacant_rows and not inv_comp_rows and not no_rent_rows:
        print("No duplicates, vacant rows, investment comps, or no-rent rows found.")
        return {
            "rows_scanned": len(rows),
            "duplicate_pairs": 0,
            "vacant_rows": 0,
            "inv_comp_rows": 0,
            "no_rent_rows": 0,
            "rows_removed": 0,
            "details": [],
        }

    if dupes_found:
        print(f"Found {len(dupes_found)} duplicate pair(s):")
        print("=" * 100)

        for keep, dupe, reason in dupes_found:
            rent_k = f"\u00a3{keep['rent_pa']:,.0f}" if keep["rent_pa"] else "N/A"
            rent_d = f"\u00a3{dupe['rent_pa']:,.0f}" if dupe["rent_pa"] else "N/A"
            print()
            print(f"  KEEP   row {keep['row']:>4}: [{keep['entry_type']}] "
                  f"{keep['tenant_raw'][:30]}, {keep['unit_raw'][:15]}, {keep['address'][:40]}")
            print(f"                   Rent: {rent_k}  (source: {keep['source_deal'][:50]})")
            print(f"  DUPE   row {dupe['row']:>4}: [{dupe['entry_type']}] "
                  f"{dupe['tenant_raw'][:30]}, {dupe['unit_raw'][:15]}, {dupe['address'][:40]}")
            print(f"                   Rent: {rent_d}  (source: {dupe['source_deal'][:50]})")
            print(f"  MATCH  {reason}")

            details.append(
                f"[{reason}] '{keep['tenant_raw'][:30]}' row {keep['row']} "
                f"<-> '{dupe['tenant_raw'][:30]}' row {dupe['row']}"
            )

            # Show what would be merged
            merge_cols = []
            for col in COL_RANGE:
                keep_val = keep["cells"][col]
                dupe_val = dupe["cells"][col]
                keep_empty = keep_val is None or str(keep_val).strip() == ""
                dupe_has = dupe_val is not None and str(dupe_val).strip() != ""
                if keep_empty and dupe_has:
                    merge_cols.append((col, dupe_val))

            if merge_cols:
                print(f"  MERGE  {len(merge_cols)} field(s) from dupe -> keep:")
                for col, val in merge_cols:
                    display = f"{val}" if not isinstance(val, float) else f"{val:,.2f}"
                    print(f"           {COL_NAMES.get(col, f'Col {col}')}: {display}")
            else:
                print(f"  MERGE  nothing to merge")

    if vacant_rows:
        print()
        print(f"Found {len(vacant_rows)} vacant row(s) to remove:")
        for r in rows:
            if r["row"] in vacant_rows:
                print(f"  row {r['row']:>4}: [{r['entry_type']}] "
                      f"'{r['tenant_raw']}', {r['unit_raw']}, {r['address'][:40]} "
                      f"(source: {r['source_deal'][:40]})")
                details.append(f"[vacant] '{r['tenant_raw']}' row {r['row']}")

    if inv_comp_rows:
        print()
        print(f"Found {len(inv_comp_rows)} investment comparable row(s) to remove:")
        for r in rows:
            if r["row"] in inv_comp_rows:
                print(f"  row {r['row']:>4}: [{r['entry_type']}] "
                      f"{r['address'][:40]}  (source: {r['source_deal'][:40]})")
                details.append(f"[inv comp] '{r['address'][:40]}' row {r['row']}")

    if no_rent_rows:
        print()
        print(f"Found {len(no_rent_rows)} row(s) with no rent (PA or PSF) to remove:")
        for r in rows:
            if r["row"] in no_rent_rows:
                print(f"  row {r['row']:>4}: [{r['entry_type']}] "
                      f"{r['tenant_raw'][:30]}, {r['address'][:40]} "
                      f"(source: {r['source_deal'][:40]})")
                details.append(f"[no rent] '{r['address'][:40]}' row {r['row']}")

    print()
    print(f"Summary: {len(dupes_found)} duplicate pair(s), "
          f"{len(vacant_rows)} vacant row(s), "
          f"{len(inv_comp_rows)} investment comp row(s), "
          f"{len(no_rent_rows)} no-rent row(s)")
    print(f"Total rows to remove: {len(all_rows_to_remove)}")

    if not fix:
        print()
        print("Dry run - no changes made. Use --fix to merge and delete.")
        return {
            "rows_scanned": len(rows),
            "duplicate_pairs": len(dupes_found),
            "vacant_rows": len(vacant_rows),
            "inv_comp_rows": len(inv_comp_rows),
            "no_rent_rows": len(no_rent_rows),
            "rows_removed": 0,
            "details": details,
        }

    # --- Apply: backup, merge, delete ---
    print()
    print("Applying fixes...")

    # Backup before destructive changes
    from email_pipeline.excel_writer import _backup_file
    _backup_file(path)

    # Re-open the actual file (not read-only this time)
    wb = load_workbook(str(path), data_only=False)
    ws = wb.active

    # 1. Merge: copy non-empty cells from dupe into keep where keep is empty
    for keep, dupe, reason in dupes_found:
        for col in COL_RANGE:
            keep_cell = ws.cell(row=keep["row"], column=col)
            dupe_cell = ws.cell(row=dupe["row"], column=col)
            keep_empty = keep_cell.value is None or str(keep_cell.value).strip() == ""
            dupe_has = dupe_cell.value is not None and str(dupe_cell.value).strip() != ""
            if keep_empty and dupe_has:
                keep_cell.value = dupe_cell.value
                print(f"  Row {keep['row']}, {COL_NAMES.get(col, f'Col {col}')}: "
                      f"<- {dupe_cell.value} (from row {dupe['row']})")

    # 2. Delete rows (dupes + vacants) in reverse order to keep row numbers stable
    for row_num in sorted(all_rows_to_remove, reverse=True):
        ws.delete_rows(row_num)
        print(f"  Deleted row {row_num}")

    # 3. Save via temp file (OneDrive safety)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        wb.close()
        shutil.copy2(tmp, str(path))
        print(f"\n  Saved: {path.name}")
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass

    print(f"  Done: {len(dupes_found)} duplicate(s) merged/removed, "
          f"{len(vacant_rows)} vacant row(s) removed, "
          f"{len(inv_comp_rows)} investment comp row(s) removed, "
          f"{len(no_rent_rows)} no-rent row(s) removed.")

    return {
        "rows_scanned": len(rows),
        "duplicate_pairs": len(dupes_found),
        "vacant_rows": len(vacant_rows),
        "inv_comp_rows": len(inv_comp_rows),
        "no_rent_rows": len(no_rent_rows),
        "rows_removed": len(all_rows_to_remove),
        "details": details,
    }


# --- CLI entry point ---

def main():
    parser = argparse.ArgumentParser(
        description="Find and optionally fix occupational comp duplicates"
    )
    parser.add_argument("--fix", action="store_true",
                        help="Merge and delete duplicates + vacants (default: dry run)")
    args = parser.parse_args()

    from config import get_occupational_comps_path

    path = get_occupational_comps_path()
    if not path or not path.exists():
        print(f"File not found: {path}")
        return

    dedup_occupational_comps(path, fix=args.fix)


if __name__ == "__main__":
    main()
