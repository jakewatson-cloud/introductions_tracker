"""
Reparse Brochures — Standalone Comparables Re-Extraction
=========================================================
Walks the archive folders, re-runs brochure parsing on every PDF/XLSX,
and rewrites investment + occupational comparables to Excel.

Does NOT touch Gmail, Pipeline Excel, or deal extraction — only comps.

Usage:
    python reparse_brochures.py              # Full run: clear old comps & rewrite
    python reparse_brochures.py --dry-run    # Preview what would be parsed
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

# Ensure project root is on path
sys.path.insert(0, str(Path(__file__).resolve().parent))

from config import (
    get_anthropic_api_key,
    get_db_path,
    get_intros_archive_path,
    get_investment_comps_path,
    get_occupational_comps_path,
)
from email_pipeline.brochure_parser import parse_brochure
from email_pipeline.database import Database
from email_pipeline.excel_writer import InvestmentCompsWriter, OccupationalCompsWriter

logger = logging.getLogger(__name__)

BROCHURE_SUFFIXES = {".pdf", ".xlsx", ".xls"}
SKIP_FILES = {"metadata.json", "email_body.txt"}

# Filename patterns (case-insensitive substrings) that indicate a file is a
# financial model or appraisal rather than a brochure.  These waste API calls
# because they contain assumptions / derived numbers, not factual comps.
_SKIP_PATTERNS = [
    "model",
    "appraisal",
    "cashflow",
    "cash flow",
    "cash_flow",
    "underwriting",
    "proforma",
    "pro forma",
    "pro-forma",
    "forecast",
    "budget",
    "valuation",
    "sensitivity",
    "irr analysis",
    "dcf",
]


def _is_financial_model(filename: str) -> bool:
    """Return True if the filename looks like a financial model, not a brochure."""
    lower = filename.lower()
    return any(pat in lower for pat in _SKIP_PATTERNS)


def discover_brochures(archive_root: Path) -> list[tuple[str, Path]]:
    """Walk archive and return (source_deal_name, brochure_path) pairs."""
    results: list[tuple[str, Path]] = []

    for property_dir in sorted(archive_root.iterdir()):
        if not property_dir.is_dir():
            continue

        deal_name = property_dir.name  # e.g. "Birmingham, Kings Road"

        # Brochures live in dated sub-folders
        for sub_dir in sorted(property_dir.iterdir()):
            if not sub_dir.is_dir():
                continue

            for f in sorted(sub_dir.iterdir()):
                if (
                    f.is_file()
                    and f.suffix.lower() in BROCHURE_SUFFIXES
                    and f.name not in SKIP_FILES
                    and not _is_financial_model(f.name)
                ):
                    results.append((deal_name, f))

    return results


def clear_pipeline_comps(inv_path: Path, occ_path: Path) -> None:
    """Remove all pipeline-written investment & occupational comps.

    For investment comps: clears rows written by the pipeline (rows 87+
    that were auto-extracted). Leaves manually-entered rows intact.

    For occupational comps: the file is entirely auto-generated, so we
    recreate it fresh with just headers.
    """
    import openpyxl
    from copy import copy

    # --- Investment comps: clear rows that were auto-generated (pipeline rows).
    #     Pipeline rows are identified by: address present (col F) but no date (col B).
    #     Manually-entered rows always have a date.
    if inv_path.exists():
        try:
            wb = openpyxl.load_workbook(str(inv_path), data_only=False)
            ws = wb["2026 Data"]

            # Find where pipeline data starts: first row with address but no date.
            first_pipeline_row = None
            for row in range(3, ws.max_row + 1):
                addr = ws.cell(row=row, column=6).value
                date = ws.cell(row=row, column=2).value

                if addr and not date:
                    if first_pipeline_row is None:
                        first_pipeline_row = row

            if first_pipeline_row:
                cleared = 0
                for row in range(first_pipeline_row, ws.max_row + 1):
                    for col in range(2, 21):  # B through T
                        ws.cell(row=row, column=col).value = None
                    cleared += 1
                wb.save(str(inv_path))
                print(f"  Cleared {cleared} investment comp rows (from row {first_pipeline_row})")
            else:
                print("  No pipeline investment comp rows found to clear")

            wb.close()
        except Exception as e:
            print(f"  WARNING: Could not clear investment comps: {e}")

    # --- Occupational comps: clear rows in-place and rewrite headers ---
    # We do NOT delete the file because OneDrive can race: it sees the
    # delete, then overwrites the newly-created file with the deletion.
    # Instead, clear all data rows and overwrite headers so the column
    # layout always matches the current code.
    if occ_path.exists():
        try:
            wb = openpyxl.load_workbook(str(occ_path), data_only=False)
            ws = wb.active
            old_count = max(0, ws.max_row - 1)

            # Clear all data rows
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None

            # Overwrite headers to match current column layout
            from email_pipeline.excel_writer import OccupationalCompsWriter
            from openpyxl.styles import Font
            for i, header in enumerate(OccupationalCompsWriter.HEADERS, 1):
                cell = ws.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True)
            # Clear any extra old header columns beyond current layout
            for col in range(len(OccupationalCompsWriter.HEADERS) + 1, ws.max_column + 1):
                ws.cell(row=1, column=col).value = None

            wb.save(str(occ_path))
            wb.close()
            print(f"  Cleared {old_count} occupational comp rows (headers updated)")
        except Exception as e:
            print(f"  WARNING: Could not clear occupational comps: {e}")
    else:
        # File doesn't exist yet — the OccupationalCompsWriter will create it
        print("  Occupational comps file does not exist — will be created on first write")


def main():
    parser = argparse.ArgumentParser(description="Re-extract comparables from archived brochures")
    parser.add_argument("--dry-run", action="store_true", help="List brochures without parsing")
    args = parser.parse_args()

    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )

    # Load config
    api_key = get_anthropic_api_key()
    archive_root = get_intros_archive_path()
    inv_path = get_investment_comps_path()
    occ_path = get_occupational_comps_path()

    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY not set in .env")
        sys.exit(1)
    if not archive_root or not archive_root.exists():
        print(f"ERROR: Archive path not found: {archive_root}")
        sys.exit(1)

    # Discover brochures
    print(f"\nScanning archive: {archive_root}")
    brochures = discover_brochures(archive_root)
    print(f"Found {len(brochures)} brochure files\n")

    if not brochures:
        print("Nothing to parse.")
        return

    # Dry run: just list
    if args.dry_run:
        for deal_name, path in brochures:
            size_kb = path.stat().st_size // 1024
            print(f"  [{deal_name}] {path.name} ({size_kb} KB)")
        print(f"\n{len(brochures)} files would be parsed. Run without --dry-run to proceed.")
        return

    # Deduplicate files: same filename + same size = same brochure
    # (avoids sending the same PDF to Claude multiple times)
    seen_files: dict[tuple[str, int], list[str]] = {}  # (name, size) -> [deal_names]
    unique_brochures: list[tuple[str, Path]] = []
    dupes_skipped = 0
    for deal_name, path in brochures:
        key = (path.name, path.stat().st_size)
        if key in seen_files:
            seen_files[key].append(deal_name)
            dupes_skipped += 1
        else:
            seen_files[key] = [deal_name]
            unique_brochures.append((deal_name, path))

    if dupes_skipped:
        print(f"Deduplicated: {dupes_skipped} duplicate files skipped "
              f"({len(unique_brochures)} unique to parse)\n")

    # Initialise scrape-tracking database
    db = Database(str(get_db_path()))

    # Real run: clear old data and reparse
    print("=" * 60)
    print("Step 1: Clear old pipeline-written comparables")
    print("=" * 60)
    if inv_path:
        clear_pipeline_comps(inv_path, occ_path)
    else:
        print("  WARNING: No investment comps path configured")
    cleared = db.clear_scraped_brochures()
    if cleared:
        print(f"  Cleared {cleared} scraped brochure records")

    print()
    print("=" * 60)
    print("Step 2: Parse brochures and extract comparables")
    print("=" * 60)

    all_inv_comps = []
    all_occ_comps = []
    errors = []

    for i, (deal_name, path) in enumerate(unique_brochures, 1):
        # Include all deal names this file applies to
        key = (path.name, path.stat().st_size)
        deal_names = seen_files[key]
        source_label = deal_names[0] if len(deal_names) == 1 else f"{deal_names[0]} (+{len(deal_names)-1} more)"

        print(f"\n[{i}/{len(unique_brochures)}] {source_label} — {path.name}")

        try:
            result = parse_brochure(
                file_path=path,
                api_key=api_key,
                source_deal=deal_name,
                extract_deal=False,           # Skip deal extraction
                extract_investment_comps=True,
                extract_occupational_comps=True,
            )

            # Stamp source provenance on investment comps
            if result.investment_comps:
                for comp in result.investment_comps:
                    comp.source_deal = deal_name
                    comp.source_file_path = str(path)
                all_inv_comps.extend(result.investment_comps)
                print(f"    ✓ {len(result.investment_comps)} investment comps")
            if result.occupational_comps:
                for comp in result.occupational_comps:
                    comp.source_file_path = str(path)
                all_occ_comps.extend(result.occupational_comps)
                print(f"    ✓ {len(result.occupational_comps)} occupational comps")
            if result.error_message:
                print(f"    ⚠ {result.error_message}")
                errors.append(f"{deal_name}: {result.error_message}")
            if not result.investment_comps and not result.occupational_comps and not result.error_message:
                print(f"    – No comparables found")

            # Record in scrape database
            db.mark_brochure_scraped(
                file_path=str(path),
                file_name=path.name,
                file_size=path.stat().st_size,
                file_modified=datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
                deal_name=deal_name,
                investment_comps_found=len(result.investment_comps),
                occupational_comps_found=len(result.occupational_comps),
            )

        except Exception as e:
            print(f"    ✗ ERROR: {e}")
            errors.append(f"{deal_name}: {e}")

    print()
    print("=" * 60)
    print("Step 3: Write comparables to Excel")
    print("=" * 60)

    inv_written = 0
    occ_written = 0

    if all_inv_comps and inv_path:
        writer = InvestmentCompsWriter(inv_path)
        inv_written = writer.append_comps(all_inv_comps)
        print(f"  Investment comps written: {inv_written} (of {len(all_inv_comps)} extracted)")

    if all_occ_comps and occ_path:
        writer = OccupationalCompsWriter(occ_path)
        occ_written = writer.append_comps(all_occ_comps)
        print(f"  Occupational comps written: {occ_written} (of {len(all_occ_comps)} extracted)")

    # Post-write: backup, snapshot, clean (once per run)
    from email_pipeline.excel_writer import _backup_file

    if inv_written > 0 and inv_path and inv_path.exists():
        _backup_file(inv_path)

    if occ_written > 0 and occ_path and occ_path.exists():
        _backup_file(occ_path)
        try:
            from email_pipeline.occ_comps_cleaner import clean_occupational_comps
            from config import get_cleaned_occupational_comps_path, get_db_path
            cleaned_path = get_cleaned_occupational_comps_path()
            db_path = get_db_path()
            if cleaned_path:
                summary = clean_occupational_comps(
                    raw_excel_path=occ_path,
                    cleaned_excel_path=cleaned_path,
                    db_path=db_path,
                )
                filled = summary.get("cells_filled", 0)
                db_rows = summary.get("db_rows", 0)
                print(f"  Cleaner: {filled} cells filled, {db_rows} rows in DB")
        except Exception as e:
            print(f"  ⚠ Occ comps cleaner failed: {e}")

    # Summary
    print()
    print("=" * 60)
    print("Summary")
    print("=" * 60)
    print(f"  Brochures parsed:        {len(unique_brochures)} ({dupes_skipped} duplicates skipped)")
    print(f"  Investment comps:        {inv_written} written ({len(all_inv_comps)} extracted)")
    print(f"  Occupational comps:      {occ_written} written ({len(all_occ_comps)} extracted)")
    print(f"  Errors:                  {len(errors)}")
    if errors:
        print()
        for err in errors:
            print(f"    - {err}")


if __name__ == "__main__":
    main()
