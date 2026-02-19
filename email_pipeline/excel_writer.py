"""
Excel Writer
=============
Writes extracted data to the three target Excel files:

1. Pipeline 2026.xlsx — "Intros" sheet (deal introductions)
2. INVESTMENT COMPARABLES MASTER.xlsx — "2026 Data" sheet
3. OCCUPATIONAL COMPARABLES.xlsx — new file (created if needed)

Handles:
- Backup before every write
- Formula preservation (data_only=False)
- Deduplication against existing rows
- Formatting copied from previous rows
- OneDrive file lock retry with backoff
"""

import difflib
import logging
import os
import re
import shutil
import tempfile
import time
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from email_pipeline.models import (
    DealExtraction,
    InvestmentComp,
    OccupationalComp,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Fuzzy deal-matching helpers (used by PipelineWriter + cross-thread dedup)
# ---------------------------------------------------------------------------

# Words that are common in property names but not distinctive
_STOP_WORDS = {
    "the", "a", "an", "and", "of", "at", "in", "on", "for",
    "site", "park", "estate", "industrial", "trading", "business",
    "investment", "fund", "intro", "introduction", "portfolio",
    "centre", "center", "works", "unit", "units", "road", "street",
    "lane", "way", "drive", "close", "ltd", "limited", "plc",
    "son", "sons",
    # Street type abbreviations
    "rd", "st", "ave", "ct", "pl", "sq",
    # Common geographic/directional
    "north", "south", "east", "west",
}


def _normalize_name(name: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace."""
    name = name.lower().strip()
    name = re.sub(r"[^\w\s]", " ", name)   # punctuation → space
    name = re.sub(r"\s+", " ", name).strip()
    return name


def _significant_words(name: str) -> set[str]:
    """Extract significant (non-stop) words from a property name."""
    normalised = _normalize_name(name)
    return {w for w in normalised.split() if w not in _STOP_WORDS and len(w) > 1}


def _is_town_match(town_a: str, town_b: str) -> bool:
    """Check if two towns match, treating empty/multi-location as wildcards."""
    a = town_a.strip().lower()
    b = town_b.strip().lower()

    # Wildcards: empty, or multi-location variants
    if not a or not b:
        return True
    if a.startswith("multi") or b.startswith("multi"):
        return True

    return a == b


def is_deal_match(
    name_a: str,
    town_a: str,
    postcode_a: str,
    name_b: str,
    town_b: str,
    postcode_b: str,
) -> tuple[bool, str]:
    """Check if two deals represent the same property.

    Returns (is_match, reason) where reason describes the match type.
    Used by both PipelineWriter._is_duplicate() and _RunDeduplicator.

    Matching tiers (first match wins):
    1. Exact name + town
    2. Postcode match (non-empty)
    3. Substring containment (normalised)
    4. Significant word overlap (≥2 words, ≥60% of shorter set)
    5. Fuzzy name match within same town (SequenceMatcher ≥ 0.65)
    """
    norm_a = _normalize_name(name_a)
    norm_b = _normalize_name(name_b)

    # Skip if either name is empty
    if not norm_a or not norm_b:
        return False, ""

    # --- Tier 1: Exact name + town ---
    if norm_a == norm_b and _is_town_match(town_a, town_b):
        return True, f"exact match"

    # --- Tier 2: Postcode match ---
    pc_a = postcode_a.strip().upper()
    pc_b = postcode_b.strip().upper()
    if pc_a and pc_b and pc_a == pc_b:
        return True, f"postcode match ({pc_a})"

    # --- Tier 3: Substring containment ---
    # Require the shorter name to have ≥2 words to avoid false positives
    # with single-word town/area names (e.g. "Warrington" matching
    # "Warrington Central Trading Estate & Causeway Park")
    shorter_norm = norm_a if len(norm_a) <= len(norm_b) else norm_b
    if len(shorter_norm.split()) >= 2:
        if norm_a in norm_b:
            return True, f'"{name_a}" contained in "{name_b}"'
        if norm_b in norm_a:
            return True, f'"{name_b}" contained in "{name_a}"'

    # --- Tier 4: Significant word overlap ---
    words_a = _significant_words(name_a)
    words_b = _significant_words(name_b)

    if words_a and words_b:
        overlap = words_a & words_b
        shorter_len = min(len(words_a), len(words_b))
        if len(overlap) >= 2 and (len(overlap) / shorter_len) >= 0.6:
            return True, f"word overlap ({', '.join(sorted(overlap))})"

    # --- Tier 5: Fuzzy name match within same town ---
    # Compare only significant words (strip "Industrial Estate", "Business Park", etc.)
    # to avoid false positives from shared property type suffixes
    if _is_town_match(town_a, town_b):
        sig_a = " ".join(sorted(words_a)) if words_a else norm_a
        sig_b = " ".join(sorted(words_b)) if words_b else norm_b
        if sig_a and sig_b:
            ratio = difflib.SequenceMatcher(None, sig_a, sig_b).ratio()
            if ratio >= 0.65:
                return True, f"fuzzy match ({ratio:.0%})"

    return False, ""

# Maximum retries for OneDrive-locked files
MAX_RETRIES = 3
RETRY_DELAY = 5  # seconds


# ---------------------------------------------------------------------------
# 1. Pipeline Excel Writer
# ---------------------------------------------------------------------------

class PipelineWriter:
    """Writes deal data to Pipeline 2026.xlsx (Intros sheet).

    Schema (headers in row 11, data from row 12):
        B=Date, C=Agent, D=Asset Name, E=Country, F=Town, G=Address,
        H=Classification, I=Area(acres), J=Area(sqft), K=Rent PA,
        L=Rent PSF, M=Asking Price, N=NY%, O=RY%, P=CapVal PSF,
        Q=Status, R=Considered, S=Initial IRR, T=Brochure Stored,
        U=Brochure Scraped, V=Comment
    """

    SHEET_NAME = "Intros"
    HEADER_ROW = 11
    DATA_START_ROW = 12

    # Column mapping (1-indexed)
    COLUMNS = {
        "date": 2,             # B
        "agent": 3,            # C
        "asset_name": 4,       # D
        "country": 5,          # E
        "town": 6,             # F
        "address": 7,          # G
        "classification": 8,   # H
        "area_acres": 9,       # I
        "area_sqft": 10,       # J
        "rent_pa": 11,         # K
        "rent_psf": 12,        # L
        "asking_price": 13,    # M
        "net_yield": 14,       # N
        "reversionary_yield": 15,  # O
        "capval_psf": 16,      # P
        "status": 17,          # Q
        "considered": 18,      # R
        "initial_irr": 19,     # S
        "brochure_stored": 20, # T
        "brochure_scraped": 21,  # U
        "comment": 22,         # V
    }

    def __init__(self, excel_path: Path):
        """Initialize with path to Pipeline Excel file.

        Parameters
        ----------
        excel_path : Path
            Path to Pipeline 2026.xlsx.
        """
        self.excel_path = Path(excel_path)

    def append_deal(
        self,
        deal: DealExtraction,
        has_brochure: bool = False,
        brochure_scraped: bool = False,
        comment: str = "",
    ) -> bool:
        """Append a deal to the Pipeline Excel.

        Parameters
        ----------
        deal : DealExtraction
            The deal data to write.
        has_brochure : bool
            Whether a brochure was saved to the archive.
        brochure_scraped : bool
            Whether brochure data was extracted.
        comment : str
            Optional comment (e.g. "Auto-imported from email pipeline").

        Returns
        -------
        bool
            True if successfully written.
        """
        return _retry_write(
            self.excel_path,
            lambda wb: self._do_append(wb, deal, has_brochure, brochure_scraped, comment),
        )

    def _do_append(
        self,
        wb: "Workbook",
        deal: DealExtraction,
        has_brochure: bool,
        brochure_scraped: bool,
        comment: str,
    ) -> bool:
        """Internal: append deal to the workbook."""
        if self.SHEET_NAME not in wb.sheetnames:
            logger.error("Sheet '%s' not found in %s", self.SHEET_NAME, self.excel_path)
            return False

        ws = wb[self.SHEET_NAME]

        # Check for duplicate (multi-tier fuzzy matching)
        is_dup, dup_reason = self._is_duplicate(ws, deal)
        if is_dup:
            logger.info("  Duplicate found (%s): %s — skipping", dup_reason, deal.asset_name)
            return False

        # Find next empty row (scan column D = asset_name)
        next_row = self._find_next_row(ws)
        logger.info("  Writing to row %d", next_row)

        # Copy formatting from the previous data row
        format_row = next_row - 1 if next_row > self.DATA_START_ROW else self.DATA_START_ROW
        self._copy_row_format(ws, format_row, next_row)

        # Write data
        col = self.COLUMNS
        # Parse date string to a proper date so Excel treats it as a date, not text
        date_value = deal.date
        if isinstance(date_value, str) and date_value.strip():
            try:
                parts = date_value.strip().split("/")
                if len(parts) == 3:
                    date_value = datetime(int(parts[2]), int(parts[1]), int(parts[0])).date()
            except (ValueError, IndexError):
                pass  # keep as string if parsing fails
        ws.cell(row=next_row, column=col["date"], value=date_value)
        ws.cell(row=next_row, column=col["agent"], value=deal.agent)
        ws.cell(row=next_row, column=col["asset_name"], value=deal.asset_name)
        ws.cell(row=next_row, column=col["country"], value=deal.country)
        ws.cell(row=next_row, column=col["town"], value=deal.town)
        # Append postcode to address if not already present
        address_with_postcode = deal.address or ""
        if deal.postcode and deal.postcode not in address_with_postcode:
            address_with_postcode = f"{address_with_postcode}, {deal.postcode}".strip(", ")
        ws.cell(row=next_row, column=col["address"], value=address_with_postcode)
        ws.cell(row=next_row, column=col["classification"], value=deal.classification)

        if deal.area_acres is not None:
            ws.cell(row=next_row, column=col["area_acres"], value=deal.area_acres)
        if deal.area_sqft is not None:
            ws.cell(row=next_row, column=col["area_sqft"], value=deal.area_sqft)
        if deal.rent_pa is not None:
            ws.cell(row=next_row, column=col["rent_pa"], value=deal.rent_pa)
        if deal.rent_psf is not None:
            ws.cell(row=next_row, column=col["rent_psf"], value=deal.rent_psf)
        if deal.asking_price is not None:
            ws.cell(row=next_row, column=col["asking_price"], value=deal.asking_price)
        if deal.net_yield is not None:
            ws.cell(row=next_row, column=col["net_yield"], value=deal.net_yield / 100)
        if deal.reversionary_yield is not None:
            ws.cell(row=next_row, column=col["reversionary_yield"], value=deal.reversionary_yield / 100)
        if deal.capval_psf is not None:
            ws.cell(row=next_row, column=col["capval_psf"], value=deal.capval_psf)

        ws.cell(row=next_row, column=col["status"], value="New")
        ws.cell(row=next_row, column=col["considered"], value="No")
        ws.cell(row=next_row, column=col["brochure_stored"], value="Yes" if has_brochure else "No")
        ws.cell(row=next_row, column=col["brochure_scraped"], value="Yes" if brochure_scraped else "No")

        if comment:
            ws.cell(row=next_row, column=col["comment"], value=comment)

        return True

    def _is_duplicate(self, ws, deal: DealExtraction) -> tuple[bool, str]:
        """Check if a deal already exists in the sheet using fuzzy matching.

        Returns (is_duplicate, reason) where reason describes match type.
        """
        # If asset_name is empty/None, we can't meaningfully match — allow the write
        if not deal.asset_name or not deal.asset_name.strip():
            logger.warning("  Empty asset_name — skipping duplicate check")
            return False, ""

        col_d = self.COLUMNS["asset_name"]  # D
        col_f = self.COLUMNS["town"]        # F
        col_g = self.COLUMNS["address"]     # G (contains postcode)

        deal_postcode = (deal.postcode or "").strip().upper()

        for row in range(self.DATA_START_ROW, ws.max_row + 1):
            existing_name = str(ws.cell(row=row, column=col_d).value or "").strip()
            existing_town = str(ws.cell(row=row, column=col_f).value or "").strip()
            existing_addr = str(ws.cell(row=row, column=col_g).value or "").strip()

            if not existing_name:
                continue

            # Extract postcode from existing address (last word if it looks like a UK postcode)
            existing_postcode = ""
            if existing_addr:
                # UK postcodes end like "XX1 1XX" — grab last ~8 chars worth
                addr_parts = existing_addr.split(",")
                last_part = addr_parts[-1].strip()
                # Simple UK postcode pattern
                pc_match = re.search(r"[A-Z]{1,2}\d[\dA-Z]?\s*\d[A-Z]{2}", last_part.upper())
                if pc_match:
                    existing_postcode = pc_match.group(0)

            matched, reason = is_deal_match(
                name_a=deal.asset_name,
                town_a=deal.town or "",
                postcode_a=deal_postcode,
                name_b=existing_name,
                town_b=existing_town,
                postcode_b=existing_postcode,
            )

            if matched:
                return True, f'{reason} with "{existing_name}" in row {row}'

        return False, ""

    def _find_next_row(self, ws) -> int:
        """Find the next empty row (scanning column D)."""
        col_d = self.COLUMNS["asset_name"]
        for row in range(self.DATA_START_ROW, ws.max_row + 100):
            if not ws.cell(row=row, column=col_d).value:
                return row
        return ws.max_row + 1

    def _copy_row_format(self, ws, source_row: int, target_row: int):
        """Copy cell formatting from source row to target row."""
        for col in range(2, 23):  # B to V
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)


# ---------------------------------------------------------------------------
# 2. Investment Comparables Writer
# ---------------------------------------------------------------------------

class InvestmentCompsWriter:
    """Writes investment comparables to INVESTMENT COMPARABLES MASTER.xlsx.

    Schema ("2026 Data" sheet, headers row 2, data from row 3):
        B=Date, C=Quarter, D=Town, E=Style, F=Address, G=Units,
        H=Area, I=Rent(pa), J=Rent(psf), K=AWULTC, L=Price,
        M=Yield(NIY), N=RY, O=Cap Val psf, P=Vendor, Q=Purchaser,
        R=Comment, S=Information Source, T=Link
    """

    SHEET_NAME = "2026 Data"
    HEADER_ROW = 2
    DATA_START_ROW = 3

    # Column mapping (1-indexed) — matches actual spreadsheet layout
    COLUMNS = {
        "date": 2,                # B
        "quarter": 3,             # C
        "town": 4,                # D
        "style": 5,               # E
        "address": 6,             # F
        "units": 7,               # G
        "area_sqft": 8,           # H
        "rent_pa": 9,             # I
        "rent_psf": 10,           # J
        "awultc": 11,             # K
        "price": 12,              # L
        "yield_niy": 13,          # M
        "reversionary_yield": 14, # N
        "capval_psf": 15,         # O
        "vendor": 16,             # P
        "purchaser": 17,          # Q
        "comment": 18,            # R
        "source_deal": 19,        # S — Information Source
        "source_file_path": 20,   # T — Link
    }

    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)

    def append_comps(self, comps: list[InvestmentComp]) -> int:
        """Append investment comparables to the Excel file.

        Parameters
        ----------
        comps : list[InvestmentComp]
            Comparables to append.

        Returns
        -------
        int
            Number of comps successfully written.
        """
        if not comps:
            return 0

        count = [0]

        def _write(wb):
            if self.SHEET_NAME not in wb.sheetnames:
                logger.error("Sheet '%s' not found in %s", self.SHEET_NAME, self.excel_path)
                return False

            ws = wb[self.SHEET_NAME]

            dupes = 0
            merged = 0
            for comp in comps:
                dup_row = self._find_duplicate_row(ws, comp)
                if dup_row is not None:
                    price_str = f" (£{comp.price:,.0f})" if comp.price else ""
                    # Merge: fill any gaps in the existing row from this comp
                    fills = self._merge_into_row(ws, dup_row, comp)
                    if fills:
                        merged += 1
                        print(f"    ⊘ Duplicate inv comp: {comp.town}, {comp.address}{price_str}"
                              f" — merged {fills} field(s) into row {dup_row}")
                    else:
                        print(f"    ⊘ Duplicate inv comp: {comp.town}, {comp.address}{price_str}")
                    dupes += 1
                    continue

                next_row = self._find_next_row(ws)
                self._write_comp(ws, next_row, comp)
                count[0] += 1

            if dupes or count[0]:
                merge_note = f", {merged} merged" if merged else ""
                print(f"  → {count[0]} inv comps written, {dupes} duplicates skipped{merge_note}")
            return count[0] > 0 or merged > 0

        _retry_write(self.excel_path, _write)

        # Post-write cleaning pass: fill gaps using arithmetic relationships
        if count[0] > 0:
            try:
                from email_pipeline.comps_cleaner import clean_investment_comps

                summary = clean_investment_comps(self.excel_path)
                if summary["cells_filled"] > 0:
                    logger.info(
                        "  Comps cleaner filled %d cells", summary["cells_filled"]
                    )
            except Exception as e:
                logger.warning("  Comps cleaner failed: %s", e)

        return count[0]

    def _write_comp(self, ws, row: int, comp: InvestmentComp):
        """Write a single comp to a row."""
        col = self.COLUMNS

        # Copy formatting from the previous data row
        format_row = row - 1 if row > self.DATA_START_ROW else self.DATA_START_ROW
        self._copy_row_format(ws, format_row, row)

        if comp.date:
            ws.cell(row=row, column=col["date"], value=comp.date)
        if comp.quarter:
            ws.cell(row=row, column=col["quarter"], value=comp.quarter)

        ws.cell(row=row, column=col["town"], value=comp.town)

        if comp.style:
            ws.cell(row=row, column=col["style"], value=comp.style)

        ws.cell(row=row, column=col["address"], value=comp.address)

        if comp.units is not None:
            ws.cell(row=row, column=col["units"], value=comp.units)
        if comp.area_sqft is not None:
            ws.cell(row=row, column=col["area_sqft"], value=comp.area_sqft)
        if comp.rent_pa is not None:
            ws.cell(row=row, column=col["rent_pa"], value=comp.rent_pa)
        if comp.rent_psf is not None:
            ws.cell(row=row, column=col["rent_psf"], value=comp.rent_psf)
        if comp.awultc is not None:
            ws.cell(row=row, column=col["awultc"], value=comp.awultc)
        if comp.price is not None:
            ws.cell(row=row, column=col["price"], value=comp.price)
        if comp.yield_niy is not None:
            ws.cell(row=row, column=col["yield_niy"], value=comp.yield_niy / 100)
        if comp.reversionary_yield is not None:
            ws.cell(row=row, column=col["reversionary_yield"], value=comp.reversionary_yield / 100)
        if comp.capval_psf is not None:
            ws.cell(row=row, column=col["capval_psf"], value=comp.capval_psf)
        if comp.vendor:
            ws.cell(row=row, column=col["vendor"], value=comp.vendor)
        if comp.purchaser:
            ws.cell(row=row, column=col["purchaser"], value=comp.purchaser)
        if comp.comment:
            ws.cell(row=row, column=col["comment"], value=comp.comment)
        if comp.source_deal:
            ws.cell(row=row, column=col["source_deal"], value=comp.source_deal)
        if comp.source_file_path:
            ws.cell(row=row, column=col["source_file_path"], value=comp.source_file_path)

    def _merge_into_row(self, ws, row: int, comp: InvestmentComp) -> int:
        """Merge data from a duplicate comp into an existing row.

        For each column: if the existing cell is empty but the comp has a
        value, copy the comp's value in.  Never overwrites existing data.

        Returns the number of cells filled.
        """
        col = self.COLUMNS
        fills = 0

        # Map: column index → comp value (matching _write_comp logic)
        field_map = {
            col["date"]: comp.date,
            col["quarter"]: comp.quarter,
            col["town"]: comp.town,
            col["style"]: comp.style,
            col["address"]: comp.address,
            col["units"]: comp.units,
            col["area_sqft"]: comp.area_sqft,
            col["rent_pa"]: comp.rent_pa,
            col["rent_psf"]: comp.rent_psf,
            col["awultc"]: comp.awultc,
            col["price"]: comp.price,
            col["yield_niy"]: comp.yield_niy / 100 if comp.yield_niy is not None else None,
            col["reversionary_yield"]: comp.reversionary_yield / 100 if comp.reversionary_yield is not None else None,
            col["capval_psf"]: comp.capval_psf,
            col["vendor"]: comp.vendor,
            col["purchaser"]: comp.purchaser,
            col["comment"]: comp.comment,
            col["source_deal"]: comp.source_deal,
            col["source_file_path"]: comp.source_file_path,
        }

        for col_idx, new_val in field_map.items():
            if new_val is None:
                continue
            if isinstance(new_val, str) and not new_val.strip():
                continue
            existing = ws.cell(row=row, column=col_idx).value
            if existing is not None and str(existing).strip() != "":
                continue  # already has data — don't overwrite
            ws.cell(row=row, column=col_idx, value=new_val)
            fills += 1

        return fills

    @staticmethod
    def _parse_quarter(q: str) -> Optional[int]:
        """Parse '2025 Q1' or 'Q1 2025' into an ordinal (year*4 + quarter) for proximity checks.

        Returns None if the string can't be parsed.
        """
        s = (q or "").strip()
        # New format: "2025 Q1"
        m = re.match(r"(\d{4})\s*Q([1-4])", s, re.IGNORECASE)
        if m:
            return int(m.group(1)) * 4 + int(m.group(2))
        # Old format: "Q1 2025"
        m = re.match(r"Q([1-4])\s*(\d{4})", s, re.IGNORECASE)
        if m:
            return int(m.group(2)) * 4 + int(m.group(1))
        return None

    @staticmethod
    def _is_price_close(price_a: Optional[float], price_b: Optional[float],
                        tolerance: float = 0.05) -> bool:
        """Check if two prices are within ±tolerance of each other.

        Returns False if either price is None or zero.
        """
        if not price_a or not price_b:
            return False
        avg = (price_a + price_b) / 2
        return abs(price_a - price_b) / avg <= tolerance

    @staticmethod
    def _is_town_close(town_a: str, town_b: str) -> bool:
        """Fuzzy town match: exact, substring containment, or high similarity."""
        a = town_a.strip().lower()
        b = town_b.strip().lower()
        if not a or not b:
            return True  # treat missing town as wildcard
        if a == b:
            return True
        # Substring: "Letchworth" in "Letchworth Garden City"
        if a in b or b in a:
            return True
        return False

    @staticmethod
    def _is_address_close(addr_a: str, addr_b: str) -> bool:
        """Fuzzy address match: normalise then check similarity."""
        a = _normalize_name(addr_a)
        b = _normalize_name(addr_b)
        if not a or not b:
            return False
        if a == b:
            return True
        # Substring (≥2 words to avoid false positives)
        shorter = a if len(a) <= len(b) else b
        if len(shorter.split()) >= 2 and (a in b or b in a):
            return True
        # Significant word overlap
        words_a = _significant_words(addr_a)
        words_b = _significant_words(addr_b)
        if words_a and words_b:
            overlap = words_a & words_b
            shorter_len = min(len(words_a), len(words_b))
            if len(overlap) >= 2 and (len(overlap) / shorter_len) >= 0.6:
                return True
        # Fuzzy ratio — 0.85 threshold avoids false positives from shared
        # suffixes like "Industrial Estate" or "Business Park"
        ratio = difflib.SequenceMatcher(None, a, b).ratio()
        if ratio >= 0.85:
            return True
        return False

    def _find_duplicate_row(self, ws, comp: InvestmentComp) -> Optional[int]:
        """Find an existing row that matches this comp using fuzzy multi-field matching.

        A comp is a duplicate if ALL of the following match an existing row:
          1. Price within ±5%  (required on both sides)
          2. Quarter within ±1 quarter  (skipped if either is missing)
          3. Address: fuzzy (normalised substring, word overlap, or SequenceMatcher ≥ 0.85)

        Returns the matching row number, or None if no match found.
        """
        comp_price = comp.price
        if not comp_price:
            return None  # can't dedup without price

        comp_quarter_ord = self._parse_quarter(comp.quarter)

        col_addr = self.COLUMNS["address"]
        col_price = self.COLUMNS["price"]
        col_quarter = self.COLUMNS["quarter"]

        for row in range(self.DATA_START_ROW, ws.max_row + 1):
            existing_addr = str(ws.cell(row=row, column=col_addr).value or "").strip()
            if not existing_addr:
                continue

            # 1. Price check (required)
            existing_price = ws.cell(row=row, column=col_price).value
            try:
                existing_price = float(existing_price)
            except (TypeError, ValueError):
                continue  # no price → can't match
            if not self._is_price_close(comp_price, existing_price):
                continue

            # 2. Quarter check (if both present, must be within ±1)
            existing_quarter = str(ws.cell(row=row, column=col_quarter).value or "").strip()
            existing_quarter_ord = self._parse_quarter(existing_quarter)
            if comp_quarter_ord is not None and existing_quarter_ord is not None:
                if abs(comp_quarter_ord - existing_quarter_ord) > 1:
                    continue

            # 3. Address check (fuzzy)
            if not self._is_address_close(comp.address or "", existing_addr):
                continue

            # All checks passed → duplicate
            logger.info("  Duplicate inv comp: %s, %s matches row %d (%s)",
                        comp.town, comp.address, row, existing_addr)
            return row

        return None

    def _find_next_row(self, ws) -> int:
        """Find the next empty row (scanning column F = address)."""
        col_addr = self.COLUMNS["address"]
        for row in range(self.DATA_START_ROW, ws.max_row + 100):
            if not ws.cell(row=row, column=col_addr).value:
                return row
        return ws.max_row + 1

    def _copy_row_format(self, ws, source_row: int, target_row: int):
        """Copy cell formatting from source row to target row."""
        for col in range(2, 21):  # B to T
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)


# ---------------------------------------------------------------------------
# 3. Occupational Comparables Writer
# ---------------------------------------------------------------------------

class OccupationalCompsWriter:
    """Writes occupational comparables to OCCUPATIONAL COMPARABLES.xlsx.

    Creates the file if it doesn't exist.

    Columns:
        A=Source Deal, B=Entry Type, C=Tenant, D=Unit, E=Address, F=Town,
        G=Postcode, H=Size(sqft), I=Rent PA, J=Rent PSF, K=Lease Start,
        L=Lease Expiry, M=Break Date, N=Review Date, O=Term(yrs),
        P=Comp Date, Q=Notes, R=Source File, S=Extraction Date
    """

    HEADERS = [
        "Source Deal", "Entry Type", "Tenant", "Unit", "Address", "Town",
        "Postcode", "Size (sqft)", "Rent PA", "Rent PSF", "Lease Start",
        "Lease Expiry", "Break Date", "Review Date", "Term (yrs)",
        "Comp Date", "Notes", "Source File", "Extraction Date",
    ]

    # Column indices (1-based)
    COL_SOURCE = 1
    COL_ENTRY_TYPE = 2
    COL_TENANT = 3
    COL_UNIT = 4
    COL_ADDRESS = 5
    COL_TOWN = 6
    COL_POSTCODE = 7
    COL_SIZE = 8
    COL_RENT_PA = 9
    COL_RENT_PSF = 10
    COL_LEASE_START = 11
    COL_LEASE_EXPIRY = 12
    COL_BREAK = 13
    COL_REVIEW = 14
    COL_TERM = 15
    COL_COMP_DATE = 16
    COL_NOTES = 17
    COL_SOURCE_FILE = 18
    COL_EXTRACTION_DATE = 19

    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)

    def append_comps(self, comps: list[OccupationalComp]) -> int:
        """Append occupational comparables to the Excel file.

        Creates the file with headers if it doesn't exist.

        Parameters
        ----------
        comps : list[OccupationalComp]
            Comparables to append.

        Returns
        -------
        int
            Number of comps successfully written.
        """
        if not comps:
            return 0

        # Create file if it doesn't exist
        if not self.excel_path.exists():
            self._create_file()

        written_comps: list[OccupationalComp] = []

        def _write(wb):
            ws = wb.active
            print(f"\n  Writing {len(comps)} occ comps to {self.excel_path.name}...")

            dupes = 0
            merged = 0

            for i, comp in enumerate(comps):
                dup_row = self._find_duplicate_row(ws, comp)
                if dup_row is not None:
                    label = comp.tenant_name or comp.address
                    fills = self._merge_into_row(ws, dup_row, comp)
                    if fills:
                        merged += 1
                        print(f"    ⊘ Duplicate occ comp: {label} "
                              f"(source: {comp.source_deal}) "
                              f"— merged {fills} field(s) into row {dup_row}")
                    else:
                        print(f"    ⊘ Duplicate: {label} (source: {comp.source_deal})")
                    logger.info("  Duplicate occ comp: %s — skipping (row %d, merged %d)",
                                label, dup_row, fills)
                    dupes += 1
                    continue

                next_row = self._find_next_row(ws)
                self._write_comp(ws, next_row, comp)
                written_comps.append(comp)

            print(f"  → {len(written_comps)} written, "
                  f"{dupes} duplicates skipped"
                  f"{f' ({merged} with merge)' if merged else ''}")
            return len(written_comps) > 0 or merged > 0

        _retry_write(self.excel_path, _write)

        # Post-save verification: OneDrive can overwrite the file with a
        # stale cloud copy after we save.  If that happens, re-save from
        # scratch using _create_file + a fresh write pass.
        if written_comps:
            self._verify_and_resave(written_comps, len(written_comps))

        return len(written_comps)

    def _verify_and_resave(self, comps: list["OccupationalComp"], expected: int):
        """Re-read the file to make sure our rows actually persisted.

        OneDrive can silently overwrite a freshly-saved file with an older
        cloud version.  If we detect the rows are missing, we rebuild the
        file contents in-memory and save again via a temp file (up to 3
        attempts with a delay to let OneDrive settle).

        Parameters
        ----------
        comps : list[OccupationalComp]
            Only the comps that were actually written (duplicates excluded).
        expected : int
            Number of rows we expect to find.
        """
        import time as _time

        for attempt in range(3):
            _time.sleep(5)  # Give OneDrive time to settle
            try:
                wb = load_workbook(str(self.excel_path), data_only=False)
                ws = wb.active
                # Count non-empty data rows (source deal is always populated)
                data_rows = 0
                for r in range(2, ws.max_row + 1):
                    if ws.cell(row=r, column=self.COL_SOURCE).value:
                        data_rows += 1
                wb.close()

                if data_rows >= expected:
                    print(f"  ✓ Verified: {data_rows} occ comp rows persisted")
                    return  # All good

                print(
                    f"  ⚠ OneDrive overwrite detected: expected {expected} rows, "
                    f"found {data_rows} — re-saving (attempt {attempt + 1}/3)"
                )

                # Re-build the file: open, clear, rewrite headers + comps
                wb = load_workbook(str(self.excel_path), data_only=False)
                ws = wb.active

                # Clear any stale data
                for row in range(2, ws.max_row + 1):
                    for col in range(1, len(self.HEADERS) + 1):
                        ws.cell(row=row, column=col).value = None

                # Re-write only the non-duplicate comps
                row_num = 2
                for comp in comps:
                    self._write_comp(ws, row_num, comp)
                    row_num += 1

                # Save via temp file to avoid another OneDrive race
                fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
                os.close(fd)
                try:
                    wb.save(tmp_path)
                    wb.close()
                    shutil.copy2(tmp_path, str(self.excel_path))
                finally:
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass
                print(f"  Re-saved {len(comps)} rows to {self.excel_path.name}")

            except Exception as e:
                print(f"  ⚠ Verification attempt {attempt + 1} failed: {e}")

    def _create_file(self):
        """Create a new Excel file with headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Occupational Comps"

        # Write headers
        from openpyxl.styles import Font
        for i, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True)

        # Auto-width columns
        for i, header in enumerate(self.HEADERS, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 2, 12)

        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        # Save via temp file so OneDrive doesn't sync a headers-only
        # version that could overwrite the populated file moments later.
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            wb.save(tmp_path)
            wb.close()
            shutil.copy2(tmp_path, str(self.excel_path))
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
        logger.info("Created new file: %s", self.excel_path)

    def _write_comp(self, ws, row: int, comp: OccupationalComp):
        """Write a single occupational comp to a row."""
        ws.cell(row=row, column=self.COL_SOURCE, value=comp.source_deal)
        ws.cell(row=row, column=self.COL_ENTRY_TYPE, value=comp.entry_type)
        ws.cell(row=row, column=self.COL_TENANT, value=comp.tenant_name)
        ws.cell(row=row, column=self.COL_UNIT, value=comp.unit_name or "")
        ws.cell(row=row, column=self.COL_ADDRESS, value=comp.address)
        ws.cell(row=row, column=self.COL_TOWN, value=comp.town)
        ws.cell(row=row, column=self.COL_POSTCODE, value=comp.postcode or "")

        if comp.size_sqft is not None:
            ws.cell(row=row, column=self.COL_SIZE, value=comp.size_sqft)
        if comp.rent_pa is not None:
            ws.cell(row=row, column=self.COL_RENT_PA, value=comp.rent_pa)
        if comp.rent_psf is not None:
            ws.cell(row=row, column=self.COL_RENT_PSF, value=comp.rent_psf)

        ws.cell(row=row, column=self.COL_LEASE_START, value=comp.lease_start or "")
        ws.cell(row=row, column=self.COL_LEASE_EXPIRY, value=comp.lease_expiry or "")
        ws.cell(row=row, column=self.COL_BREAK, value=comp.break_date or "")
        ws.cell(row=row, column=self.COL_REVIEW, value=comp.rent_review_date or "")

        if comp.lease_term_years is not None:
            ws.cell(row=row, column=self.COL_TERM, value=comp.lease_term_years)

        ws.cell(row=row, column=self.COL_COMP_DATE, value=comp.comp_date or "")
        ws.cell(row=row, column=self.COL_NOTES, value=comp.notes or "")
        if comp.source_file_path:
            ws.cell(row=row, column=self.COL_SOURCE_FILE, value=comp.source_file_path)
        ws.cell(row=row, column=self.COL_EXTRACTION_DATE, value=datetime.now().strftime("%d/%m/%Y"))

    # --- Dedup helpers (occupational comps) ---

    @staticmethod
    def _normalise_tenant(name: str) -> str:
        """Normalise a tenant name for comparison.

        Lowercase, strip punctuation, collapse whitespace, and remove
        common suffixes (Ltd, Limited, PLC, Inc) that vary between sources.
        """
        n = _normalize_name(name)
        # Strip common company suffixes
        n = re.sub(r'\b(ltd|limited|plc|inc|llp|llc)\b', '', n)
        return re.sub(r'\s+', ' ', n).strip()

    @staticmethod
    def _normalise_unit(name: str) -> str:
        """Normalise a unit name for comparison.

        Lowercase, strip punctuation, strip 'unit'/'plot' prefixes,
        strip leading zeros.
        """
        n = _normalize_name(name)
        n = re.sub(r'\b0+(\d)', r'\1', n)        # "01" → "1"
        n = re.sub(r'\bunit\b\s*', '', n)         # strip "unit" prefix
        n = re.sub(r'\bplot\b\s*', '', n)         # strip "plot" prefix
        return n.strip()

    @staticmethod
    def _is_rent_close(rent_a: Optional[float], rent_b: Optional[float],
                       tolerance: float = 0.005) -> bool:
        """Check if two rents are within ±tolerance of each other.

        Default 0.5% tolerance handles rounding (e.g. £178,875 vs £178,876).
        Returns False if either rent is None or zero.
        """
        if not rent_a or not rent_b:
            return False
        avg = (rent_a + rent_b) / 2
        if avg == 0:
            return False
        return abs(rent_a - rent_b) / avg <= tolerance

    def _find_duplicate_row(self, ws, comp: OccupationalComp) -> Optional[int]:
        """Find an existing row that matches this comp.

        Three-phase dedup:

        Phase 1 — Normalised tenant name + rent PA within ±0.5%
                  (catches same tenant from different source brochures)

        Phase 2 — Exact normalised unit + rent PA within ±0.5%
                  (catches same unit from different source brochures
                   where tenant name may differ slightly)

        Phase 3 — Fuzzy tenant name (SequenceMatcher ≥ 0.90) + rent PA
                  within ±0.5%  (catches near-misspellings like
                  "Planetbloom" vs "Planet Bloom")

        Rows where tenant is "Vacant" are skipped entirely — they
        are not meaningful for dedup.

        Returns the matching row number, or None if no match found.
        """
        comp_tenant = self._normalise_tenant(comp.tenant_name or "")
        comp_unit = self._normalise_unit(comp.unit_name or "")

        # Skip explicitly "vacant" entries — not useful for dedup.
        # Empty tenant is fine (comparables have no tenant).
        if comp_tenant in ("vacant", "vacant under offer"):
            return None

        comp_rent = comp.rent_pa
        comp_rent_psf = comp.rent_psf

        for row in range(2, ws.max_row + 1):
            existing_source = ws.cell(row=row, column=self.COL_SOURCE).value
            if not existing_source:
                continue

            existing_tenant_raw = str(ws.cell(row=row, column=self.COL_TENANT).value or "").strip()
            existing_tenant = self._normalise_tenant(existing_tenant_raw)

            # Skip explicitly vacant rows in the sheet too
            if existing_tenant in ("vacant", "vacant under offer"):
                continue

            existing_rent = ws.cell(row=row, column=self.COL_RENT_PA).value
            try:
                existing_rent = float(existing_rent) if existing_rent else None
            except (TypeError, ValueError):
                existing_rent = None

            existing_rent_psf = ws.cell(row=row, column=self.COL_RENT_PSF).value
            try:
                existing_rent_psf = float(existing_rent_psf) if existing_rent_psf else None
            except (TypeError, ValueError):
                existing_rent_psf = None

            # Helper: rent PA match, falling back to PSF when both PA are None
            def _rents_match() -> bool:
                if self._is_rent_close(comp_rent, existing_rent):
                    return True
                if comp_rent is None and existing_rent is None:
                    return self._is_rent_close(comp_rent_psf, existing_rent_psf)
                return False

            # Phase 1: tenant name + rent
            if comp_tenant and existing_tenant == comp_tenant:
                if _rents_match():
                    return row

            # Phase 2: unit + rent (exact)
            if comp_unit:
                existing_unit = self._normalise_unit(
                    str(ws.cell(row=row, column=self.COL_UNIT).value or "").strip()
                )
                if existing_unit and comp_unit == existing_unit:
                    if _rents_match():
                        return row

            # Phase 3: fuzzy tenant name + rent
            if (comp_tenant and existing_tenant
                    and _rents_match()):
                ratio = difflib.SequenceMatcher(
                    None, comp_tenant, existing_tenant
                ).ratio()
                if ratio >= 0.90:
                    logger.info(
                        "  Fuzzy tenant match (%.0f%%): '%s' ~ '%s' (row %d)",
                        ratio * 100, comp_tenant, existing_tenant, row,
                    )
                    return row

        return None

    def _merge_into_row(self, ws, row: int, comp: OccupationalComp) -> int:
        """Merge data from a duplicate comp into an existing row.

        For each column: if the existing cell is empty but the comp has a
        value, copy the comp's value in.  Never overwrites existing data.

        Returns the number of cells filled.
        """
        fills = 0

        # Map: column index → comp value (matching _write_comp logic)
        field_map = {
            self.COL_TENANT: comp.tenant_name,
            self.COL_UNIT: comp.unit_name,
            self.COL_ADDRESS: comp.address,
            self.COL_TOWN: comp.town,
            self.COL_POSTCODE: comp.postcode,
            self.COL_SIZE: comp.size_sqft,
            self.COL_RENT_PA: comp.rent_pa,
            self.COL_RENT_PSF: comp.rent_psf,
            self.COL_LEASE_START: comp.lease_start,
            self.COL_LEASE_EXPIRY: comp.lease_expiry,
            self.COL_BREAK: comp.break_date,
            self.COL_REVIEW: comp.rent_review_date,
            self.COL_TERM: comp.lease_term_years,
            self.COL_COMP_DATE: comp.comp_date,
            self.COL_NOTES: comp.notes,
        }

        col_names = {
            self.COL_TENANT: "Tenant", self.COL_UNIT: "Unit",
            self.COL_ADDRESS: "Address", self.COL_TOWN: "Town",
            self.COL_POSTCODE: "Postcode", self.COL_SIZE: "Size",
            self.COL_RENT_PA: "Rent PA", self.COL_RENT_PSF: "Rent PSF",
            self.COL_LEASE_START: "Lease Start", self.COL_LEASE_EXPIRY: "Lease Expiry",
            self.COL_BREAK: "Break", self.COL_REVIEW: "Review",
            self.COL_TERM: "Term", self.COL_COMP_DATE: "Comp Date",
            self.COL_NOTES: "Notes",
        }

        for col_idx, new_val in field_map.items():
            if new_val is None:
                continue
            if isinstance(new_val, str) and not new_val.strip():
                continue
            existing = ws.cell(row=row, column=col_idx).value
            if existing is not None and str(existing).strip() != "":
                continue  # already has data — don't overwrite
            ws.cell(row=row, column=col_idx, value=new_val)
            fills += 1
            logger.debug("  Row %d, %s: ← %s", row,
                         col_names.get(col_idx, f"Col {col_idx}"), new_val)

        return fills

    def _find_next_row(self, ws) -> int:
        """Find the next empty row (scanning column A = source deal, always populated)."""
        for row in range(2, ws.max_row + 100):
            if not ws.cell(row=row, column=self.COL_SOURCE).value:
                return row
        return ws.max_row + 1


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _backup_file(file_path: Path) -> Optional[Path]:
    """Create a timestamped backup of an Excel file.

    Parameters
    ----------
    file_path : Path
        Path to the file to back up.

    Returns
    -------
    Path or None
        Path to the backup file, or None if failed.
    """
    if not file_path.exists():
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = file_path.parent / "backups"
    backup_dir.mkdir(exist_ok=True)

    backup_name = f"{file_path.stem}_{timestamp}{file_path.suffix}"
    backup_path = backup_dir / backup_name

    try:
        shutil.copy2(str(file_path), str(backup_path))
        logger.info("  Backup created: %s", backup_path.name)
    except Exception as e:
        logger.warning("  Failed to create backup: %s", e)
        return None

    # Prune old backups — keep only the 5 most recent per source file
    _prune_backups(backup_dir, file_path.stem, keep=5)

    return backup_path


def _prune_backups(backup_dir: Path, stem: str, keep: int = 5):
    """Delete old backups, keeping only the *keep* most recent per source file.

    Backups are matched by filename prefix (the original file's stem) and
    sorted by modification time.  Only ``.xlsx`` files are considered.
    """
    try:
        pattern = f"{stem}_*.xlsx"
        backups = sorted(
            backup_dir.glob(pattern),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        for old in backups[keep:]:
            old.unlink()
            logger.info("  Pruned old backup: %s", old.name)
    except Exception as e:
        logger.warning("  Failed to prune backups: %s", e)


def _retry_write(file_path: Path, write_fn, max_retries: int = MAX_RETRIES) -> bool:
    """Open an Excel file, call write_fn, and save with retry logic.

    Saves to a temp file outside OneDrive first, then copies back to
    avoid OneDrive silently overwriting the freshly-saved file with a
    stale cloud copy.

    Parameters
    ----------
    file_path : Path
        Path to the Excel file.
    write_fn : callable
        Function that takes a Workbook and returns bool (True if changes made).
    max_retries : int
        Maximum number of retries.

    Returns
    -------
    bool
        True if changes were actually written (write_fn returned True and save succeeded).
        False if skipped (duplicate), file locked, or error.
    """
    for attempt in range(max_retries):
        try:
            # Open workbook (preserve formulas)
            wb = load_workbook(str(file_path), data_only=False)

            # Call the write function
            changes_made = write_fn(wb)

            if changes_made:
                # Save to a temp file outside OneDrive, then copy back.
                # This prevents OneDrive from racing to overwrite the
                # target with a stale cloud copy during the save.
                fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
                os.close(fd)
                try:
                    wb.save(tmp_path)
                    wb.close()
                    shutil.copy2(tmp_path, str(file_path))
                    logger.info("  Saved: %s", file_path.name)
                finally:
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass
            else:
                logger.info("  No changes to save")
                wb.close()

            return bool(changes_made)

        except PermissionError as e:
            if attempt < max_retries - 1:
                wait = RETRY_DELAY * (2 ** attempt)
                logger.warning(
                    "  File locked (attempt %d/%d), retrying in %ds: %s",
                    attempt + 1, max_retries, wait, e,
                )
                time.sleep(wait)
            else:
                logger.error("  File locked after %d attempts: %s", max_retries, e)
                return False

        except Exception as e:
            print(f"  ✖ Excel write error: {e}")
            logger.error("  Excel write error: %s", e)
            return False

    return False
