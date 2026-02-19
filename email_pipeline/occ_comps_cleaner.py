"""
Occupational Comps Cleaner
==========================
Post-extraction cleaning pipeline for occupational comparables.

Reads the raw OCCUPATIONAL COMPARABLES.xlsx, applies cleaning rules in
cascade order, and outputs:
  1. A cleaned Excel file (OCCUPATIONAL COMPARABLES - CLEANED.xlsx)
  2. Rows in the SQLite cleaned_occupational_comps table

Rules (applied in cascade order):
    1. Date normalisation          — all date columns to ISO YYYY-MM-DD
    1b. Address from source deal   — fill blank address/town from source deal column
    2. Postcode from address       — regex-extract UK postcode from address string
    3. Town from postcode          — postcodes.io bulk lookup
    3b. Postcode from Sonnet       — Claude infers postcode from address/town
    4. Acres to sqft               — parse notes for site area in acres
    5. Rent arithmetic             — fill 3rd value from 2 known (size, PA, PSF)
    7. Comp date derivation        — fill comp_date from lease_start + term logic
    8. Build total_address         — concatenate address + town + postcode

Post-enrichment filters:
    - Remove rows still missing unit size after all derivation rules

Called automatically after OccupationalCompsWriter.append_comps() completes.
Also callable standalone from the GUI "Clean Occ Comps" button.
"""

import csv
import json
import logging
import os
import re
import shutil
import sqlite3
import tempfile
import time
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants — must match OccupationalCompsWriter in excel_writer.py
# ---------------------------------------------------------------------------

COL_SOURCE = 1
COL_ENTRY_TYPE = 2
COL_TENANT = 3
COL_UNIT = 4
COL_ADDRESS = 5
COL_TOWN = 6
COL_POSTCODE = 7
COL_SIZE = 8
COL_RENT_PA = 9
COL_RENT_PSF = 10
COL_LEASE_START = 11
COL_LEASE_EXPIRY = 12
COL_BREAK = 13
COL_REVIEW = 14
COL_TERM = 15
COL_COMP_DATE = 16
COL_NOTES = 17
COL_SOURCE_FILE = 18
COL_EXTRACTION_DATE = 19
COL_TOTAL_ADDRESS = 20  # New column in cleaned file

CLEANED_HEADERS = [
    "Source Deal", "Entry Type", "Tenant", "Unit", "Address", "Town",
    "Postcode", "Size (sqft)", "Rent PA", "Rent PSF", "Lease Start",
    "Lease Expiry", "Break Date", "Review Date", "Term (yrs)",
    "Comp Date", "Notes", "Source File", "Extraction Date",
    "Total Address",
]

DATE_COLUMNS = ["lease_start", "lease_expiry", "break_date",
                "rent_review_date", "comp_date"]

# UK postcode regex (same as email_archiver.py)
_UK_POSTCODE_RE = re.compile(
    r"[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}", re.IGNORECASE
)

# Acres pattern for storage yard / land site areas
_ACRES_RE = re.compile(
    r"(?:c\.?\s*|approx\.?\s*|circa\s*|~\s*)?(\d+(?:[.,]\d+)?)\s*(?:-?\s*)?(?:acres?|ac)\b",
    re.IGNORECASE,
)

SQFT_PER_ACRE = 43_560

# Month names for date parsing
_MONTH_MAP = {
    "jan": 1, "january": 1, "feb": 2, "february": 2,
    "mar": 3, "march": 3, "apr": 4, "april": 4,
    "may": 5, "jun": 6, "june": 6,
    "jul": 7, "july": 7, "aug": 8, "august": 8,
    "sep": 9, "september": 9, "oct": 10, "october": 10,
    "nov": 11, "november": 11, "dec": 12, "december": 12,
}

# Quarter start months
_QUARTER_MONTH = {"q1": 1, "q2": 4, "q3": 7, "q4": 10}

# Retry settings for OneDrive file locking
MAX_RETRIES = 3
RETRY_DELAY = 5

# postcodes.io
_POSTCODES_IO_URL = "https://api.postcodes.io/postcodes"
_POSTCODES_IO_PLACES_URL = "https://api.postcodes.io/places"
_POSTCODES_IO_BATCH_SIZE = 100


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def clean_occupational_comps(
    raw_excel_path: Path,
    cleaned_excel_path: Path,
    db_path: Path,
) -> dict:
    """Run the cleaning pipeline over raw occupational comparables.

    Parameters
    ----------
    raw_excel_path : Path
        Path to the raw OCCUPATIONAL COMPARABLES.xlsx.
    cleaned_excel_path : Path
        Path to write the cleaned Excel file.
    db_path : Path
        Path to the SQLite database.

    Returns
    -------
    dict
        Summary with keys: rows_scanned, cells_filled, details, db_rows.
    """
    raw_excel_path = Path(raw_excel_path)
    cleaned_excel_path = Path(cleaned_excel_path)
    db_path = Path(db_path)

    summary: dict = {
        "rows_scanned": 0,
        "cells_filled": 0,
        "details": [],
        "db_rows": 0,
    }

    if not raw_excel_path.exists():
        print(f"  Raw file not found: {raw_excel_path}")
        return summary

    # Step 1: Read raw rows
    print(f"  Reading raw data from {raw_excel_path.name}...")
    rows = _read_raw_rows(raw_excel_path)
    if not rows:
        print("  No data rows found")
        return summary

    summary["rows_scanned"] = len(rows)
    print(f"  {len(rows)} rows loaded")

    # Step 2: Batch postcode lookup for rows needing town
    postcodes_needing_town = set()
    for row in rows:
        pc = (row.get("postcode") or "").strip()
        town = (row.get("town") or "").strip()
        if pc and not town:
            postcodes_needing_town.add(_normalise_postcode(pc))
        # Also check if address (or source_deal fallback) contains a postcode
        if not pc:
            addr = (row.get("address") or "").strip()
            # If address is empty, source_deal will be used as address by Rule 1b
            if not addr:
                addr = (row.get("source_deal") or "").strip()
            match = _UK_POSTCODE_RE.search(addr)
            if match:
                extracted = _normalise_postcode(match.group())
                if extracted:
                    postcodes_needing_town.add(extracted)

    postcode_cache: dict[str, dict] = {}
    if postcodes_needing_town:
        print(f"  Looking up {len(postcodes_needing_town)} postcodes via postcodes.io...")
        postcode_cache = _batch_postcode_lookup(list(postcodes_needing_town))
        print(f"  {len(postcode_cache)} postcodes resolved")

    # Step 2b: Build places cache from source_deal comma segments.
    # For rows missing location data, source_deal often contains the town
    # (e.g. "Warrington, Gateway 49 Trade Park").  We extract unique candidate
    # place names and validate them against postcodes.io /places endpoint.
    candidate_places: set[str] = set()
    for row in rows:
        addr = (row.get("address") or "").strip()
        town = (row.get("town") or "").strip()
        pc = (row.get("postcode") or "").strip()
        source_deal = (row.get("source_deal") or "").strip()

        # Rows where source_deal could help fill gaps
        if source_deal and "," in source_deal and (not addr or not town):
            for part in source_deal.split(","):
                clean = part.strip()
                # Only consider parts that look like place names
                if (len(clean) >= 3
                        and clean[0].isalpha()
                        and not _UK_POSTCODE_RE.search(clean)
                        and not any(c.isdigit() for c in clean)):
                    candidate_places.add(clean.lower())

    places_cache: dict[str, dict] = {}
    if candidate_places:
        print(f"  Validating {len(candidate_places)} place name(s) via postcodes.io...")
        places_cache = _batch_places_lookup(list(candidate_places))
        print(f"  {len(places_cache)} confirmed as UK places")

    # Step 3: Clean each row
    print("  Applying cleaning rules...")
    changes = 0
    for i, row in enumerate(rows):
        row_num = i + 2  # Excel row (1-indexed, row 1 is headers)
        row_changes = _clean_row(row, row_num, postcode_cache, places_cache,
                                 summary["details"])
        changes += row_changes

    # Step 3b: Use Haiku to infer postcodes for rows still missing them.
    # Collect unique (address, town) combos that need a postcode.
    needs_postcode: dict[tuple[str, str], list[int]] = {}
    for i, row in enumerate(rows):
        pc = (row.get("postcode") or "").strip()
        addr = (row.get("address") or "").strip()
        town_val = (row.get("town") or "").strip()
        if not pc and (addr or town_val):
            key = (addr, town_val)
            if key not in needs_postcode:
                needs_postcode[key] = []
            needs_postcode[key].append(i)

    if needs_postcode:
        print(f"  {len(needs_postcode)} unique locations still missing postcodes, "
              f"asking Haiku...")
        haiku_postcodes = _haiku_postcode_lookup(list(needs_postcode.keys()))
        resolved = 0

        # Validate Haiku results via postcodes.io and fill rows
        haiku_pcs_to_validate = [
            pc for pc in haiku_postcodes.values()
            if pc and pc not in postcode_cache
        ]
        if haiku_pcs_to_validate:
            extra_cache = _batch_postcode_lookup(list(set(haiku_pcs_to_validate)))
            postcode_cache.update(extra_cache)

        for (addr, town_val), row_indices in needs_postcode.items():
            raw_pc = haiku_postcodes.get((addr, town_val))
            if not raw_pc:
                continue
            norm_pc = _normalise_postcode(raw_pc)
            # Only accept if postcodes.io confirms it's valid
            if norm_pc not in postcode_cache:
                logger.info("  Haiku suggested '%s' for '%s, %s' "
                            "but postcodes.io didn't recognise it — skipping",
                            norm_pc, addr, town_val)
                continue

            for idx in row_indices:
                row = rows[idx]
                row_num = idx + 2
                row["postcode"] = norm_pc
                changes += 1
                resolved += 1
                summary["details"].append(
                    f"Row {row_num}: filled postcode '{norm_pc}' via Haiku "
                    f"(from '{addr}, {town_val}')"
                )
                # Also fill town if still missing
                if not (row.get("town") or "").strip():
                    lookup = postcode_cache.get(norm_pc, {})
                    if lookup.get("town"):
                        row["town"] = lookup["town"]
                        changes += 1
                        summary["details"].append(
                            f"Row {row_num}: filled town '{lookup['town']}' "
                            f"from validated postcode {norm_pc}"
                        )
                # Rebuild total_address with new postcode
                row["total_address"] = _rule_build_total_address(
                    row.get("address") or "",
                    row.get("town") or "",
                    row.get("postcode") or "",
                )

        print(f"  {resolved} postcode(s) filled via Haiku "
              f"({len(needs_postcode) - len(haiku_postcodes)} unresolved)")

    # Step 3c: Write enriched location data back to the raw file so that
    # Haiku / places lookups don't repeat on future runs.  We only write
    # address, town, and postcode — columns that may have been blank
    # and are now filled by rules 1b, 1c, 2, 3, and 3b.
    _write_back_locations(rows, raw_excel_path)

    summary["cells_filled"] = changes
    print(f"  {changes} cells filled across {len(rows)} rows")

    # Step 3d: Remove rows still missing unit size after all enrichment.
    # Size can be derived from rent arithmetic (Rule 5) or acres (Rule 4),
    # so we only remove rows that are still blank after those rules have run.
    before_count = len(rows)
    rows = [r for r in rows if r.get("size_sqft") is not None]
    no_size_removed = before_count - len(rows)
    if no_size_removed > 0:
        print(f"  Removed {no_size_removed} row(s) with no unit size")
        summary["details"].append(
            f"Removed {no_size_removed} row(s) with no unit size (post-enrichment)"
        )
    summary["no_size_removed"] = no_size_removed

    # Step 4: Write cleaned Excel
    print(f"  Writing cleaned data to {cleaned_excel_path.name}...")
    _write_cleaned_excel(rows, cleaned_excel_path)

    # Step 5: Insert into database
    print(f"  Inserting into database...")
    db_rows = _insert_into_db(rows, db_path)
    summary["db_rows"] = db_rows
    print(f"  {db_rows} rows upserted into cleaned_occupational_comps")

    return summary


# ---------------------------------------------------------------------------
# CSV snapshot (audit trail for raw data)
# ---------------------------------------------------------------------------

def snapshot_raw_csv(raw_excel_path: Path):
    """Save a timestamped CSV snapshot of the raw occupational comps.

    Creates a snapshots/ subfolder next to the raw Excel file and writes
    a CSV named like ``occ_comps_raw_20260217_143022.csv``.  These are
    lightweight, append-only, and give a full audit trail of every
    extraction run.
    """
    raw_excel_path = Path(raw_excel_path)
    if not raw_excel_path.exists():
        return

    snapshot_dir = raw_excel_path.parent / "snapshots"
    snapshot_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = snapshot_dir / f"occ_comps_raw_{timestamp}.csv"

    rows = _read_raw_rows(raw_excel_path)
    if not rows:
        return

    # Use the raw Excel headers (excluding Total Address which is cleaned-only)
    fieldnames = [
        "source_deal", "entry_type", "tenant_name", "unit_name",
        "address", "town", "postcode", "size_sqft", "rent_pa", "rent_psf",
        "lease_start", "lease_expiry", "break_date", "rent_review_date",
        "lease_term_years", "comp_date", "notes", "source_file_path",
        "extraction_date",
    ]

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    logger.info("  Raw CSV snapshot: %s (%d rows)", csv_path.name, len(rows))
    print(f"  Raw CSV snapshot: {csv_path.name} ({len(rows)} rows)")


# ---------------------------------------------------------------------------
# Read raw rows from Excel
# ---------------------------------------------------------------------------

def _read_raw_rows(excel_path: Path) -> list[dict]:
    """Read all data rows from the raw occupational comps Excel."""
    rows = []

    for attempt in range(MAX_RETRIES):
        try:
            wb = load_workbook(str(excel_path), data_only=False)
            ws = wb.active

            for r in range(2, ws.max_row + 1):
                # Skip empty rows
                if not ws.cell(row=r, column=COL_SOURCE).value:
                    continue

                row = {
                    "source_deal": str(ws.cell(row=r, column=COL_SOURCE).value or ""),
                    "entry_type": str(ws.cell(row=r, column=COL_ENTRY_TYPE).value or "tenancy"),
                    "tenant_name": str(ws.cell(row=r, column=COL_TENANT).value or ""),
                    "unit_name": str(ws.cell(row=r, column=COL_UNIT).value or ""),
                    "address": str(ws.cell(row=r, column=COL_ADDRESS).value or ""),
                    "town": str(ws.cell(row=r, column=COL_TOWN).value or ""),
                    "postcode": str(ws.cell(row=r, column=COL_POSTCODE).value or ""),
                    "size_sqft": _to_number(ws.cell(row=r, column=COL_SIZE).value),
                    "rent_pa": _to_number(ws.cell(row=r, column=COL_RENT_PA).value),
                    "rent_psf": _to_number(ws.cell(row=r, column=COL_RENT_PSF).value),
                    "lease_start": ws.cell(row=r, column=COL_LEASE_START).value,
                    "lease_expiry": ws.cell(row=r, column=COL_LEASE_EXPIRY).value,
                    "break_date": ws.cell(row=r, column=COL_BREAK).value,
                    "rent_review_date": ws.cell(row=r, column=COL_REVIEW).value,
                    "lease_term_years": _to_number(ws.cell(row=r, column=COL_TERM).value),
                    "comp_date": ws.cell(row=r, column=COL_COMP_DATE).value,
                    "notes": str(ws.cell(row=r, column=COL_NOTES).value or ""),
                    "source_file_path": str(ws.cell(row=r, column=COL_SOURCE_FILE).value or ""),
                    "extraction_date": str(ws.cell(row=r, column=COL_EXTRACTION_DATE).value or ""),
                }
                rows.append(row)

            wb.close()
            return rows

        except PermissionError as e:
            if attempt < MAX_RETRIES - 1:
                wait = RETRY_DELAY * (2 ** attempt)
                print(f"  File locked, retrying in {wait}s: {e}")
                time.sleep(wait)
            else:
                print(f"  File locked after {MAX_RETRIES} attempts: {e}")
                return rows

        except Exception as e:
            print(f"  Error reading raw file: {e}")
            return rows

    return rows


# ---------------------------------------------------------------------------
# Row-level cleaning
# ---------------------------------------------------------------------------

def _clean_row(
    row: dict,
    row_num: int,
    postcode_cache: dict[str, dict],
    places_cache: dict[str, dict],
    details: list[str],
) -> int:
    """Apply all cleaning rules to a single row. Returns count of cells changed."""
    changes = 0

    # --- Rule 1: Date normalisation ---
    for col_name in DATE_COLUMNS:
        raw_val = row.get(col_name)
        if raw_val is not None and raw_val != "":
            normalised = _rule_normalise_date(raw_val)
            if normalised and normalised != str(raw_val):
                row[col_name] = normalised
                changes += 1
                details.append(
                    f"Row {row_num}: normalised {col_name} '{raw_val}' -> '{normalised}'"
                )
        elif raw_val == "" or raw_val is None:
            row[col_name] = None

    # --- Rule 1b: Fill address/town from source_deal when location fields are empty ---
    address = (row.get("address") or "").strip()
    town = (row.get("town") or "").strip()
    postcode = (row.get("postcode") or "").strip()

    if not address and not town and not postcode:
        source_deal = (row.get("source_deal") or "").strip()
        if source_deal:
            # Try to split town from address using the places cache.
            # Source deals are typically "Town, Property Name" or
            # "Property Name, Town" — test each comma-separated part
            # against the places cache to find the town.
            parts = [p.strip() for p in source_deal.split(",")]
            town_found = None
            town_idx = None

            if len(parts) >= 2 and places_cache:
                for idx, part in enumerate(parts):
                    clean = part.strip()
                    if clean.lower() in places_cache:
                        town_found = places_cache[clean.lower()]["name"]
                        town_idx = idx
                        break

            if town_found is not None:
                remaining = [p.strip() for i, p in enumerate(parts) if i != town_idx]
                row["address"] = ", ".join(remaining)
                row["town"] = town_found
                address = row["address"]
                town = town_found
                changes += 2
                details.append(
                    f"Row {row_num}: filled address '{row['address']}' "
                    f"and town '{town_found}' from source deal"
                )
            else:
                # No place match — use the whole source_deal as address
                row["address"] = source_deal
                address = source_deal
                changes += 1
                details.append(
                    f"Row {row_num}: filled address from source deal '{source_deal}'"
                )

    # --- Rule 1c: Fill town from source_deal when address exists but town is blank ---
    if not town and not postcode:
        source_deal = (row.get("source_deal") or "").strip()
        parts = [p.strip() for p in source_deal.split(",")]
        if len(parts) >= 2 and places_cache:
            for part in parts:
                clean = part.strip()
                if clean.lower() in places_cache:
                    row["town"] = places_cache[clean.lower()]["name"]
                    town = row["town"]
                    changes += 1
                    details.append(
                        f"Row {row_num}: filled town '{town}' from source deal"
                    )
                    break

    # --- Rule 2: Extract postcode from address ---
    postcode = (row.get("postcode") or "").strip()
    address = (row.get("address") or "").strip()

    if not postcode and address:
        match = _UK_POSTCODE_RE.search(address)
        if match:
            extracted = _normalise_postcode(match.group())
            if extracted:
                row["postcode"] = extracted
                # Remove postcode from address to avoid duplication
                row["address"] = address[:match.start()].rstrip(", ") + address[match.end():]
                row["address"] = row["address"].strip().rstrip(",").strip()
                postcode = extracted
                changes += 1
                details.append(
                    f"Row {row_num}: extracted postcode '{extracted}' from address"
                )

    # Normalise existing postcode
    if postcode:
        normalised_pc = _normalise_postcode(postcode)
        if normalised_pc and normalised_pc != postcode:
            row["postcode"] = normalised_pc
            postcode = normalised_pc

    # --- Rule 3: Town from postcode (via postcodes.io cache) ---
    town = (row.get("town") or "").strip()
    if not town and postcode and postcode in postcode_cache:
        lookup = postcode_cache[postcode]
        if lookup.get("town"):
            row["town"] = lookup["town"]
            changes += 1
            details.append(
                f"Row {row_num}: filled town '{lookup['town']}' from postcode {postcode}"
            )

    # --- Rule 4: Acres to sqft from notes ---
    notes = (row.get("notes") or "").strip()
    size = row.get("size_sqft")

    if size is None and notes:
        converted = _rule_acres_to_sqft(notes)
        if converted is not None:
            row["size_sqft"] = converted
            size = converted
            changes += 1
            details.append(
                f"Row {row_num}: converted {_ACRES_RE.search(notes).group()} "
                f"-> {converted:,.0f} sqft from notes"
            )

    # --- Rule 5: Rent arithmetic ---
    rent_pa = row.get("rent_pa")
    rent_psf = row.get("rent_psf")

    if rent_pa is None and size and rent_psf:
        rent_pa = round(rent_psf * size, 2)
        row["rent_pa"] = rent_pa
        changes += 1
        details.append(
            f"Row {row_num}: derived Rent PA {rent_pa:,.0f} from PSF & Size"
        )

    if rent_psf is None and rent_pa and size:
        rent_psf = round(rent_pa / size, 2)
        row["rent_psf"] = rent_psf
        changes += 1
        details.append(
            f"Row {row_num}: derived Rent PSF {rent_psf:.2f} from PA & Size"
        )

    if size is None and rent_pa and rent_psf:
        size = round(rent_pa / rent_psf, 0)
        row["size_sqft"] = size
        changes += 1
        details.append(
            f"Row {row_num}: derived Size {size:,.0f} sqft from PA & PSF"
        )

    # --- Rule 7: Derive comp_date (source-of-truth date for the transaction) ---
    # If comp_date is empty, derive it from lease_start and term:
    #   - If term > 5 years AND review date is empty, use min(lease_start + 5y, today)
    #     (captures rent reviewed at 5-year mark)
    #   - Otherwise, use lease_start as-is
    comp_date = row.get("comp_date")
    lease_start = row.get("lease_start")

    if not comp_date and lease_start:
        lease_start_str = str(lease_start).strip()
        try:
            ls_dt = datetime.strptime(lease_start_str[:10], "%Y-%m-%d")
        except (ValueError, TypeError):
            ls_dt = None

        if ls_dt:
            term_raw = row.get("lease_term_years")
            review_raw = row.get("rent_review_date")
            has_review = review_raw is not None and str(review_raw).strip() != ""

            try:
                term_years = float(term_raw) if term_raw else None
            except (TypeError, ValueError):
                term_years = None

            if term_years and term_years > 5 and not has_review:
                # Use lease start + 5 years, but only if that date is in the past
                review_dt = ls_dt.replace(year=ls_dt.year + 5)
                today = datetime.now()
                if review_dt <= today:
                    derived = review_dt.strftime("%Y-%m-%d")
                    row["comp_date"] = derived
                    changes += 1
                    details.append(
                        f"Row {row_num}: derived comp_date '{derived}' "
                        f"(lease start + 5yr review)"
                    )
                else:
                    # Review date would be in future — use lease start
                    derived = ls_dt.strftime("%Y-%m-%d")
                    row["comp_date"] = derived
                    changes += 1
                    details.append(
                        f"Row {row_num}: derived comp_date '{derived}' "
                        f"(lease start, 5yr review is future)"
                    )
            else:
                # Term <= 5 or has review date — use lease start
                derived = ls_dt.strftime("%Y-%m-%d")
                row["comp_date"] = derived
                changes += 1
                details.append(
                    f"Row {row_num}: derived comp_date '{derived}' from lease start"
                )

    # --- Rule 8: Build total_address ---
    row["total_address"] = _rule_build_total_address(
        row.get("address") or "",
        row.get("town") or "",
        row.get("postcode") or "",
    )

    return changes


# ---------------------------------------------------------------------------
# Individual rule functions
# ---------------------------------------------------------------------------

def _rule_normalise_date(value) -> Optional[str]:
    """Normalise a date value to ISO YYYY-MM-DD format.

    Handles:
    - datetime objects (openpyxl date cells)
    - DD/MM/YYYY or DD-MM-YYYY
    - MM/YYYY
    - Mon YYYY or Month YYYY (e.g. "Jan 2025", "January 2025")
    - Q2 2025, Q2/2025, Q2-2025
    - YYYY (year only)
    - YYYY-MM-DD (ISO passthrough)
    """
    if value is None:
        return None

    # Handle datetime objects from openpyxl
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    s = str(value).strip()
    if not s:
        return None

    # ISO passthrough: YYYY-MM-DD
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s

    # DD/MM/YYYY or DD-MM-YYYY
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$", s)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(year, month, day).strftime("%Y-%m-%d")
        except ValueError:
            return s  # Invalid date, return as-is

    # MM/YYYY
    m = re.match(r"^(\d{1,2})[/\-](\d{4})$", s)
    if m:
        month, year = int(m.group(1)), int(m.group(2))
        try:
            return datetime(year, month, 1).strftime("%Y-%m-%d")
        except ValueError:
            return s

    # Month YYYY (e.g. "Jan 2025", "January 2025")
    m = re.match(r"^([A-Za-z]+)\s+(\d{4})$", s)
    if m:
        month_str = m.group(1).lower()
        year = int(m.group(2))
        month_num = _MONTH_MAP.get(month_str)
        if month_num:
            return datetime(year, month_num, 1).strftime("%Y-%m-%d")

    # Quarter: Q2 2025, Q2/2025, Q2-2025
    m = re.match(r"^(Q[1-4])\s*[/\-]?\s*(\d{4})$", s, re.IGNORECASE)
    if m:
        quarter = m.group(1).lower()
        year = int(m.group(2))
        month = _QUARTER_MONTH.get(quarter, 1)
        return datetime(year, month, 1).strftime("%Y-%m-%d")

    # Year only: YYYY
    m = re.match(r"^(\d{4})$", s)
    if m:
        return f"{m.group(1)}-01-01"

    # Couldn't parse — return as-is
    return s


def _rule_acres_to_sqft(notes: str) -> Optional[float]:
    """Parse notes for site area in acres and convert to sqft.

    Returns sqft value or None if no acres reference found.
    """
    if not notes:
        return None

    match = _ACRES_RE.search(notes)
    if match:
        acres_str = match.group(1).replace(",", ".")
        try:
            acres = float(acres_str)
            if acres > 0:
                return round(acres * SQFT_PER_ACRE, 0)
        except ValueError:
            pass

    return None


def _rule_build_total_address(address: str, town: str, postcode: str) -> str:
    """Concatenate non-empty address components, deduplicating overlaps."""
    parts = []

    address = address.strip()
    town = town.strip()
    postcode = postcode.strip()

    if address:
        parts.append(address)

    # Only add town if it's not already in the address
    if town and town.lower() not in address.lower():
        parts.append(town)

    # Only add postcode if it's not already in the address
    if postcode and postcode.lower() not in address.lower():
        parts.append(postcode)

    return ", ".join(parts)


# ---------------------------------------------------------------------------
# Postcode helpers
# ---------------------------------------------------------------------------

def _normalise_postcode(raw: str) -> str:
    """Normalise a UK postcode: uppercase, single space before last 3 chars."""
    cleaned = re.sub(r"\s+", "", raw.strip().upper())
    if not _UK_POSTCODE_RE.match(cleaned):
        return raw.strip().upper()
    # Insert space before last 3 characters
    if len(cleaned) > 3:
        return cleaned[:-3] + " " + cleaned[-3:]
    return cleaned


def _batch_postcode_lookup(postcodes: list[str]) -> dict[str, dict]:
    """Look up postcodes via postcodes.io bulk endpoint.

    Returns dict mapping normalised postcode to {town, county, region}.
    """
    cache: dict[str, dict] = {}

    for batch_start in range(0, len(postcodes), _POSTCODES_IO_BATCH_SIZE):
        batch = postcodes[batch_start:batch_start + _POSTCODES_IO_BATCH_SIZE]

        payload = json.dumps({"postcodes": batch}).encode("utf-8")
        req = urllib.request.Request(
            _POSTCODES_IO_URL,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )

        for attempt in range(3):
            try:
                with urllib.request.urlopen(req, timeout=30) as resp:
                    data = json.loads(resp.read().decode("utf-8"))

                for item in data.get("result", []):
                    query_pc = item.get("query", "")
                    result = item.get("result")
                    if result:
                        cache[_normalise_postcode(query_pc)] = {
                            "town": result.get("admin_district", ""),
                            "county": result.get("admin_county", ""),
                            "region": result.get("region", ""),
                        }
                break  # Success

            except urllib.error.HTTPError as e:
                if e.code == 429 and attempt < 2:
                    wait = 2 ** (attempt + 1)
                    logger.warning("postcodes.io rate limited, waiting %ds", wait)
                    time.sleep(wait)
                else:
                    logger.warning("postcodes.io HTTP error: %s", e)
                    break

            except Exception as e:
                logger.warning("postcodes.io lookup failed: %s", e)
                break

    return cache


def _batch_places_lookup(candidates: list[str]) -> dict[str, dict]:
    """Validate place names via postcodes.io /places endpoint.

    Takes a list of lowercase candidate place names and returns a dict
    mapping confirmed place names to {name, county, region, outcode}.
    Only returns entries where the API confirms a match.
    """
    cache: dict[str, dict] = {}

    # Place types that indicate a settlement (not a building, park, etc.)
    _SETTLEMENT_TYPES = {
        "City", "Town", "Suburban Area", "Village", "Hamlet",
        "Other Settlement", "Section of Named Road",
    }

    for candidate in candidates:
        url = f"{_POSTCODES_IO_PLACES_URL}?q={urllib.request.quote(candidate)}&limit=1"
        req = urllib.request.Request(url, method="GET")

        for attempt in range(3):
            try:
                with urllib.request.urlopen(req, timeout=15) as resp:
                    data = json.loads(resp.read().decode("utf-8"))

                results = data.get("result") or []
                if results:
                    place = results[0]
                    name = place.get("name_1", "")
                    local_type = place.get("local_type", "")

                    # Only accept settlement types — skip "Industrial Estate" etc.
                    if local_type in _SETTLEMENT_TYPES:
                        cache[candidate] = {
                            "name": name,
                            "county": place.get("county_unitary", ""),
                            "region": place.get("region", ""),
                            "outcode": place.get("outcode", ""),
                        }
                break  # Success (even if no result)

            except urllib.error.HTTPError as e:
                if e.code == 429 and attempt < 2:
                    wait = 2 ** (attempt + 1)
                    logger.warning("postcodes.io places rate limited, waiting %ds", wait)
                    time.sleep(wait)
                else:
                    logger.warning("postcodes.io places HTTP error: %s", e)
                    break

            except Exception as e:
                logger.warning("postcodes.io places lookup failed: %s", e)
                break

    return cache


def _haiku_postcode_lookup(
    locations: list[tuple[str, str]],
) -> dict[tuple[str, str], str]:
    """Use Claude Haiku to infer UK postcodes from address/town pairs.

    Sends a single batched prompt with all locations and asks for the most
    likely UK postcode for each.  Results are validated against postcodes.io
    by the caller — this function just returns raw suggestions.

    Parameters
    ----------
    locations : list of (address, town) tuples

    Returns
    -------
    dict mapping (address, town) -> suggested postcode string
    """
    if not locations:
        return {}

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        logger.warning("ANTHROPIC_API_KEY not set — skipping Haiku postcode lookup")
        return {}

    # Build a numbered list for the prompt
    location_lines = []
    for i, (addr, town) in enumerate(locations, 1):
        parts = [p for p in (addr, town) if p]
        location_lines.append(f"{i}. {', '.join(parts)}")

    prompt = (
        "You are a UK commercial property expert. For each location below, "
        "provide the most likely UK postcode. These are industrial estates, "
        "business parks, and commercial properties.\n\n"
        "Respond ONLY with a JSON object mapping the line number to the "
        "postcode. If you are unsure, omit that entry. Use the format:\n"
        '{"1": "WA5 3UL", "2": "B97 6RH"}\n\n'
        "Locations:\n" + "\n".join(location_lines)
    )

    try:
        import anthropic

        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        response_text = message.content[0].text.strip()

        # Strip markdown code block if present
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            response_text = "\n".join(
                l for l in lines
                if not l.startswith("```")
            ).strip()

        result_map = json.loads(response_text)

        cache: dict[tuple[str, str], str] = {}
        for key, pc in result_map.items():
            idx = int(key) - 1
            if 0 <= idx < len(locations):
                normalised = _normalise_postcode(str(pc))
                if _UK_POSTCODE_RE.match(normalised.replace(" ", "")):
                    cache[locations[idx]] = normalised

        logger.info("Haiku suggested %d/%d postcodes", len(cache), len(locations))
        return cache

    except Exception as e:
        logger.warning("Haiku postcode lookup failed: %s", e)
        print(f"  ⚠ Haiku postcode lookup failed: {e}")
        return {}


# ---------------------------------------------------------------------------
# Numeric helper
# ---------------------------------------------------------------------------

def _to_number(value) -> Optional[float]:
    """Convert a cell value to a number, or None if not numeric / zero.

    Same pattern as comps_cleaner.py.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value != 0 else None
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("\u00a3", "").replace(" ", "").strip()
        if not cleaned:
            return None
        try:
            result = float(cleaned)
            return result if result != 0 else None
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Output: Cleaned Excel
# ---------------------------------------------------------------------------

def _write_back_locations(rows: list[dict], raw_excel_path: Path):
    """Write enriched address/town/postcode back to the raw Excel file.

    Only touches cells that were blank in the original and are now filled
    by the cleaning rules.  This means Haiku and places lookups won't
    repeat on future runs.  Uses a temp file + copy for OneDrive safety.
    """
    raw_excel_path = Path(raw_excel_path)
    if not raw_excel_path.exists():
        return

    try:
        wb = load_workbook(str(raw_excel_path), data_only=False)
        ws = wb.active

        updates = 0
        for i, row in enumerate(rows):
            excel_row = i + 2  # 1-indexed, row 1 is headers

            for col, key in (
                (COL_ADDRESS, "address"),
                (COL_TOWN, "town"),
                (COL_POSTCODE, "postcode"),
            ):
                current = ws.cell(row=excel_row, column=col).value
                new_val = (row.get(key) or "").strip()
                if (current is None or str(current).strip() == "") and new_val:
                    ws.cell(row=excel_row, column=col).value = new_val
                    updates += 1

        if updates > 0:
            fd, tmp = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            try:
                wb.save(tmp)
                wb.close()
                shutil.copy2(tmp, str(raw_excel_path))
                print(f"  Wrote {updates} location cell(s) back to {raw_excel_path.name}")
            finally:
                try:
                    os.unlink(tmp)
                except OSError:
                    pass
        else:
            wb.close()

    except Exception as e:
        logger.warning("Failed to write locations back to raw file: %s", e)
        print(f"  ⚠ Write-back to raw file failed: {e}")


def _write_cleaned_excel(rows: list[dict], excel_path: Path):
    """Write cleaned rows to a new Excel file.

    Keeps one rolling backup: if the cleaned file already exists, it is
    renamed to ``<name> - BACKUP.xlsx`` before the new file is written.
    Only one backup is kept (overwritten each time).
    """
    # Rolling backup: keep one previous copy
    if excel_path.exists():
        backup_path = excel_path.with_name(
            excel_path.stem + " - BACKUP" + excel_path.suffix
        )
        try:
            shutil.copy2(str(excel_path), str(backup_path))
            logger.info("  Cleaned backup: %s", backup_path.name)
        except Exception as e:
            logger.warning("  Failed to create cleaned backup: %s", e)

    wb = Workbook()
    ws = wb.active
    ws.title = "Occupational Comps (Cleaned)"

    # Headers
    for i, header in enumerate(CLEANED_HEADERS, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=COL_SOURCE, value=row.get("source_deal", ""))
        ws.cell(row=r_idx, column=COL_ENTRY_TYPE, value=row.get("entry_type", ""))
        ws.cell(row=r_idx, column=COL_TENANT, value=row.get("tenant_name", ""))
        ws.cell(row=r_idx, column=COL_UNIT, value=row.get("unit_name", ""))
        ws.cell(row=r_idx, column=COL_ADDRESS, value=row.get("address", ""))
        ws.cell(row=r_idx, column=COL_TOWN, value=row.get("town", ""))
        ws.cell(row=r_idx, column=COL_POSTCODE, value=row.get("postcode", ""))

        size = row.get("size_sqft")
        if size is not None:
            ws.cell(row=r_idx, column=COL_SIZE, value=size)
        rent_pa = row.get("rent_pa")
        if rent_pa is not None:
            ws.cell(row=r_idx, column=COL_RENT_PA, value=rent_pa)
        rent_psf = row.get("rent_psf")
        if rent_psf is not None:
            ws.cell(row=r_idx, column=COL_RENT_PSF, value=rent_psf)

        ws.cell(row=r_idx, column=COL_LEASE_START, value=row.get("lease_start") or "")
        ws.cell(row=r_idx, column=COL_LEASE_EXPIRY, value=row.get("lease_expiry") or "")
        ws.cell(row=r_idx, column=COL_BREAK, value=row.get("break_date") or "")
        ws.cell(row=r_idx, column=COL_REVIEW, value=row.get("rent_review_date") or "")

        term = row.get("lease_term_years")
        if term is not None:
            ws.cell(row=r_idx, column=COL_TERM, value=term)

        ws.cell(row=r_idx, column=COL_COMP_DATE, value=row.get("comp_date") or "")
        ws.cell(row=r_idx, column=COL_NOTES, value=row.get("notes", ""))
        ws.cell(row=r_idx, column=COL_SOURCE_FILE, value=row.get("source_file_path", ""))
        ws.cell(row=r_idx, column=COL_EXTRACTION_DATE, value=row.get("extraction_date", ""))
        ws.cell(row=r_idx, column=COL_TOTAL_ADDRESS, value=row.get("total_address", ""))

    # Auto-width columns
    for i, header in enumerate(CLEANED_HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 2, 12)

    # Save via temp file (OneDrive protection)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp_path)
        wb.close()
        shutil.copy2(tmp_path, str(excel_path))
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


# ---------------------------------------------------------------------------
# Output: SQLite database
# ---------------------------------------------------------------------------

def _insert_into_db(rows: list[dict], db_path: Path) -> int:
    """Replace the cleaned_occupational_comps table with the current rows.

    Truncates the table first so the DB exactly mirrors the cleaned Excel
    output — no stale rows from previous runs are kept.
    """
    cleaned_at = datetime.now().isoformat()
    count = 0

    with sqlite3.connect(str(db_path)) as conn:
        conn.execute("DELETE FROM cleaned_occupational_comps")
        for row in rows:
            try:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO cleaned_occupational_comps (
                        source_deal, source_file_path, entry_type,
                        tenant_name, unit_name,
                        address, town, postcode, total_address,
                        size_sqft, rent_pa, rent_psf,
                        lease_start, lease_expiry, break_date,
                        rent_review_date, lease_term_years, comp_date,
                        notes, extraction_date, cleaned_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row.get("source_deal", ""),
                        row.get("source_file_path", ""),
                        row.get("entry_type", "tenancy"),
                        row.get("tenant_name", ""),
                        row.get("unit_name", ""),
                        row.get("address", ""),
                        row.get("town", ""),
                        row.get("postcode", ""),
                        row.get("total_address", ""),
                        row.get("size_sqft"),
                        row.get("rent_pa"),
                        row.get("rent_psf"),
                        row.get("lease_start"),
                        row.get("lease_expiry"),
                        row.get("break_date"),
                        row.get("rent_review_date"),
                        row.get("lease_term_years"),
                        row.get("comp_date"),
                        row.get("notes", ""),
                        row.get("extraction_date", ""),
                        cleaned_at,
                    ),
                )
                count += 1
            except Exception as e:
                logger.warning("Failed to insert row: %s", e)

        conn.commit()

    return count
