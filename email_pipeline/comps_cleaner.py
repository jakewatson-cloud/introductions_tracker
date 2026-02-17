"""
Comps Cleaner
=============
Post-write cleaning pass over the INVESTMENT COMPARABLES MASTER spreadsheet.
Fills gaps using arithmetic relationships between columns.

Called automatically after InvestmentCompsWriter.append_comps() completes.
Also callable standalone from the GUI and reparse script.

Rules (applied in cascade order):
    1. Quarter from Date         (C from B)
    2. Rent PA from Price & Yield (I from L, M)  — (price * 1.068) * yield
    3. Rent PA from PSF & Area   (I from J, H)  — psf * area
    4. Rent PSF from PA & Area   (J from I, H)  — rent_pa / area
    5. Area from Rent PA & PSF   (H from I, J)  — rent_pa / rent_psf
    6. Cap Val PSF from Price & Area (O from L, H) — price / area
"""

import logging
import re
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants — must match InvestmentCompsWriter in excel_writer.py
# ---------------------------------------------------------------------------

SHEET_NAME = "2026 Data"
HEADER_ROW = 2
DATA_START_ROW = 3

# Column indices (1-based)
COL_DATE = 2          # B
COL_QUARTER = 3       # C
COL_AREA = 8          # H
COL_RENT_PA = 9       # I
COL_RENT_PSF = 10     # J
COL_PRICE = 12        # L
COL_YIELD_NIY = 13    # M  (stored as decimal, e.g. 0.065)
COL_CAPVAL_PSF = 15   # O

# Acquisition cost multiplier for deriving rent from price * yield.
# ~6.8% covers stamp duty + fees on a UK commercial property purchase.
ACQUISITION_COST_FACTOR = 1.068

# Retry settings for OneDrive file locking
MAX_RETRIES = 3
RETRY_DELAY = 5  # seconds (doubles each retry)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def clean_investment_comps(excel_path: Path) -> dict:
    """Run cleaning rules over the investment comps spreadsheet.

    Opens the workbook, iterates all data rows, applies gap-filling
    rules in cascade order, and saves.  Uses retry logic for OneDrive
    file locking.

    Parameters
    ----------
    excel_path : Path
        Path to INVESTMENT COMPARABLES MASTER.xlsx.

    Returns
    -------
    dict
        Summary with keys:
        - rows_scanned (int)
        - cells_filled (int)
        - details (list[str]) — human-readable description of each change
    """
    excel_path = Path(excel_path)
    summary: dict = {"rows_scanned": 0, "cells_filled": 0, "details": []}

    for attempt in range(MAX_RETRIES):
        try:
            wb = load_workbook(str(excel_path), data_only=False)

            if SHEET_NAME not in wb.sheetnames:
                logger.error("Sheet '%s' not found in %s", SHEET_NAME, excel_path)
                wb.close()
                return summary

            ws = wb[SHEET_NAME]
            changes = 0

            for row in range(DATA_START_ROW, ws.max_row + 1):
                # Skip empty rows (no address in col F)
                if not ws.cell(row=row, column=6).value:
                    continue

                summary["rows_scanned"] += 1

                # Read all relevant cell values
                date_val = ws.cell(row=row, column=COL_DATE).value
                quarter_val = ws.cell(row=row, column=COL_QUARTER).value
                area = _to_number(ws.cell(row=row, column=COL_AREA).value)
                rent_pa = _to_number(ws.cell(row=row, column=COL_RENT_PA).value)
                rent_psf = _to_number(ws.cell(row=row, column=COL_RENT_PSF).value)
                price = _to_number(ws.cell(row=row, column=COL_PRICE).value)
                yield_niy = _to_number(ws.cell(row=row, column=COL_YIELD_NIY).value)
                capval = _to_number(ws.cell(row=row, column=COL_CAPVAL_PSF).value)

                # --- Rule 1: Quarter from Date (C from B) ---
                if not quarter_val and date_val:
                    derived = _derive_quarter_from_cell(date_val)
                    if derived:
                        # Copy quarter to col C but KEEP the original date in col B.
                        # Clearing col B would make the row look like a pipeline row
                        # (no date) and get deleted on the next "clear old" run.
                        ws.cell(row=row, column=COL_QUARTER, value=derived)
                        quarter_val = derived
                        changes += 1
                        summary["details"].append(
                            f"Row {row}: derived Quarter '{derived}' from Date '{date_val}'"
                        )

                # --- Rule 2: Rent PA from Price & Yield (I from L, M) ---
                if not rent_pa and price and yield_niy:
                    rent_pa = round((price * ACQUISITION_COST_FACTOR) * yield_niy, 2)
                    ws.cell(row=row, column=COL_RENT_PA, value=rent_pa)
                    changes += 1
                    summary["details"].append(
                        f"Row {row}: derived Rent PA {rent_pa:,.0f} from Price & Yield"
                    )

                # --- Rule 3: Rent PA from PSF & Area (I from J, H) ---
                if not rent_pa and rent_psf and area:
                    rent_pa = round(rent_psf * area, 2)
                    ws.cell(row=row, column=COL_RENT_PA, value=rent_pa)
                    changes += 1
                    summary["details"].append(
                        f"Row {row}: derived Rent PA {rent_pa:,.0f} from PSF & Area"
                    )

                # --- Rule 4: Rent PSF from PA & Area (J from I, H) ---
                if not rent_psf and rent_pa and area:
                    rent_psf = round(rent_pa / area, 2)
                    ws.cell(row=row, column=COL_RENT_PSF, value=rent_psf)
                    changes += 1
                    summary["details"].append(
                        f"Row {row}: derived Rent PSF {rent_psf:.2f} from PA & Area"
                    )

                # --- Rule 5: Area from Rent PA & PSF (H from I, J) ---
                if not area and rent_pa and rent_psf:
                    area = round(rent_pa / rent_psf, 0)
                    ws.cell(row=row, column=COL_AREA, value=area)
                    changes += 1
                    summary["details"].append(
                        f"Row {row}: derived Area {area:,.0f} from Rent PA & PSF"
                    )

                # --- Rule 6: Cap Val PSF from Price & Area (O from L, H) ---
                if not capval and price and area:
                    capval = round(price / area, 2)
                    ws.cell(row=row, column=COL_CAPVAL_PSF, value=capval)
                    changes += 1
                    summary["details"].append(
                        f"Row {row}: derived Cap Val PSF {capval:.2f} from Price & Area"
                    )

            summary["cells_filled"] = changes

            if changes > 0:
                wb.save(str(excel_path))
                logger.info(
                    "Comps cleaner: filled %d cells across %d rows",
                    changes,
                    summary["rows_scanned"],
                )
            else:
                logger.info(
                    "Comps cleaner: no gaps to fill (%d rows scanned)",
                    summary["rows_scanned"],
                )

            wb.close()
            return summary

        except PermissionError as e:
            if attempt < MAX_RETRIES - 1:
                wait = RETRY_DELAY * (2 ** attempt)
                logger.warning(
                    "File locked (attempt %d/%d), retrying in %ds: %s",
                    attempt + 1,
                    MAX_RETRIES,
                    wait,
                    e,
                )
                time.sleep(wait)
            else:
                logger.error("File locked after %d attempts: %s", MAX_RETRIES, e)
                return summary

        except Exception as e:
            logger.error("Comps cleaner error: %s", e)
            return summary

    return summary


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_number(value) -> float | None:
    """Convert a cell value to a number, or None if not numeric / zero."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value != 0 else None
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("\u00a3", "").replace(" ", "").strip()
        if not cleaned:
            return None
        try:
            result = float(cleaned)
            return result if result != 0 else None
        except ValueError:
            return None
    return None


def _is_quarter_string(value: str) -> bool:
    """Check if a string is a quarter reference like '2025 Q2' or 'Q2 2025'."""
    s = value.strip()
    return bool(
        re.match(r"^\d{4}\s*Q[1-4]$", s, re.IGNORECASE)
        or re.match(r"^Q[1-4]\s*[-/]?\s*\d{4}$", s, re.IGNORECASE)
    )


def _derive_quarter_from_cell(value) -> str | None:
    """Derive quarter from a cell value.

    Handles:
    - datetime objects (openpyxl may deserialise date cells)
    - "2025 Q2" or "Q2 2025" or "Q2-2025" → "2025 Q2" (normalised)
    - "15/03/2025"            → "2025 Q1"
    - "03/2025"               → "2025 Q1"
    """
    if value is None:
        return None

    # Handle datetime objects
    if isinstance(value, datetime):
        quarter = (value.month - 1) // 3 + 1
        return f"{value.year} Q{quarter}"

    if not isinstance(value, str):
        return None

    value = value.strip()
    if not value:
        return None

    # New format: "2025 Q2"
    m = re.match(r"^(\d{4})\s*Q([1-4])$", value, re.IGNORECASE)
    if m:
        return f"{m.group(1)} Q{m.group(2)}"

    # Old format: "Q2 2025", "Q2-2025", "Q2/2025"
    m = re.match(r"^Q([1-4])\s*[-/]?\s*(\d{4})$", value, re.IGNORECASE)
    if m:
        return f"{m.group(2)} Q{m.group(1)}"

    # DD/MM/YYYY
    parts = value.split("/")
    if len(parts) == 3:
        try:
            month = int(parts[1])
            year = int(parts[2])
            if 1 <= month <= 12 and 2000 <= year <= 2100:
                quarter = (month - 1) // 3 + 1
                return f"{year} Q{quarter}"
        except (ValueError, IndexError):
            pass

    # MM/YYYY
    if len(parts) == 2:
        try:
            month = int(parts[0])
            year = int(parts[1])
            if 1 <= month <= 12 and 2000 <= year <= 2100:
                quarter = (month - 1) // 3 + 1
                return f"{year} Q{quarter}"
        except (ValueError, IndexError):
            pass

    return None
