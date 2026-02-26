"""
Brochure Parser
===============
Extracts investment and occupational comparables from brochure files (PDF/Excel).

Two extraction modes:
- Investment comparables: from comparable evidence / transaction sections
- Occupational comparables: from tenancy schedules / letting details

Uses pdfplumber for PDF table extraction (primary), PyMuPDF for text fallback,
and openpyxl for Excel files. For image-based/scanned PDFs where text extraction
fails, falls back to Claude's vision API (renders pages as images via PyMuPDF).
Claude API for structured extraction.
"""

import base64
import json
import logging
import re
from pathlib import Path
from typing import Optional

import anthropic

from email_pipeline.models import (
    BrochureResult,
    DealExtraction,
    InvestmentComp,
    OccupationalComp,
)

logger = logging.getLogger(__name__)

# Minimum characters of extracted text to consider "meaningful".
# Below this threshold, the PDF is likely image-based and we fall back to vision.
_MIN_TEXT_CHARS = 100

# Vision settings
_VISION_DPI = 150         # DPI for rendering PDF pages to images
_VISION_MAX_PAGES = 20    # Max pages to render (keeps API costs reasonable)

# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------

def extract_text_from_pdf(pdf_path: Path) -> str:
    """Extract text from a PDF file.

    Tries pdfplumber first (good for tables), falls back to PyMuPDF.

    Parameters
    ----------
    pdf_path : Path
        Path to the PDF file.

    Returns
    -------
    str
        Extracted text content.
    """
    text = ""

    # Try pdfplumber first (better for tables)
    try:
        import pdfplumber

        with pdfplumber.open(str(pdf_path)) as pdf:
            pages_text = []
            for i, page in enumerate(pdf.pages):
                page_text = page.extract_text() or ""

                # Also extract tables
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if row:
                                cells = [str(c).strip() if c else "" for c in row]
                                page_text += "\n" + " | ".join(cells)

                if page_text.strip():
                    pages_text.append(f"--- Page {i + 1} ---\n{page_text}")

            text = "\n\n".join(pages_text)

    except ImportError:
        logger.warning("pdfplumber not installed, trying PyMuPDF")
    except Exception as e:
        logger.warning("pdfplumber failed on %s: %s, trying PyMuPDF", pdf_path.name, e)

    # Fallback to PyMuPDF
    if not text.strip():
        try:
            import fitz  # PyMuPDF

            doc = fitz.open(str(pdf_path))
            pages_text = []
            for i, page in enumerate(doc):
                page_text = page.get_text()
                if page_text.strip():
                    pages_text.append(f"--- Page {i + 1} ---\n{page_text}")
            doc.close()
            text = "\n\n".join(pages_text)

        except ImportError:
            logger.error("Neither pdfplumber nor PyMuPDF available")
        except Exception as e:
            logger.error("PyMuPDF failed on %s: %s", pdf_path.name, e)

    return text


def extract_text_from_excel(excel_path: Path) -> str:
    """Extract text from an Excel file.

    Reads all sheets and formats as text tables.

    Parameters
    ----------
    excel_path : Path
        Path to the Excel file.

    Returns
    -------
    str
        Extracted text content.
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(excel_path), data_only=True, read_only=True)
        sheets_text = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):  # Skip empty rows
                    rows.append(" | ".join(cells))

            if rows:
                sheets_text.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))

        wb.close()
        return "\n\n".join(sheets_text)

    except ImportError:
        logger.error("openpyxl not installed")
        return ""
    except Exception as e:
        logger.error("Failed to read Excel %s: %s", excel_path.name, e)
        return ""


def extract_text(file_path: Path) -> str:
    """Extract text from a brochure file (PDF or Excel).

    Parameters
    ----------
    file_path : Path
        Path to the file.

    Returns
    -------
    str
        Extracted text content.
    """
    suffix = file_path.suffix.lower()

    if suffix == ".pdf":
        return extract_text_from_pdf(file_path)
    elif suffix in (".xlsx", ".xls"):
        return extract_text_from_excel(file_path)
    else:
        logger.warning("Unsupported file type: %s", suffix)
        return ""


# ---------------------------------------------------------------------------
# Vision fallback — render PDF pages as images for Claude's vision API
# ---------------------------------------------------------------------------

def _render_pdf_pages(pdf_path: Path, dpi: int = _VISION_DPI, max_pages: int = _VISION_MAX_PAGES) -> list[bytes]:
    """Render PDF pages to PNG image bytes using PyMuPDF.

    Parameters
    ----------
    pdf_path : Path
        Path to the PDF file.
    dpi : int
        Resolution for rendering (default 150).
    max_pages : int
        Maximum number of pages to render.

    Returns
    -------
    list[bytes]
        List of PNG image bytes, one per page.
    """
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(str(pdf_path))
        total_pages = len(doc)
        page_images = []
        num_pages = min(total_pages, max_pages)

        for i in range(num_pages):
            page = doc[i]
            # Render at specified DPI (default 72 → scale factor = dpi/72)
            zoom = dpi / 72.0
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            page_images.append(pix.tobytes("png"))

        doc.close()

        if total_pages > max_pages:
            logger.info("  PDF has %d pages, rendered first %d for vision", total_pages, max_pages)

        return page_images

    except ImportError:
        logger.error("PyMuPDF not available for vision rendering")
        return []
    except Exception as e:
        logger.error("Failed to render PDF %s as images: %s", pdf_path.name, e)
        return []


def _build_vision_content(page_images: list[bytes], prompt_text: str) -> list[dict]:
    """Build Claude API content blocks: images first, then text prompt.

    Parameters
    ----------
    page_images : list[bytes]
        PNG image bytes for each page.
    prompt_text : str
        The extraction prompt (text-only, no {text} placeholder).

    Returns
    -------
    list[dict]
        Content blocks for the Claude messages API.
    """
    blocks: list[dict] = []

    for png_bytes in page_images:
        b64 = base64.standard_b64encode(png_bytes).decode("utf-8")
        blocks.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": b64,
            },
        })

    blocks.append({
        "type": "text",
        "text": prompt_text,
    })

    return blocks


# ---------------------------------------------------------------------------
# AI extraction prompts
# ---------------------------------------------------------------------------

INVESTMENT_COMPS_PROMPT = """You are a commercial property data extraction tool. Your task is to extract **investment comparable evidence** from the brochure text below using a strict two-step process.

CRITICAL RULES:
- ONLY extract transactions that are explicitly described in the text with specific details (a named property, a price or yield, and a location).
- NEVER invent, estimate, or fabricate any data. Every value you return must come directly from the text.
- If a field is not explicitly stated in the text, use null. Do NOT calculate or derive values.
- If the brochure contains NO investment comparable evidence at all, return an empty JSON result (see format below).
- Do NOT confuse the SUBJECT PROPERTY being marketed with comparable evidence. The property being sold in this brochure is NOT a comparable — comparables are OTHER transactions referenced for benchmarking.
- General market commentary (e.g. "yields have compressed to 5%") is NOT a comparable transaction. You need a specific property name and at least a price or yield.

Investment comparables are recent property transactions (sales) referenced in the brochure as benchmarking evidence. They are typically found in sections labelled "Comparable Evidence", "Investment Comparables", "Market Transactions", "Recent Sales", or similar.

## TWO-STEP EXTRACTION PROCESS

**STEP 1 — EVIDENCE**: For each comparable, find and quote the EXACT sentence(s) or table row from the brochure that describes the transaction. Copy the text verbatim. If you cannot find a direct quote, do NOT create a comparable entry.

**STEP 2 — EXTRACTION**: Based ONLY on the quoted evidence from Step 1, extract these fields:
- **town**: Town/city (as stated in the quote)
- **address**: Property name or address (as stated in the quote)
- **date**: Transaction date (DD/MM/YYYY or MM/YYYY or year). null if not in the quote.
- **style**: "Multi-Let", "Single-Let", "Industrial", "Office", "Retail", "Logistics", "Mixed Use", "Portfolio", or null
- **units**: Number of units. null if not in the quote.
- **area_sqft**: Total area in sq ft (convert from sq m if needed: 1 sq m = 10.764 sq ft). null if not in the quote.
- **rent_pa**: Passing rent per annum in £. null if not in the quote.
- **rent_psf**: Rent per sq ft. null if not in the quote.
- **awultc**: Average weighted unexpired lease term (years). null if not in the quote.
- **price**: Sale price in £. null if not in the quote.
- **yield_niy**: Net Initial Yield as a percentage (e.g. 6.5). null if not in the quote.
- **reversionary_yield**: Reversionary yield as a percentage. null if not in the quote.
- **capval_psf**: Capital value per sq ft in £. null if not in the quote.
- **vendor**: Vendor/seller name. null if not in the quote.
- **purchaser**: Purchaser/buyer name. null if not in the quote.

Return ONLY a valid JSON object with this structure:
{{
  "comparables": [
    {{
      "evidence": "the exact verbatim quote from the brochure that describes this transaction",
      "town": "...",
      "address": "...",
      "date": null,
      "style": null,
      "units": null,
      "area_sqft": null,
      "rent_pa": null,
      "rent_psf": null,
      "awultc": null,
      "price": null,
      "yield_niy": null,
      "reversionary_yield": null,
      "capval_psf": null,
      "vendor": null,
      "purchaser": null
    }}
  ]
}}

If no comparables found, return: {{"comparables": []}}

## Brochure text:

{text}
"""

OCCUPATIONAL_COMPS_PROMPT = """You are a commercial property data extraction tool. Extract ALL occupational / letting data from the brochure text below.

There are TWO types of occupational data to look for. You must extract BOTH:

**TYPE 1 — TENANCY SCHEDULE**: The subject property's own current leases.
Found in sections labelled "Tenancy Schedule", "Income Schedule", "Current Tenancies", or similar.
These have tenant names, unit numbers, lease dates, passing rent, ERV, etc.

**TYPE 2 — OCCUPATIONAL / RENTAL COMPARABLES**: Recent lettings at OTHER properties, used as benchmarking evidence.
Found in sections labelled "Occupational Comparables", "Occupational Rental Comparables", "Letting Evidence", "Rental Evidence", "Letting Comparables", or similar.
These typically have just a date, address, size, and rent (no tenant name or lease dates).
They may appear as a small table next to or below the tenancy schedule.

IMPORTANT: A brochure may contain MULTIPLE tenancy schedules and MULTIPLE comparable tables (e.g. one per estate in a portfolio). Extract ALL of them.

EXCLUDE the following — these are NOT individual tenancies or occupational comparables:
- **Portfolio / estate summary rows**: Aggregate lines that summarise an entire estate or portfolio (e.g. "Multiple tenants", "Various (15 units)", total WAULT, overall reversion %, overall vacancy %). These describe the whole property, not a single letting.
- **Totals / subtotals rows**: Lines labelled "Total", "Sub-total", "Aggregate", or similar.
- **The subject property's headline metrics** from the executive summary (e.g. total rent, total area, blended yield). These are deal-level figures, not individual tenancies.
- **INVESTMENT COMPARABLES**: Tables or rows showing property SALES with columns like Net Initial Yield (NIY), Capital Value (Cap Val), AWULTC, Sale Price, or Purchaser/Vendor. These are investment transactions, NOT occupational lettings. Even if they appear on the same page as letting data, do NOT extract them here. If a table has yield, price, or cap val columns it is an investment comparable, not an occupational one.

For each INDIVIDUAL entry (whether tenancy or comparable), extract:
- **tenant_name**: Tenant / occupier name (null for rental comparables)
- **unit_name**: Unit identifier or name (null if not stated)
- **address**: Property address
- **town**: Town/city
- **postcode**: Postcode (null if not stated)
- **size_sqft**: Unit size in sq ft (convert from sq m if needed)
- **rent_pa**: Annual rent in £ (null if only rent PSF given)
- **rent_psf**: Rent per sq ft
- **lease_start**: Lease start date (DD/MM/YYYY) (null for rental comparables)
- **lease_expiry**: Lease expiry date (DD/MM/YYYY) (null for rental comparables)
- **break_date**: Break date (DD/MM/YYYY or null)
- **rent_review_date**: Next rent review date (DD/MM/YYYY or null)
- **lease_term_years**: Total lease term in years (null for rental comparables)
- **notes**: Any other relevant notes
- **entry_type**: "tenancy" for Type 1 entries, "comparable" for Type 2 entries
- **comp_date**: For rental comparables, the transaction date (e.g. "Q4 2025" or "01/2025"). null for tenancy entries.

Return ONLY a valid JSON array. Use null for unknown values. If no data found, return: []

## Brochure text:

{text}
"""

DEAL_FROM_BROCHURE_PROMPT = """You are a commercial property analyst. Extract the main **deal / property details** from this investment brochure.

This is a brochure for a commercial property being marketed for sale. Extract the key investment metrics.

Extract:
- **asset_name**: Property or estate name
- **town**: Town/city
- **address**: Full address
- **classification**: One of "Multi-Let Industrial", "Single-Let Industrial", "Multi-Let Office", "Single-Let Office", "Retail", "Mixed Use", "Logistics", "Land", "Portfolio", "Other"
- **area_acres**: Site area in acres (null if not stated)
- **area_sqft**: Total building area in sq ft (convert from sq m if needed)
- **rent_pa**: Total passing rent per annum in £
- **rent_psf**: Rent per sq ft
- **asking_price**: Asking/quoting price in £
- **net_yield**: Net Initial Yield as percentage (e.g. 6.5)
- **reversionary_yield**: Reversionary yield as percentage
- **confidence**: Your confidence in the extraction (0.0 to 1.0)

Return ONLY valid JSON:
{{
    "asset_name": "...",
    "town": "...",
    "address": "...",
    "classification": "...",
    "area_acres": null,
    "area_sqft": null,
    "rent_pa": null,
    "rent_psf": null,
    "asking_price": null,
    "net_yield": null,
    "reversionary_yield": null,
    "confidence": 0.8
}}

## Brochure text:

{text}
"""

# Vision-mode prompt suffixes (no {text} placeholder — images are sent as content blocks)
_DEAL_VISION_PROMPT = """You are a commercial property analyst. The images above are pages from an investment brochure for a commercial property being marketed for sale. Extract the key deal / property details.

Extract:
- **asset_name**: Property or estate name
- **town**: Town/city
- **address**: Full address
- **classification**: One of "Multi-Let Industrial", "Single-Let Industrial", "Multi-Let Office", "Single-Let Office", "Retail", "Mixed Use", "Logistics", "Land", "Portfolio", "Other"
- **area_acres**: Site area in acres (null if not stated)
- **area_sqft**: Total building area in sq ft (convert from sq m if needed)
- **rent_pa**: Total passing rent per annum in £
- **rent_psf**: Rent per sq ft
- **asking_price**: Asking/quoting price in £
- **net_yield**: Net Initial Yield as percentage (e.g. 6.5)
- **reversionary_yield**: Reversionary yield as percentage
- **confidence**: Your confidence in the extraction (0.0 to 1.0)

Return ONLY valid JSON:
{
    "asset_name": "...",
    "town": "...",
    "address": "...",
    "classification": "...",
    "area_acres": null,
    "area_sqft": null,
    "rent_pa": null,
    "rent_psf": null,
    "asking_price": null,
    "net_yield": null,
    "reversionary_yield": null,
    "confidence": 0.8
}"""

_INVESTMENT_COMPS_VISION_PROMPT = """You are a commercial property data extraction tool. The images above are pages from an investment brochure. Extract ONLY investment comparable evidence that is EXPLICITLY SHOWN in these pages, using a strict two-step process.

CRITICAL RULES:
- ONLY extract transactions that are explicitly shown with specific details (a named property, a price or yield, and a location).
- NEVER invent, estimate, or fabricate any data. Every value must be visible in the images.
- If a field is not visible, use null. Do NOT calculate or derive values.
- If these pages contain NO investment comparable evidence, return: {"comparables": []}
- Do NOT confuse the SUBJECT PROPERTY being marketed with comparable evidence. Comparables are OTHER transactions referenced for benchmarking.
- General market commentary is NOT a comparable transaction.

Look for sections labelled "Comparable Evidence", "Investment Comparables", "Market Transactions", "Recent Sales", or similar.

## TWO-STEP EXTRACTION PROCESS

**STEP 1 — EVIDENCE**: For each comparable, transcribe the EXACT text or table row visible in the image that describes the transaction. If you cannot read specific text describing a transaction, do NOT create a comparable entry.

**STEP 2 — EXTRACTION**: Based ONLY on the transcribed evidence from Step 1, extract these fields:
- **town**: Town/city. null if not in the evidence.
- **address**: Property name or address. null if not in the evidence.
- **date**: Transaction date (DD/MM/YYYY or MM/YYYY or year). null if not in the evidence.
- **style**: "Multi-Let", "Single-Let", "Industrial", "Office", "Retail", "Logistics", "Mixed Use", "Portfolio", or null
- **units**: Number of units. null if not in the evidence.
- **area_sqft**: Total area in sq ft (convert from sq m if needed). null if not in the evidence.
- **rent_pa**: Passing rent per annum in £. null if not in the evidence.
- **rent_psf**: Rent per sq ft. null if not in the evidence.
- **awultc**: Average weighted unexpired lease term (years). null if not in the evidence.
- **price**: Sale price in £. null if not in the evidence.
- **yield_niy**: Net Initial Yield as a percentage (e.g. 6.5). null if not in the evidence.
- **reversionary_yield**: Reversionary yield as a percentage. null if not in the evidence.
- **capval_psf**: Capital value per sq ft in £. null if not in the evidence.
- **vendor**: Vendor/seller name. null if not in the evidence.
- **purchaser**: Purchaser/buyer name. null if not in the evidence.

Return ONLY a valid JSON object with this structure:
{
  "comparables": [
    {
      "evidence": "the exact text transcribed from the image that describes this transaction",
      "town": "...",
      "address": "...",
      "date": null,
      "style": null,
      "units": null,
      "area_sqft": null,
      "rent_pa": null,
      "rent_psf": null,
      "awultc": null,
      "price": null,
      "yield_niy": null,
      "reversionary_yield": null,
      "capval_psf": null,
      "vendor": null,
      "purchaser": null
    }
  ]
}

If no comparables found, return: {"comparables": []}"""

_OCCUPATIONAL_COMPS_VISION_PROMPT = """You are a commercial property data extraction tool. The images above are pages from an investment brochure. Extract ALL occupational / letting data.

There are TWO types of occupational data to look for. You must extract BOTH:

**TYPE 1 — TENANCY SCHEDULE**: The subject property's own current leases.
Found in sections labelled "Tenancy Schedule", "Income Schedule", "Current Tenancies", or similar.
These have tenant names, unit numbers, lease dates, passing rent, ERV, etc.

**TYPE 2 — OCCUPATIONAL / RENTAL COMPARABLES**: Recent lettings at OTHER properties, used as benchmarking evidence.
Found in sections labelled "Occupational Comparables", "Occupational Rental Comparables", "Letting Evidence", "Rental Evidence", "Letting Comparables", or similar.
These typically have just a date, address, size, and rent (no tenant name or lease dates).
They may appear as a small table next to or below the tenancy schedule.

IMPORTANT: A brochure may contain MULTIPLE tenancy schedules and MULTIPLE comparable tables (e.g. one per estate in a portfolio). Extract ALL of them.

EXCLUDE the following — these are NOT individual tenancies or occupational comparables:
- **Portfolio / estate summary rows**: Aggregate lines that summarise an entire estate or portfolio (e.g. "Multiple tenants", "Various (15 units)", total WAULT, overall reversion %, overall vacancy %). These describe the whole property, not a single letting.
- **Totals / subtotals rows**: Lines labelled "Total", "Sub-total", "Aggregate", or similar.
- **The subject property's headline metrics** from the executive summary (e.g. total rent, total area, blended yield). These are deal-level figures, not individual tenancies.
- **INVESTMENT COMPARABLES**: Tables or rows showing property SALES with columns like Net Initial Yield (NIY), Capital Value (Cap Val), AWULTC, Sale Price, or Purchaser/Vendor. These are investment transactions, NOT occupational lettings. Even if they appear on the same page as letting data, do NOT extract them here. If a table has yield, price, or cap val columns it is an investment comparable, not an occupational one.

For each INDIVIDUAL entry (whether tenancy or comparable), extract:
- **tenant_name**: Tenant / occupier name (null for rental comparables)
- **unit_name**: Unit identifier or name (null if not stated)
- **address**: Property address
- **town**: Town/city
- **postcode**: Postcode (null if not stated)
- **size_sqft**: Unit size in sq ft (convert from sq m if needed)
- **rent_pa**: Annual rent in £ (null if only rent PSF given)
- **rent_psf**: Rent per sq ft
- **lease_start**: Lease start date (DD/MM/YYYY) (null for rental comparables)
- **lease_expiry**: Lease expiry date (DD/MM/YYYY) (null for rental comparables)
- **break_date**: Break date (DD/MM/YYYY or null)
- **rent_review_date**: Next rent review date (DD/MM/YYYY or null)
- **lease_term_years**: Total lease term in years (null for rental comparables)
- **notes**: Any other relevant notes
- **entry_type**: "tenancy" for Type 1 entries, "comparable" for Type 2 entries
- **comp_date**: For rental comparables, the transaction date (e.g. "Q4 2025" or "01/2025"). null for tenancy entries.

Return ONLY valid JSON array. Use null for unknown values. If no data found, return: []"""

# ---------------------------------------------------------------------------
# Judge / verification prompts
# ---------------------------------------------------------------------------

_INVESTMENT_COMPS_JUDGE_PROMPT = """You are a verification assistant for commercial property data extraction. You have been given the ORIGINAL BROCHURE TEXT and a set of EXTRACTED COMPARABLES (in JSON).

Your job is to verify each comparable against the original text. For each comparable, check:

1. **EXISTS**: Does this transaction actually appear in the brochure text? Is there a real sentence or table row describing it?
2. **EVIDENCE CHECK**: Does the "evidence" field match actual text from the brochure (not fabricated)?
3. **DATA ACCURACY**: Do the extracted values (price, yield, area, rent, etc.) match what the evidence text says?
4. **NOT THE SUBJECT PROPERTY**: Is this a COMPARABLE transaction, not the property being marketed in this brochure?

For each comparable, return a verdict:
- "keep" — the comparable is genuine and accurately extracted
- "remove" — the comparable appears fabricated, is the subject property, or has no supporting evidence in the text

Return ONLY a valid JSON object:
{{
  "verdicts": [
    {{"index": 0, "verdict": "keep", "reason": "Transaction found in text on page 5"}},
    {{"index": 1, "verdict": "remove", "reason": "No mention of this property anywhere in the brochure"}}
  ]
}}

## Original brochure text:

{text}

## Extracted comparables to verify:

{comps_json}
"""

_INVESTMENT_COMPS_JUDGE_VISION_PROMPT = """You are a verification assistant for commercial property data extraction. You have been shown the ORIGINAL BROCHURE PAGES (as images) and below is a set of EXTRACTED COMPARABLES (in JSON).

Your job is to verify each comparable against the original pages. For each comparable, check:

1. **EXISTS**: Does this transaction actually appear in the brochure pages? Is there a real sentence or table row describing it?
2. **EVIDENCE CHECK**: Does the "evidence" field match actual text visible in the pages (not fabricated)?
3. **DATA ACCURACY**: Do the extracted values (price, yield, area, rent, etc.) match what is visible in the pages?
4. **NOT THE SUBJECT PROPERTY**: Is this a COMPARABLE transaction, not the property being marketed in this brochure?

For each comparable, return a verdict:
- "keep" — the comparable is genuine and accurately extracted
- "remove" — the comparable appears fabricated, is the subject property, or has no supporting evidence in the pages

Return ONLY a valid JSON object:
{
  "verdicts": [
    {"index": 0, "verdict": "keep", "reason": "Transaction visible in table on page 5"},
    {"index": 1, "verdict": "remove", "reason": "No mention of this property anywhere in the brochure"}
  ]
}

## Extracted comparables to verify:

"""


# ---------------------------------------------------------------------------
# Main extraction functions
# ---------------------------------------------------------------------------

_DEFAULT_JUDGE_MODEL = None  # None = use same model as extractor


def parse_brochure(
    file_path: Path,
    api_key: str,
    source_deal: str = "",
    model: str = "claude-sonnet-4-6",
    judge_model: str = _DEFAULT_JUDGE_MODEL,
    extract_deal: bool = True,
    extract_investment_comps: bool = True,
    extract_occupational_comps: bool = True,
    verify_comps: bool = True,
) -> BrochureResult:
    """Parse a brochure file and extract all comparables.

    Parameters
    ----------
    file_path : Path
        Path to the brochure file (PDF or Excel).
    api_key : str
        Anthropic API key.
    source_deal : str
        Name of the source deal (for occupational comp tracking).
    model : str
        Claude model to use for extraction (Sonnet recommended).
    judge_model : str or None
        Model for verification judge. Defaults to None which uses the
        same model as the extractor (recommended — prompt caching means
        the cost overhead is minimal). Set to a different model ID to
        use a different model, or set verify_comps=False to disable.
    extract_deal : bool
        Whether to extract deal/property details.
    extract_investment_comps : bool
        Whether to extract investment comparables.
    extract_occupational_comps : bool
        Whether to extract occupational comparables.
    verify_comps : bool
        Whether to run the judge verification step on investment comps.
        Requires judge_model to be set.

    Returns
    -------
    BrochureResult
        Extraction results.
    """
    file_path = Path(file_path)
    result = BrochureResult(file_path=str(file_path))

    # Step 1: Extract text
    logger.info("Extracting text from %s...", file_path.name)
    text = extract_text(file_path)

    use_vision = False

    if not text.strip() or len(text.strip()) < _MIN_TEXT_CHARS:
        # Text extraction failed or returned too little — try vision for PDFs
        if file_path.suffix.lower() == ".pdf":
            logger.info("  Text extraction returned %d chars (below %d threshold) — trying vision mode",
                        len(text.strip()), _MIN_TEXT_CHARS)
            page_images = _render_pdf_pages(file_path)
            if page_images:
                use_vision = True
                logger.info("  Rendered %d pages as images for vision API", len(page_images))
            else:
                result.error_message = f"No text and vision rendering failed for {file_path.name}"
                logger.warning(result.error_message)
                return result
        else:
            result.error_message = f"No text extracted from {file_path.name}"
            logger.warning(result.error_message)
            return result

    if not use_vision:
        logger.info("  Extracted %d characters of text", len(text))
        # Truncate very large documents to stay within token limits.
        # Sonnet's context is 200K tokens (~800K chars). 60K chars ≈ 15K
        # tokens, which leaves ample room for prompts + response.
        # The old limit of 15K was silently dropping tenancy schedules
        # and comparable tables from longer portfolio brochures.
        max_chars = 60000
        if len(text) > max_chars:
            logger.warning("  Text is %d chars — truncating to %d", len(text), max_chars)
            text = text[:max_chars] + "\n\n[... truncated ...]"

    client = anthropic.Anthropic(api_key=api_key)

    # Step 2: Extract deal details (if requested)
    if extract_deal:
        try:
            if use_vision:
                result.deal_extraction = _extract_deal_from_vision(client, page_images, model)
            else:
                result.deal_extraction = _extract_deal_from_text(client, text, model)
            if result.deal_extraction and not source_deal:
                source_deal = result.deal_extraction.asset_name or file_path.stem
            logger.info("  Deal extraction: %s",
                        result.deal_extraction.asset_name if result.deal_extraction else "None")
        except Exception as e:
            logger.error("  Deal extraction failed: %s", e)

    # Step 3: Extract investment comparables (if requested)
    if extract_investment_comps:
        try:
            if use_vision:
                comps, raw_items = _extract_investment_comps_vision(client, page_images, model)
            else:
                comps, raw_items = _extract_investment_comps(client, text, model)
            logger.info("  Investment comps: %d extracted", len(comps))

            # Step 3b: Judge verification (if enabled and comps were found)
            if verify_comps and comps:
                _judge = judge_model or model  # Default: same model as extractor
                logger.info("  Running judge verification (%s)...", _judge)
                if use_vision:
                    comps = _verify_investment_comps_vision(
                        client, page_images, comps, raw_items, _judge
                    )
                else:
                    comps = _verify_investment_comps(
                        client, text, comps, raw_items, _judge
                    )

            result.investment_comps = comps
        except Exception as e:
            logger.error("  Investment comp extraction failed: %s", e)

    # Step 4: Extract occupational comparables (if requested)
    if extract_occupational_comps:
        try:
            if use_vision:
                result.occupational_comps = _extract_occupational_comps_vision(
                    client, page_images, source_deal or file_path.stem, model
                )
            else:
                result.occupational_comps = _extract_occupational_comps(
                    client, text, source_deal or file_path.stem, model
                )
            logger.info("  Occupational comps: %d found", len(result.occupational_comps))
        except Exception as e:
            logger.error("  Occupational comp extraction failed: %s", e)

    return result


# ---------------------------------------------------------------------------
# Internal extraction functions
# ---------------------------------------------------------------------------

def _extract_deal_from_text(
    client: anthropic.Anthropic,
    text: str,
    model: str,
) -> Optional[DealExtraction]:
    """Extract deal details from brochure text using Claude API."""
    prompt = DEAL_FROM_BROCHURE_PROMPT.format(text=text)

    message = client.messages.create(
        model=model,
        max_tokens=1000,
        messages=[{"role": "user", "content": prompt}],
    )

    data = _parse_json_response(message.content[0].text.strip())

    return DealExtraction(
        date="",  # Not available from brochure
        agent="",  # Not available from brochure
        asset_name=_clean_str(data.get("asset_name", "")),
        country="England",
        town=_clean_str(data.get("town", "")),
        address=_clean_str(data.get("address", "")),
        classification=_clean_str(data.get("classification", "")),
        area_acres=_to_float(data.get("area_acres")),
        area_sqft=_to_float(data.get("area_sqft")),
        rent_pa=_to_float(data.get("rent_pa")),
        rent_psf=_to_float(data.get("rent_psf")),
        asking_price=_to_float(data.get("asking_price")),
        net_yield=_to_float(data.get("net_yield")),
        reversionary_yield=_to_float(data.get("reversionary_yield")),
        confidence=data.get("confidence", 0.7),
        raw_source="brochure",
    )


def _extract_investment_comps(
    client: anthropic.Anthropic,
    text: str,
    model: str,
) -> tuple[list[InvestmentComp], list[dict]]:
    """Extract investment comparables from brochure text using Claude API.

    Returns
    -------
    tuple[list[InvestmentComp], list[dict]]
        (comps, raw_items) — the raw_items include the 'evidence' field
        for use by the judge verification step.
    """
    prompt = INVESTMENT_COMPS_PROMPT.format(text=text)

    message = client.messages.create(
        model=model,
        max_tokens=3000,
        temperature=0,
        messages=[{
            "role": "user",
            "content": [{
                "type": "text",
                "text": prompt,
                "cache_control": {"type": "ephemeral"},
            }],
        }],
    )

    data = _parse_json_response(message.content[0].text.strip())

    # Handle CoVe format {"comparables": [...]} or legacy bare array [...]
    items = _unwrap_comps_response(data)
    if items is None:
        return [], []

    comps = []
    for item in items:
        date_str = item.get("date")
        evidence = item.get("evidence", "")
        if evidence:
            logger.debug("  Evidence: %s", evidence[:120])
        comps.append(
            InvestmentComp(
                town=item.get("town", ""),
                address=item.get("address", ""),
                date=date_str,
                quarter=_derive_quarter(date_str),
                style=item.get("style"),
                units=_to_int(item.get("units")),
                area_sqft=_to_float(item.get("area_sqft")),
                rent_pa=_to_float(item.get("rent_pa")),
                rent_psf=_to_float(item.get("rent_psf")),
                awultc=_to_float(item.get("awultc")),
                price=_to_float(item.get("price")),
                yield_niy=_to_float(item.get("yield_niy")),
                reversionary_yield=_to_float(item.get("reversionary_yield")),
                capval_psf=_to_float(item.get("capval_psf")),
                vendor=item.get("vendor"),
                purchaser=item.get("purchaser"),
            )
        )

    return comps, items


def _extract_occupational_comps(
    client: anthropic.Anthropic,
    text: str,
    source_deal: str,
    model: str,
) -> list[OccupationalComp]:
    """Extract occupational comparables from brochure text using Claude API.

    Extracts both tenancy schedule entries (the subject property's own leases)
    and occupational rental comparables (external lettings used as evidence).
    """
    prompt = OCCUPATIONAL_COMPS_PROMPT.format(text=text)

    message = client.messages.create(
        model=model,
        max_tokens=6000,
        temperature=0,
        messages=[{"role": "user", "content": prompt}],
    )

    data = _parse_json_response(message.content[0].text.strip())

    if not isinstance(data, list):
        return []

    comps = []
    for item in data:
        entry_type = item.get("entry_type", "tenancy")
        if entry_type not in ("tenancy", "comparable"):
            entry_type = "tenancy"  # Default to tenancy if unrecognised

        comps.append(
            OccupationalComp(
                source_deal=source_deal,
                tenant_name=item.get("tenant_name") or "",
                entry_type=entry_type,
                unit_name=item.get("unit_name"),
                address=item.get("address", ""),
                town=item.get("town", ""),
                postcode=item.get("postcode"),
                size_sqft=_to_float(item.get("size_sqft")),
                rent_pa=_to_float(item.get("rent_pa")),
                rent_psf=_to_float(item.get("rent_psf")),
                lease_start=item.get("lease_start"),
                lease_expiry=item.get("lease_expiry"),
                break_date=item.get("break_date"),
                rent_review_date=item.get("rent_review_date"),
                lease_term_years=_to_float(item.get("lease_term_years")),
                comp_date=item.get("comp_date"),
                notes=item.get("notes"),
            )
        )

    return comps


# ---------------------------------------------------------------------------
# Judge / verification functions
# ---------------------------------------------------------------------------

def _verify_investment_comps(
    client: anthropic.Anthropic,
    text: str,
    comps: list[InvestmentComp],
    raw_items: list[dict],
    model: str,
) -> list[InvestmentComp]:
    """Verify extracted investment comps against the original brochure text.

    Uses a cheaper/faster model as a 'judge' to cross-check each comparable
    against the source material. Prompt caching means the brochure text
    (already cached from the extractor call) costs only 10% of normal input.

    Parameters
    ----------
    client : anthropic.Anthropic
        Anthropic API client.
    text : str
        Original brochure text (same as passed to extractor).
    comps : list[InvestmentComp]
        Extracted comparables from the extractor.
    raw_items : list[dict]
        Raw JSON items (with 'evidence' field) from the extractor response.
    model : str
        Model to use for verification (e.g. claude-haiku-4-5).

    Returns
    -------
    list[InvestmentComp]
        Filtered list with hallucinations removed.
    """
    if not comps:
        return comps

    comps_json = json.dumps(raw_items, indent=2)
    prompt = _INVESTMENT_COMPS_JUDGE_PROMPT.format(text=text, comps_json=comps_json)

    try:
        message = client.messages.create(
            model=model,
            max_tokens=1500,
            temperature=0,
            messages=[{
                "role": "user",
                "content": [{
                    "type": "text",
                    "text": prompt,
                    "cache_control": {"type": "ephemeral"},
                }],
            }],
        )

        verdicts_data = _parse_json_response(message.content[0].text.strip())

        verdicts = verdicts_data.get("verdicts", [])
        if not verdicts:
            logger.warning("  Judge returned no verdicts — keeping all comps")
            return comps

        # Build removal set
        remove_indices = set()
        for v in verdicts:
            idx = v.get("index")
            verdict = v.get("verdict", "").lower()
            reason = v.get("reason", "")
            if verdict == "remove":
                remove_indices.add(idx)
                logger.info("  Judge REMOVED comp %d (%s): %s",
                           idx, comps[idx].address if idx < len(comps) else "?", reason)
            else:
                logger.debug("  Judge kept comp %d: %s", idx, reason)

        if remove_indices:
            filtered = [c for i, c in enumerate(comps) if i not in remove_indices]
            logger.info("  Judge: %d/%d comps kept (%d removed)",
                       len(filtered), len(comps), len(remove_indices))
            return filtered
        else:
            logger.info("  Judge: all %d comps verified", len(comps))
            return comps

    except Exception as e:
        logger.warning("  Judge verification failed (%s) — keeping all comps", e)
        return comps


def _verify_investment_comps_vision(
    client: anthropic.Anthropic,
    page_images: list[bytes],
    comps: list[InvestmentComp],
    raw_items: list[dict],
    model: str,
) -> list[InvestmentComp]:
    """Verify extracted investment comps against the original brochure page images.

    Vision-mode judge: sends the same page images + the extracted JSON to a
    cheaper model for cross-checking. Prompt caching on the images means the
    second call costs ~10% of the image input tokens.

    Parameters
    ----------
    client : anthropic.Anthropic
        Anthropic API client.
    page_images : list[bytes]
        PNG image bytes for each brochure page.
    comps : list[InvestmentComp]
        Extracted comparables from the extractor.
    raw_items : list[dict]
        Raw JSON items (with 'evidence' field) from the extractor response.
    model : str
        Model to use for verification (e.g. claude-haiku-4-5).

    Returns
    -------
    list[InvestmentComp]
        Filtered list with hallucinations removed.
    """
    if not comps:
        return comps

    comps_json = json.dumps(raw_items, indent=2)

    # Build content blocks: images (cached) + judge prompt with comps JSON
    blocks: list[dict] = []
    for png_bytes in page_images:
        b64 = base64.standard_b64encode(png_bytes).decode("utf-8")
        blocks.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": b64,
            },
        })

    # Mark the last image block for caching (covers all images)
    if blocks:
        blocks[-1]["cache_control"] = {"type": "ephemeral"}

    blocks.append({
        "type": "text",
        "text": _INVESTMENT_COMPS_JUDGE_VISION_PROMPT + comps_json,
    })

    try:
        message = client.messages.create(
            model=model,
            max_tokens=1500,
            temperature=0,
            messages=[{"role": "user", "content": blocks}],
        )

        verdicts_data = _parse_json_response(message.content[0].text.strip())

        verdicts = verdicts_data.get("verdicts", [])
        if not verdicts:
            logger.warning("  Judge (vision) returned no verdicts — keeping all comps")
            return comps

        remove_indices = set()
        for v in verdicts:
            idx = v.get("index")
            verdict = v.get("verdict", "").lower()
            reason = v.get("reason", "")
            if verdict == "remove":
                remove_indices.add(idx)
                logger.info("  Judge REMOVED comp %d (%s): %s",
                           idx, comps[idx].address if idx < len(comps) else "?", reason)
            else:
                logger.debug("  Judge kept comp %d: %s", idx, reason)

        if remove_indices:
            filtered = [c for i, c in enumerate(comps) if i not in remove_indices]
            logger.info("  Judge (vision): %d/%d comps kept (%d removed)",
                       len(filtered), len(comps), len(remove_indices))
            return filtered
        else:
            logger.info("  Judge (vision): all %d comps verified", len(comps))
            return comps

    except Exception as e:
        logger.warning("  Judge (vision) verification failed (%s) — keeping all comps", e)
        return comps


# ---------------------------------------------------------------------------
# Vision-mode extraction functions (image-based PDFs)
# ---------------------------------------------------------------------------

def _extract_deal_from_vision(
    client: anthropic.Anthropic,
    page_images: list[bytes],
    model: str,
) -> Optional[DealExtraction]:
    """Extract deal details from brochure page images using Claude's vision API."""
    content = _build_vision_content(page_images, _DEAL_VISION_PROMPT)

    message = client.messages.create(
        model=model,
        max_tokens=1000,
        messages=[{"role": "user", "content": content}],
    )

    data = _parse_json_response(message.content[0].text.strip())

    return DealExtraction(
        date="",
        agent="",
        asset_name=_clean_str(data.get("asset_name", "")),
        country="England",
        town=_clean_str(data.get("town", "")),
        address=_clean_str(data.get("address", "")),
        classification=_clean_str(data.get("classification", "")),
        area_acres=_to_float(data.get("area_acres")),
        area_sqft=_to_float(data.get("area_sqft")),
        rent_pa=_to_float(data.get("rent_pa")),
        rent_psf=_to_float(data.get("rent_psf")),
        asking_price=_to_float(data.get("asking_price")),
        net_yield=_to_float(data.get("net_yield")),
        reversionary_yield=_to_float(data.get("reversionary_yield")),
        confidence=data.get("confidence", 0.7),
        raw_source="brochure_vision",
    )


def _extract_investment_comps_vision(
    client: anthropic.Anthropic,
    page_images: list[bytes],
    model: str,
) -> tuple[list[InvestmentComp], list[dict]]:
    """Extract investment comparables from brochure page images using vision API.

    Returns
    -------
    tuple[list[InvestmentComp], list[dict]]
        (comps, raw_items) — the raw_items include the 'evidence' field
        for use by the judge verification step.
    """
    # Build content with cache_control on the last image block
    # so the judge call can reuse the cached images
    blocks: list[dict] = []
    for png_bytes in page_images:
        b64 = base64.standard_b64encode(png_bytes).decode("utf-8")
        blocks.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": b64,
            },
        })

    # Mark last image for caching (covers all preceding images)
    if blocks:
        blocks[-1]["cache_control"] = {"type": "ephemeral"}

    blocks.append({
        "type": "text",
        "text": _INVESTMENT_COMPS_VISION_PROMPT,
    })

    message = client.messages.create(
        model=model,
        max_tokens=3000,
        temperature=0,
        messages=[{"role": "user", "content": blocks}],
    )

    data = _parse_json_response(message.content[0].text.strip())

    # Handle CoVe format {"comparables": [...]} or legacy bare array [...]
    items = _unwrap_comps_response(data)
    if items is None:
        return [], []

    comps = []
    for item in items:
        date_str = item.get("date")
        evidence = item.get("evidence", "")
        if evidence:
            logger.debug("  Evidence (vision): %s", evidence[:120])
        comps.append(
            InvestmentComp(
                town=item.get("town", ""),
                address=item.get("address", ""),
                date=date_str,
                quarter=_derive_quarter(date_str),
                style=item.get("style"),
                units=_to_int(item.get("units")),
                area_sqft=_to_float(item.get("area_sqft")),
                rent_pa=_to_float(item.get("rent_pa")),
                rent_psf=_to_float(item.get("rent_psf")),
                awultc=_to_float(item.get("awultc")),
                price=_to_float(item.get("price")),
                yield_niy=_to_float(item.get("yield_niy")),
                reversionary_yield=_to_float(item.get("reversionary_yield")),
                capval_psf=_to_float(item.get("capval_psf")),
                vendor=item.get("vendor"),
                purchaser=item.get("purchaser"),
            )
        )

    return comps, items


def _extract_occupational_comps_vision(
    client: anthropic.Anthropic,
    page_images: list[bytes],
    source_deal: str,
    model: str,
) -> list[OccupationalComp]:
    """Extract occupational comparables from brochure page images using vision API.

    Extracts both tenancy schedule entries and occupational rental comparables.
    """
    content = _build_vision_content(page_images, _OCCUPATIONAL_COMPS_VISION_PROMPT)

    message = client.messages.create(
        model=model,
        max_tokens=6000,
        temperature=0,
        messages=[{"role": "user", "content": content}],
    )

    data = _parse_json_response(message.content[0].text.strip())

    if not isinstance(data, list):
        return []

    comps = []
    for item in data:
        entry_type = item.get("entry_type", "tenancy")
        if entry_type not in ("tenancy", "comparable"):
            entry_type = "tenancy"

        comps.append(
            OccupationalComp(
                source_deal=source_deal,
                tenant_name=item.get("tenant_name") or "",
                entry_type=entry_type,
                unit_name=item.get("unit_name"),
                address=item.get("address", ""),
                town=item.get("town", ""),
                postcode=item.get("postcode"),
                size_sqft=_to_float(item.get("size_sqft")),
                rent_pa=_to_float(item.get("rent_pa")),
                rent_psf=_to_float(item.get("rent_psf")),
                lease_start=item.get("lease_start"),
                lease_expiry=item.get("lease_expiry"),
                break_date=item.get("break_date"),
                rent_review_date=item.get("rent_review_date"),
                lease_term_years=_to_float(item.get("lease_term_years")),
                comp_date=item.get("comp_date"),
                notes=item.get("notes"),
            )
        )

    return comps


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _derive_quarter(date_str: Optional[str]) -> Optional[str]:
    """Derive a quarter string (e.g. '2025 Q1') from a date string.

    Handles formats: DD/MM/YYYY, MM/YYYY, YYYY, Q1 2025, 2025 Q1, or None.
    """
    if not date_str or not isinstance(date_str, str):
        return None

    date_str = date_str.strip()

    # Already a quarter string: "2025 Q1" (new format)
    m = re.match(r"^(\d{4})\s*Q([1-4])$", date_str, re.IGNORECASE)
    if m:
        return f"{m.group(1)} Q{m.group(2)}"

    # Already a quarter string: "Q1 2025" (old format) → convert
    m = re.match(r"^Q([1-4])\s*[-/]?\s*(\d{4})$", date_str, re.IGNORECASE)
    if m:
        return f"{m.group(2)} Q{m.group(1)}"

    month = None
    year = None

    # DD/MM/YYYY
    parts = date_str.split("/")
    if len(parts) == 3:
        try:
            month = int(parts[1])
            year = int(parts[2])
        except (ValueError, IndexError):
            pass
    # MM/YYYY
    elif len(parts) == 2:
        try:
            month = int(parts[0])
            year = int(parts[1])
        except (ValueError, IndexError):
            pass
    # Just a year (e.g. "2025")
    elif len(date_str) == 4 and date_str.isdigit():
        year = int(date_str)

    if year is None:
        return None

    if month is None:
        return None  # Can't determine quarter without month

    quarter = (month - 1) // 3 + 1
    return f"{year} Q{quarter}"


def _strip_code_block(text: str) -> str:
    """Strip markdown code block markers from text."""
    if text.startswith("```"):
        lines = text.split("\n")
        json_lines = [l for l in lines if not l.strip().startswith("```")]
        return "\n".join(json_lines)
    return text


def _parse_json_response(text: str):
    """Robustly extract JSON from Claude's response.

    Handles:
    - Clean JSON (object or array)
    - JSON wrapped in ```json ... ``` code blocks
    - Leading/trailing commentary text outside the JSON

    Returns a Python dict or list.
    """
    # 1. Strip code block markers
    if "```" in text:
        lines = text.split("\n")
        json_lines = []
        in_block = False
        for line in lines:
            if line.strip().startswith("```"):
                in_block = not in_block
                continue
            if in_block:
                json_lines.append(line)
        if json_lines:
            text = "\n".join(json_lines)

    # 2. Try parsing the whole thing
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 3. Find the first top-level { ... } or [ ... ] via brace matching
    for opener, closer in [("{", "}"), ("[", "]")]:
        start = text.find(opener)
        if start == -1:
            continue
        depth = 0
        in_str = False
        escape = False
        for i in range(start, len(text)):
            ch = text[i]
            if escape:
                escape = False
                continue
            if ch == "\\":
                escape = True
                continue
            if ch == '"':
                in_str = not in_str
                continue
            if in_str:
                continue
            if ch == opener:
                depth += 1
            elif ch == closer:
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[start:i + 1])
                    except json.JSONDecodeError:
                        break

    raise json.JSONDecodeError("No valid JSON found in response", text, 0)


def _unwrap_comps_response(data) -> Optional[list]:
    """Unwrap the Chain-of-Verify response format.

    Handles both:
    - CoVe format: {"comparables": [...]}
    - Legacy bare array: [...]

    Returns the list of comparable items, or None if invalid.
    """
    if isinstance(data, list):
        # Legacy format — bare array
        return data
    if isinstance(data, dict):
        comps = data.get("comparables")
        if isinstance(comps, list):
            return comps
    return None


def _clean_str(value) -> str:
    """Sanitise a string value from Claude's JSON — convert None/null to empty string."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("null", "none", "n/a"):
        return ""
    return s


def _to_float(value) -> Optional[float]:
    """Convert to float, handling strings with commas/symbols."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace("£", "").replace(",", "").replace(" ", "").strip()
        if not cleaned or cleaned.lower() in ("null", "none", "n/a", "tbc", "poa"):
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _to_int(value) -> Optional[int]:
    """Convert to int, handling strings."""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned or cleaned.lower() in ("null", "none", "n/a"):
            return None
        try:
            return int(float(cleaned))
        except ValueError:
            return None
    return None
