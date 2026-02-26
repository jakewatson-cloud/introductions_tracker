"""
Raw Occupational Comparables — Database-Primary Store
=====================================================
Handles insert, dedup, merge, export, and backup for the
raw_occupational_comps table.  Replaces the Excel-primary
OccupationalCompsWriter for data storage (Excel is kept as
a convenience export only).
"""

import difflib
import logging
import re
import shutil
import sqlite3
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from email_pipeline.models import OccupationalComp
from email_pipeline.occ_comps_columns import RAW_HEADERS

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Name normalisation (ported from excel_writer._normalize_name and friends)
# ---------------------------------------------------------------------------

def _normalize_name(name: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace."""
    name = name.lower().strip()
    name = re.sub(r"[^\w\s]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def normalise_tenant(name: str) -> str:
    """Normalise a tenant name for dedup comparison."""
    n = _normalize_name(name)
    n = re.sub(r'\b(ltd|limited|plc|inc|llp|llc)\b', '', n)
    return re.sub(r'\s+', ' ', n).strip()


def normalise_unit(name: str) -> str:
    """Normalise a unit name for dedup comparison."""
    n = _normalize_name(name)
    n = re.sub(r'\b0+(\d)', r'\1', n)
    n = re.sub(r'\bunit\b\s*', '', n)
    n = re.sub(r'\bplot\b\s*', '', n)
    return n.strip()


def _is_rent_close(rent_a: Optional[float], rent_b: Optional[float],
                   tolerance: float = 0.005) -> bool:
    """Check if two rents are within +/-tolerance of each other."""
    if not rent_a or not rent_b:
        return False
    avg = (rent_a + rent_b) / 2
    if avg == 0:
        return False
    return abs(rent_a - rent_b) / avg <= tolerance


# ---------------------------------------------------------------------------
# DB column names for the raw table (order matches RAW_HEADERS)
# ---------------------------------------------------------------------------

_RAW_DB_COLS = [
    "source_deal", "entry_type", "tenant_name", "unit_name",
    "address", "town", "postcode",
    "size_sqft", "rent_pa", "rent_psf",
    "lease_start", "lease_expiry", "break_date",
    "rent_review_date", "lease_term_years", "comp_date",
    "notes", "source_file_path", "extraction_date",
]

# Columns eligible for merge (fill-blank-only on duplicate)
_MERGE_FIELDS = [
    "tenant_name", "unit_name", "address", "town", "postcode",
    "size_sqft", "rent_pa", "rent_psf",
    "lease_start", "lease_expiry", "break_date",
    "rent_review_date", "lease_term_years", "comp_date", "notes",
]


class RawOccCompsDB:
    """Database-primary store for raw occupational comparables."""

    def __init__(self, db_path: Path):
        self.db_path = Path(db_path)

    # ------------------------------------------------------------------
    # Connection helper
    # ------------------------------------------------------------------

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        return conn

    # ------------------------------------------------------------------
    # Insert / dedup / merge
    # ------------------------------------------------------------------

    def insert_comp(self, comp: OccupationalComp) -> tuple[str, int]:
        """Insert a comp, deduplicating against existing rows.

        Returns ('inserted', id), ('merged', id), or ('duplicate', id).
        """
        now = datetime.now().isoformat()
        tenant_norm = normalise_tenant(comp.tenant_name or "")
        unit_norm = normalise_unit(comp.unit_name or "")

        with self._connect() as conn:
            dup_id = self._find_duplicate(conn, comp, tenant_norm, unit_norm)

            if dup_id is not None:
                fills = self._merge_into_row(conn, dup_id, comp, now)
                if fills > 0:
                    self._log_change(conn, "merge", dup_id,
                                     comp.source_deal, comp.tenant_name,
                                     context=f"{fills} fields filled")
                    conn.commit()
                    return ("merged", dup_id)
                else:
                    self._log_change(conn, "skip_duplicate", dup_id,
                                     comp.source_deal, comp.tenant_name)
                    conn.commit()
                    return ("duplicate", dup_id)

            # New row
            row_id = self._insert_new(conn, comp, tenant_norm, unit_norm, now)
            self._log_change(conn, "insert", row_id,
                             comp.source_deal, comp.tenant_name)
            conn.commit()
            return ("inserted", row_id)

    def _insert_new(self, conn: sqlite3.Connection, comp: OccupationalComp,
                    tenant_norm: str, unit_norm: str, now: str) -> int:
        """Insert a brand-new row. Returns the new row ID."""
        cur = conn.execute(
            """
            INSERT INTO raw_occupational_comps (
                source_deal, entry_type, tenant_name, tenant_name_norm,
                unit_name, unit_name_norm,
                address, town, postcode,
                size_sqft, rent_pa, rent_psf,
                lease_start, lease_expiry, break_date,
                rent_review_date, lease_term_years, comp_date,
                notes, source_file_path, extraction_date,
                created_at, updated_at
            ) VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )
            """,
            (
                comp.source_deal,
                comp.entry_type or "tenancy",
                comp.tenant_name or "",
                tenant_norm,
                comp.unit_name or "",
                unit_norm,
                comp.address or "",
                comp.town or "",
                comp.postcode or "",
                comp.size_sqft,
                comp.rent_pa,
                comp.rent_psf,
                comp.lease_start or "",
                comp.lease_expiry or "",
                comp.break_date or "",
                comp.rent_review_date or "",
                comp.lease_term_years,
                comp.comp_date or "",
                comp.notes or "",
                comp.source_file_path or "",
                datetime.now().strftime("%Y-%m-%d"),
                now,
                now,
            ),
        )
        return cur.lastrowid

    def _find_duplicate(self, conn: sqlite3.Connection,
                        comp: OccupationalComp,
                        tenant_norm: str, unit_norm: str) -> Optional[int]:
        """Three-phase dedup matching the Excel writer logic.

        Phase 1: normalised tenant + rent PA within +/-0.5%
        Phase 2: exact normalised unit + rent PA within +/-0.5%
        Phase 3: fuzzy tenant (>=90%) + rent PA within +/-0.5%
        """
        # Skip vacant entries
        if tenant_norm in ("vacant", "vacant under offer"):
            return None

        comp_rent = comp.rent_pa
        comp_rent_psf = comp.rent_psf

        def _rents_match(existing_rent_pa, existing_rent_psf):
            if _is_rent_close(comp_rent, existing_rent_pa):
                return True
            if comp_rent is None and existing_rent_pa is None:
                return _is_rent_close(comp_rent_psf, existing_rent_psf)
            return False

        # Phase 1: tenant name + rent
        if tenant_norm:
            rows = conn.execute(
                """SELECT id, rent_pa, rent_psf, tenant_name_norm
                   FROM raw_occupational_comps
                   WHERE tenant_name_norm = ?
                     AND tenant_name_norm NOT IN ('vacant', 'vacant under offer')
                """,
                (tenant_norm,),
            ).fetchall()
            for row in rows:
                if _rents_match(row["rent_pa"], row["rent_psf"]):
                    return row["id"]

        # Phase 2: unit + rent
        if unit_norm:
            rows = conn.execute(
                """SELECT id, rent_pa, rent_psf
                   FROM raw_occupational_comps
                   WHERE unit_name_norm = ?
                     AND unit_name_norm != ''
                """,
                (unit_norm,),
            ).fetchall()
            for row in rows:
                if _rents_match(row["rent_pa"], row["rent_psf"]):
                    return row["id"]

        # Phase 3: fuzzy tenant + rent
        if tenant_norm:
            candidates = conn.execute(
                """SELECT id, tenant_name_norm, rent_pa, rent_psf
                   FROM raw_occupational_comps
                   WHERE tenant_name_norm != ''
                     AND tenant_name_norm NOT IN ('vacant', 'vacant under offer')
                """,
            ).fetchall()
            for row in candidates:
                if not _rents_match(row["rent_pa"], row["rent_psf"]):
                    continue
                ratio = difflib.SequenceMatcher(
                    None, tenant_norm, row["tenant_name_norm"]
                ).ratio()
                if ratio >= 0.90:
                    logger.info(
                        "  Fuzzy tenant match (%.0f%%): '%s' ~ '%s' (id %d)",
                        ratio * 100, tenant_norm, row["tenant_name_norm"],
                        row["id"],
                    )
                    return row["id"]

        return None

    def _merge_into_row(self, conn: sqlite3.Connection, row_id: int,
                        comp: OccupationalComp, now: str) -> int:
        """Fill blank fields from a duplicate comp. Never overwrites. Returns fill count."""
        existing = conn.execute(
            "SELECT * FROM raw_occupational_comps WHERE id = ?", (row_id,)
        ).fetchone()
        if not existing:
            return 0

        field_to_comp = {
            "tenant_name": comp.tenant_name,
            "unit_name": comp.unit_name,
            "address": comp.address,
            "town": comp.town,
            "postcode": comp.postcode,
            "size_sqft": comp.size_sqft,
            "rent_pa": comp.rent_pa,
            "rent_psf": comp.rent_psf,
            "lease_start": comp.lease_start,
            "lease_expiry": comp.lease_expiry,
            "break_date": comp.break_date,
            "rent_review_date": comp.rent_review_date,
            "lease_term_years": comp.lease_term_years,
            "comp_date": comp.comp_date,
            "notes": comp.notes,
        }

        updates = {}
        for field, new_val in field_to_comp.items():
            if new_val is None:
                continue
            if isinstance(new_val, str) and not new_val.strip():
                continue
            old_val = existing[field]
            if old_val is not None and str(old_val).strip() != "":
                continue
            updates[field] = new_val

        if not updates:
            return 0

        # Also refresh normalised fields if tenant/unit changed
        if "tenant_name" in updates:
            updates["tenant_name_norm"] = normalise_tenant(str(updates["tenant_name"]))
        if "unit_name" in updates:
            updates["unit_name_norm"] = normalise_unit(str(updates["unit_name"]))
        updates["updated_at"] = now

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        conn.execute(
            f"UPDATE raw_occupational_comps SET {set_clause} WHERE id = ?",
            (*updates.values(), row_id),
        )

        fills = len(updates) - 1  # don't count updated_at
        if "tenant_name_norm" in updates:
            fills -= 1
        if "unit_name_norm" in updates:
            fills -= 1

        for field, new_val in updates.items():
            if field in ("updated_at", "tenant_name_norm", "unit_name_norm"):
                continue
            self._log_change(
                conn, "merge", row_id,
                existing["source_deal"], existing["tenant_name"],
                field_name=field,
                old_value=str(existing[field] or ""),
                new_value=str(new_val),
            )

        return fills

    # ------------------------------------------------------------------
    # Read
    # ------------------------------------------------------------------

    def get_all_raw(self) -> list[dict]:
        """Read all raw rows as dicts (same format as _read_raw_rows).

        Includes 'id' for DB row identification.
        """
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM raw_occupational_comps ORDER BY id"
            ).fetchall()

        result = []
        for row in rows:
            d = {
                "id": row["id"],
                "source_deal": str(row["source_deal"] or ""),
                "entry_type": str(row["entry_type"] or "tenancy"),
                "tenant_name": str(row["tenant_name"] or ""),
                "unit_name": str(row["unit_name"] or ""),
                "address": str(row["address"] or ""),
                "town": str(row["town"] or ""),
                "postcode": str(row["postcode"] or ""),
                "size_sqft": row["size_sqft"],
                "rent_pa": row["rent_pa"],
                "rent_psf": row["rent_psf"],
                "lease_start": row["lease_start"] or "",
                "lease_expiry": row["lease_expiry"] or "",
                "break_date": row["break_date"] or "",
                "rent_review_date": row["rent_review_date"] or "",
                "lease_term_years": row["lease_term_years"],
                "comp_date": row["comp_date"] or "",
                "notes": str(row["notes"] or ""),
                "source_file_path": str(row["source_file_path"] or ""),
                "extraction_date": str(row["extraction_date"] or ""),
            }
            result.append(d)
        return result

    def row_count(self) -> int:
        with self._connect() as conn:
            return conn.execute(
                "SELECT COUNT(*) FROM raw_occupational_comps"
            ).fetchone()[0]

    # ------------------------------------------------------------------
    # Location enrichment (replaces _write_back_locations)
    # ------------------------------------------------------------------

    def update_locations(self, updates: list[dict]) -> int:
        """Batch update address/town/postcode for rows identified by 'id'.

        Each dict in updates must have 'id' and optionally 'address',
        'town', 'postcode'.  Only fills blank fields.

        Returns count of rows updated.
        """
        if not updates:
            return 0

        now = datetime.now().isoformat()
        count = 0

        with self._connect() as conn:
            for upd in updates:
                row_id = upd["id"]
                existing = conn.execute(
                    "SELECT address, town, postcode, source_deal, tenant_name "
                    "FROM raw_occupational_comps WHERE id = ?",
                    (row_id,),
                ).fetchone()
                if not existing:
                    continue

                sets = {}
                for field in ("address", "town", "postcode"):
                    new_val = upd.get(field, "")
                    if not new_val or not str(new_val).strip():
                        continue
                    old_val = existing[field]
                    if old_val and str(old_val).strip():
                        continue
                    sets[field] = str(new_val).strip()

                if not sets:
                    continue

                sets["updated_at"] = now
                set_clause = ", ".join(f"{k} = ?" for k in sets)
                conn.execute(
                    f"UPDATE raw_occupational_comps SET {set_clause} WHERE id = ?",
                    (*sets.values(), row_id),
                )

                for field, new_val in sets.items():
                    if field == "updated_at":
                        continue
                    self._log_change(
                        conn, "enrich", row_id,
                        existing["source_deal"], existing["tenant_name"],
                        field_name=field,
                        old_value="",
                        new_value=new_val,
                    )
                count += 1

            conn.commit()
        return count

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def export_to_excel(self, excel_path: Path) -> int:
        """Export all raw rows to an Excel file (viewing copy).

        Uses temp-file-then-copy for OneDrive safety.
        Returns row count written.
        """
        rows = self.get_all_raw()

        wb = Workbook()
        ws = wb.active
        ws.title = "Occupational Comps"

        # Headers
        bold = Font(bold=True)
        for i, header in enumerate(RAW_HEADERS, 1):
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = bold

        # Data
        for r_idx, row in enumerate(rows, 2):
            ws.cell(row=r_idx, column=1, value=row.get("source_deal", ""))
            ws.cell(row=r_idx, column=2, value=row.get("entry_type", ""))
            ws.cell(row=r_idx, column=3, value=row.get("tenant_name", ""))
            ws.cell(row=r_idx, column=4, value=row.get("unit_name", ""))
            ws.cell(row=r_idx, column=5, value=row.get("address", ""))
            ws.cell(row=r_idx, column=6, value=row.get("town", ""))
            ws.cell(row=r_idx, column=7, value=row.get("postcode", ""))
            size = row.get("size_sqft")
            if size is not None:
                ws.cell(row=r_idx, column=8, value=size)
            rent_pa = row.get("rent_pa")
            if rent_pa is not None:
                ws.cell(row=r_idx, column=9, value=rent_pa)
            rent_psf = row.get("rent_psf")
            if rent_psf is not None:
                ws.cell(row=r_idx, column=10, value=rent_psf)
            ws.cell(row=r_idx, column=11, value=row.get("lease_start") or "")
            ws.cell(row=r_idx, column=12, value=row.get("lease_expiry") or "")
            ws.cell(row=r_idx, column=13, value=row.get("break_date") or "")
            ws.cell(row=r_idx, column=14, value=row.get("rent_review_date") or "")
            term = row.get("lease_term_years")
            if term is not None:
                ws.cell(row=r_idx, column=15, value=term)
            ws.cell(row=r_idx, column=16, value=row.get("comp_date") or "")
            ws.cell(row=r_idx, column=17, value=row.get("notes", ""))
            ws.cell(row=r_idx, column=18, value=row.get("source_file_path", ""))
            ws.cell(row=r_idx, column=19, value=row.get("extraction_date", ""))

        # Auto-fit column widths
        for col_idx in range(1, len(RAW_HEADERS) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for r in range(2, min(len(rows) + 2, 50)):
                val = ws.cell(row=r, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        # Save via temp file for OneDrive safety
        excel_path = Path(excel_path)
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(excel_path.parent))
        try:
            import os
            os.close(fd)
            wb.save(tmp)
            shutil.copy2(tmp, str(excel_path))
        finally:
            try:
                Path(tmp).unlink(missing_ok=True)
            except OSError:
                pass

        return len(rows)

    # ------------------------------------------------------------------
    # Backup
    # ------------------------------------------------------------------

    def backup_db(self) -> Optional[Path]:
        """Pre-write backup: copy the SQLite file to data/backups/.

        Returns the backup path, or None if the DB file doesn't exist.
        """
        if not self.db_path.exists():
            return None

        backup_dir = self.db_path.parent / "backups"
        backup_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = backup_dir / f"introductions_tracker_{timestamp}.db"
        shutil.copy2(str(self.db_path), str(dest))
        logger.info("DB backup: %s", dest)

        # Prune old backups (keep 5 most recent)
        backups = sorted(
            backup_dir.glob("introductions_tracker_*.db"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        for old in backups[5:]:
            try:
                old.unlink()
                logger.debug("Pruned old DB backup: %s", old.name)
            except OSError:
                pass

        return dest

    # ------------------------------------------------------------------
    # Full dedup pass (replaces find_occ_dupes.dedup_occupational_comps)
    # ------------------------------------------------------------------

    def run_full_dedup(self, fix: bool = False) -> dict:
        """Six-phase dedup scan over all raw rows in the DB.

        Phases 1-3: duplicate detection (tenant/unit/fuzzy + rent match)
        Phase 4: vacant rows
        Phase 5: investment comp rows (NIY, yield, cap val in notes)
        Phase 6: rows with no rent (PA or PSF)

        When fix=True, merges data from dupe rows into keep rows and
        deletes flagged rows from the DB.

        Returns summary dict.
        """
        import re as _re
        from difflib import SequenceMatcher as _SM

        with self._connect() as conn:
            all_rows = conn.execute(
                "SELECT * FROM raw_occupational_comps ORDER BY id"
            ).fetchall()

        rows = []
        for r in all_rows:
            tenant_raw = str(r["tenant_name"] or "").strip()
            unit_raw = str(r["unit_name"] or "").strip()
            notes_raw = str(r["notes"] or "").strip()
            rows.append({
                "id": r["id"],
                "source_deal": str(r["source_deal"] or "").strip(),
                "entry_type": str(r["entry_type"] or "").strip(),
                "tenant_raw": tenant_raw,
                "tenant_norm": r["tenant_name_norm"] or normalise_tenant(tenant_raw),
                "unit_raw": unit_raw,
                "unit_norm": r["unit_name_norm"] or normalise_unit(unit_raw),
                "address": str(r["address"] or "").strip(),
                "town": str(r["town"] or "").strip(),
                "notes_raw": notes_raw,
                "rent_pa": r["rent_pa"],
                "rent_psf": r["rent_psf"],
                "db_row": r,
            })

        print(f"Loaded {len(rows)} data rows from DB")

        def _is_vacant(tenant):
            if not tenant or not tenant.strip():
                return False
            return bool(_re.search(r'(?i)\bvacant\b', tenant))

        def _is_inv_comp(notes):
            return bool(
                _re.search(r'(?i)investment\s+comp', notes)
                or _re.search(r'(?i)\bNIY\b', notes)
                or _re.search(r'(?i)\byield\b', notes)
                or _re.search(r'(?i)\bcap\s*val', notes)
            )

        def _rents_match_rows(a, b, tol=0.005):
            if _is_rent_close(a["rent_pa"], b["rent_pa"], tol):
                return True
            if a["rent_pa"] is None and b["rent_pa"] is None:
                return _is_rent_close(a["rent_psf"], b["rent_psf"], tol)
            return False

        # Flag rows for removal
        vacant_ids = set()
        inv_comp_ids = set()
        no_rent_ids = set()
        for r in rows:
            if _is_vacant(r["tenant_raw"]):
                vacant_ids.add(r["id"])
            if not r["tenant_raw"] and _re.search(r'(?i)\bvacant\b', r["notes_raw"]):
                vacant_ids.add(r["id"])
            if _is_inv_comp(r["notes_raw"]):
                inv_comp_ids.add(r["id"])
            if r["rent_pa"] is None and r["rent_psf"] is None:
                no_rent_ids.add(r["id"])

        # Find duplicate pairs (Phases 1-3)
        skip_ids = vacant_ids | inv_comp_ids | no_rent_ids
        dupes_found = []  # (keep, dupe, reason)
        seen = set()

        for i, a in enumerate(rows):
            if a["id"] in seen or a["id"] in skip_ids:
                continue
            for j, b in enumerate(rows):
                if j <= i:
                    continue
                if b["id"] in seen or b["id"] in skip_ids:
                    continue
                # Phase 1: tenant + rent
                if (a["tenant_norm"] and b["tenant_norm"]
                        and a["tenant_norm"] == b["tenant_norm"]
                        and _rents_match_rows(a, b)):
                    dupes_found.append((a, b, "tenant+rent"))
                    seen.add(b["id"])
                    continue
                # Phase 2: unit + rent
                if (a["unit_norm"] and b["unit_norm"]
                        and a["unit_norm"] == b["unit_norm"]
                        and _rents_match_rows(a, b)):
                    dupes_found.append((a, b, "unit+rent"))
                    seen.add(b["id"])
                    continue
                # Phase 3: fuzzy tenant + rent
                if (a["tenant_norm"] and b["tenant_norm"]
                        and _rents_match_rows(a, b)):
                    ratio = _SM(None, a["tenant_norm"], b["tenant_norm"]).ratio()
                    if ratio >= 0.90:
                        dupes_found.append((a, b, f"fuzzy({ratio:.0%})+rent"))
                        seen.add(b["id"])
                        continue

        all_ids_to_remove = seen | vacant_ids | inv_comp_ids | no_rent_ids
        details = []

        if dupes_found:
            print(f"Found {len(dupes_found)} duplicate pair(s)")
            for keep, dupe, reason in dupes_found:
                details.append(f"[{reason}] '{keep['tenant_raw'][:30]}' id {keep['id']} "
                               f"<-> '{dupe['tenant_raw'][:30]}' id {dupe['id']}")
        if vacant_ids:
            print(f"Found {len(vacant_ids)} vacant row(s)")
        if inv_comp_ids:
            print(f"Found {len(inv_comp_ids)} investment comp row(s)")
        if no_rent_ids:
            print(f"Found {len(no_rent_ids)} no-rent row(s)")

        print(f"Total to remove: {len(all_ids_to_remove)}")

        if not fix:
            return {
                "rows_scanned": len(rows),
                "duplicate_pairs": len(dupes_found),
                "vacant_rows": len(vacant_ids),
                "inv_comp_rows": len(inv_comp_ids),
                "no_rent_rows": len(no_rent_ids),
                "rows_removed": 0,
                "details": details,
            }

        # Apply: merge dupe data into keep rows, then delete
        with self._connect() as conn:
            # Merge
            _MERGE_DB_FIELDS = [
                "tenant_name", "unit_name", "address", "town", "postcode",
                "size_sqft", "rent_pa", "rent_psf",
                "lease_start", "lease_expiry", "break_date",
                "rent_review_date", "lease_term_years", "comp_date", "notes",
            ]
            for keep, dupe, reason in dupes_found:
                keep_row = keep["db_row"]
                dupe_row = dupe["db_row"]
                updates = {}
                for field in _MERGE_DB_FIELDS:
                    keep_val = keep_row[field]
                    dupe_val = dupe_row[field]
                    keep_empty = keep_val is None or str(keep_val).strip() == ""
                    dupe_has = dupe_val is not None and str(dupe_val).strip() != ""
                    if keep_empty and dupe_has:
                        updates[field] = dupe_val
                if updates:
                    updates["updated_at"] = datetime.now().isoformat()
                    set_clause = ", ".join(f"{k} = ?" for k in updates)
                    conn.execute(
                        f"UPDATE raw_occupational_comps SET {set_clause} WHERE id = ?",
                        (*updates.values(), keep["id"]),
                    )

            # Delete
            if all_ids_to_remove:
                placeholders = ",".join("?" * len(all_ids_to_remove))
                conn.execute(
                    f"DELETE FROM raw_occupational_comps WHERE id IN ({placeholders})",
                    tuple(all_ids_to_remove),
                )
                for rid in all_ids_to_remove:
                    self._log_change(conn, "dedup_remove", rid, context="full dedup pass")

            conn.commit()

        print(f"Removed {len(all_ids_to_remove)} row(s) from DB")

        return {
            "rows_scanned": len(rows),
            "duplicate_pairs": len(dupes_found),
            "vacant_rows": len(vacant_ids),
            "inv_comp_rows": len(inv_comp_ids),
            "no_rent_rows": len(no_rent_ids),
            "rows_removed": len(all_ids_to_remove),
            "details": details,
        }

    # ------------------------------------------------------------------
    # Change log
    # ------------------------------------------------------------------

    def _log_change(self, conn: sqlite3.Connection, action: str,
                    raw_comp_id: Optional[int] = None,
                    source_deal: Optional[str] = None,
                    tenant_name: Optional[str] = None,
                    field_name: Optional[str] = None,
                    old_value: Optional[str] = None,
                    new_value: Optional[str] = None,
                    context: Optional[str] = None):
        conn.execute(
            """INSERT INTO occ_comps_change_log
               (timestamp, action, raw_comp_id, source_deal, tenant_name,
                field_name, old_value, new_value, context)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                datetime.now().isoformat(),
                action,
                raw_comp_id,
                source_deal,
                tenant_name,
                field_name,
                old_value,
                new_value,
                context,
            ),
        )
